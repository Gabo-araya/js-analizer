#!/usr/bin/env python3
from functools import wraps
from flask import Flask, render_template, jsonify, request, redirect, url_for, flash, send_file, make_response, session, Response
import sqlite3
import json
import os
import requests
import re
import time
import csv
import io
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup
from datetime import datetime
import pytz
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from flask_wtf.csrf import CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
from security_config import rate_limit, log_security_event

# Import Fase 2 enhanced detection systems
try:
    from library_signatures import detect_libraries_by_content, get_library_info
    CONTENT_DETECTION_AVAILABLE = True
    print("🎯 Dashboard: Content-based library detection enabled")
except ImportError:
    print("⚠️ Dashboard: Content-based detection not available")
    CONTENT_DETECTION_AVAILABLE = False


try:
    from cdn_analyzer import analyze_cdn_url, get_cdn_recommendations, cdn_analyzer
    CDN_ANALYZER_AVAILABLE = True
    supported_cdns = len(cdn_analyzer.get_supported_cdns())
    print(f"🌐 Dashboard: CDN analyzer enabled ({supported_cdns} CDNs)")
except ImportError:
    print("⚠️ Dashboard: CDN analyzer not available")
    CDN_ANALYZER_AVAILABLE = False

app = Flask(__name__)
csrf = CSRFProtect(app)

# Configuración de timezone para Chile
CHILE_TZ = pytz.timezone('America/Santiago')

def get_chile_time():
    """Obtiene la fecha y hora actual en timezone de Chile."""
    return datetime.now(CHILE_TZ)

def format_chile_time(dt_obj=None, fmt='%d-%m-%Y %H:%M'):
    """Formatea fecha/hora en timezone de Chile."""
    if dt_obj is None:
        dt_obj = get_chile_time()
    elif dt_obj.tzinfo is None:
        # Si el datetime no tiene timezone, asumimos UTC y convertimos a Chile
        dt_obj = pytz.UTC.localize(dt_obj).astimezone(CHILE_TZ)
    elif dt_obj.tzinfo != CHILE_TZ:
        # Si tiene otro timezone, convertir a Chile
        dt_obj = dt_obj.astimezone(CHILE_TZ)
    return dt_obj.strftime(fmt)

# Secure secret key configuration
import os
import secrets
app.secret_key = os.environ.get('FLASK_SECRET_KEY') or secrets.token_hex(32)

# Configuración de logging para depuración
# TEMPORALMENTE DESHABILITADO para resolver conflictos de concurrencia
ENABLE_ACTION_LOGGING = os.environ.get('ENABLE_ACTION_LOGGING', 'false').lower() == 'true'
LOGGING_DEBUG = os.environ.get('LOGGING_DEBUG', 'false').lower() == 'true'

print(f"🔧 Action Logging: {'ENABLED' if ENABLE_ACTION_LOGGING else 'DISABLED'}")
print(f"🔍 Debug Logging: {'ENABLED' if LOGGING_DEBUG else 'DISABLED'}")

# Add custom Jinja2 filters
@app.template_filter('tojsonfilter')
def to_json_filter(obj):
    """Convert object to JSON string for use in JavaScript - SAFE VERSION"""
    # Since all data is now pre-converted, this should just serialize
    import json
    try:
        return json.dumps(obj)
    except Exception as e:
        # Fallback: convert any remaining Row objects
        def convert_any_rows(item):
            if hasattr(item, 'keys') and not isinstance(item, dict):
                return dict(item)
            elif isinstance(item, (list, tuple)):
                return [convert_any_rows(sub_item) for sub_item in item]
            elif isinstance(item, dict):
                return {k: convert_any_rows(v) for k, v in item.items()}
            return item

        converted = convert_any_rows(obj)
        return json.dumps(converted)

@app.template_filter('has_vulnerability')
def has_vulnerability_filter(current_version, safe_version):
    """Template filter to check if a library version has vulnerabilities"""
    return has_vulnerability(current_version, safe_version)

@app.template_global()
def check_vulnerability_with_global(current_version, individual_safe, global_safe):
    """
    Template global function to check vulnerability with global fallback
    Usage in template: {{ check_vulnerability_with_global(lib.version, lib.latest_safe_version, lib.gl_latest_safe_version) }}
    """
    return has_vulnerability(current_version, individual_safe, global_safe_version=global_safe)

@app.template_global()
def get_effective_safe_version(individual_safe, global_safe):
    """
    Get the effective safe version (individual first, then global)
    Usage in template: {{ get_effective_safe_version(lib.latest_safe_version, lib.gl_latest_safe_version) }}
    """
    return individual_safe or global_safe

@app.template_filter('truncate_left')
def truncate_left(value, length=40):
    """Truncate from the left, showing the end of the string"""
    if len(value) <= length:
        return value
    return '...' + value[-(length-3):]

@app.template_filter('check_vulnerability_with_global')
def check_vulnerability_with_global_filter(current_version, individual_safe, global_safe):
    """Template filter to check vulnerability with global library data"""
    return check_vulnerability_with_global(current_version, individual_safe, global_safe)

@app.template_filter('get_effective_safe_version')
def get_effective_safe_version_filter(individual_safe, global_safe):
    """Template filter to get effective safe version"""
    return get_effective_safe_version(individual_safe, global_safe)

def analyze_security_headers(headers):
    """
    Analyzes HTTP headers for security best practices (Enhanced 2024)
    Returns dict with present headers, missing headers, warnings, and weighted security score
    """
    # Headers clasificados por prioridad de seguridad
    security_headers = {
        # CRÍTICOS (40% del score)
        'strict-transport-security': {
            'name': 'Strict-Transport-Security',
            'description': 'Enforces HTTPS connections',
            'recommendation': 'max-age=31536000; includeSubDomains; preload',
            'priority': 'critical',
            'weight': 20
        },
        'content-security-policy': {
            'name': 'Content-Security-Policy',
            'description': 'Controls resource loading and prevents XSS',
            'recommendation': "default-src 'self'; script-src 'self'; style-src 'self' 'unsafe-inline'",
            'priority': 'critical',
            'weight': 20
        },
        
        # ALTOS (30% del score)
        'x-frame-options': {
            'name': 'X-Frame-Options',
            'description': 'Prevents clickjacking attacks',
            'recommendation': 'DENY or SAMEORIGIN',
            'priority': 'high',
            'weight': 10
        },
        'x-content-type-options': {
            'name': 'X-Content-Type-Options',
            'description': 'Prevents MIME type sniffing',
            'recommendation': 'nosniff',
            'priority': 'high',
            'weight': 10
        },
        'cross-origin-embedder-policy': {
            'name': 'Cross-Origin-Embedder-Policy',
            'description': 'Controls cross-origin resource embedding',
            'recommendation': 'require-corp',
            'priority': 'high',
            'weight': 5
        },
        'cross-origin-opener-policy': {
            'name': 'Cross-Origin-Opener-Policy',
            'description': 'Isolates browsing context from cross-origin windows',
            'recommendation': 'same-origin',
            'priority': 'high',
            'weight': 5
        },
        
        # MEDIOS (20% del score)
        'referrer-policy': {
            'name': 'Referrer-Policy',
            'description': 'Controls referrer information leakage',
            'recommendation': 'strict-origin-when-cross-origin',
            'priority': 'medium',
            'weight': 7
        },
        'permissions-policy': {
            'name': 'Permissions-Policy',
            'description': 'Controls browser API access',
            'recommendation': 'geolocation=(), microphone=(), camera=()',
            'priority': 'medium',
            'weight': 7
        },
        'cross-origin-resource-policy': {
            'name': 'Cross-Origin-Resource-Policy',
            'description': 'Controls cross-origin resource sharing',
            'recommendation': 'cross-origin',
            'priority': 'medium',
            'weight': 6
        },
        
        # BAJOS (10% del score)
        'x-xss-protection': {
            'name': 'X-XSS-Protection',
            'description': 'Legacy XSS filtering (deprecated, use CSP)',
            'recommendation': '0 (deprecated, use CSP)',
            'priority': 'low',
            'weight': 3
        },
        'expect-ct': {
            'name': 'Expect-CT',
            'description': 'Certificate Transparency monitoring',
            'recommendation': 'max-age=86400, enforce',
            'priority': 'low',
            'weight': 4
        },
        'origin-agent-cluster': {
            'name': 'Origin-Agent-Cluster',
            'description': 'Requests origin-keyed agent clustering',
            'recommendation': '?1',
            'priority': 'low',
            'weight': 3
        }
    }

    # Convert headers to lowercase for case-insensitive comparison
    headers_lower = {k.lower(): v for k, v in headers.items()}

    present_headers = []
    missing_headers = []
    warnings = []
    total_weight = sum(header_info['weight'] for header_info in security_headers.values())
    achieved_weight = 0

    for header_key, header_info in security_headers.items():
        if header_key in headers_lower:
            header_value = headers_lower[header_key]
            present_headers.append({
                'name': header_info['name'],
                'value': header_value,
                'description': header_info['description'],
                'priority': header_info['priority']
            })
            achieved_weight += header_info['weight']

            # Validación inteligente por header específico
            header_warnings = _validate_header_value(header_key, header_value, header_info)
            warnings.extend(header_warnings)
            
        else:
            missing_headers.append({
                'name': header_info['name'],
                'description': header_info['description'],
                'recommendation': header_info['recommendation'],
                'priority': header_info['priority']
            })

    # Calcular puntuación ponderada por prioridad
    weighted_score = round((achieved_weight / total_weight) * 100) if total_weight > 0 else 0

    return {
        'present': present_headers,
        'missing': missing_headers,
        'warnings': warnings,
        'security_score': weighted_score,
        'priority_breakdown': _calculate_priority_breakdown(present_headers, security_headers)
    }


def _validate_header_value(header_key, header_value, header_info):
    """
    Validación inteligente de valores de headers de seguridad
    """
    warnings = []
    
    if header_key == 'content-security-policy':
        # Validación CSP inteligente
        csp_warnings = _validate_csp_policy(header_value)
        warnings.extend(csp_warnings)
        
    elif header_key == 'strict-transport-security':
        # Validar HSTS
        if 'max-age' not in header_value.lower():
            warnings.append("HSTS header missing max-age directive")
        elif 'max-age=0' in header_value.lower():
            warnings.append("HSTS disabled with max-age=0")
        elif 'includesubdomains' not in header_value.lower():
            warnings.append("HSTS missing includeSubDomains for better security")
            
    elif header_key == 'x-xss-protection':
        # X-XSS-Protection es legacy y puede causar vulnerabilidades
        if header_value != '0':
            warnings.append("X-XSS-Protection is deprecated and may introduce vulnerabilities. Set to '0' and use CSP instead")
            
    elif header_key == 'x-frame-options':
        if header_value.upper() == 'ALLOWALL':
            warnings.append("X-Frame-Options set to ALLOWALL - vulnerable to clickjacking")
        elif header_value.upper() not in ['DENY', 'SAMEORIGIN']:
            warnings.append(f"X-Frame-Options value '{header_value}' may not be secure")
            
    elif header_key == 'referrer-policy':
        unsafe_policies = ['unsafe-url', 'no-referrer-when-downgrade', 'origin-when-cross-origin']
        if header_value.lower() in unsafe_policies:
            warnings.append(f"Referrer-Policy '{header_value}' may leak sensitive information")
    
    return warnings


def _validate_csp_policy(csp_value):
    """
    Validación inteligente de Content Security Policy
    """
    warnings = []
    csp_lower = csp_value.lower()
    
    # Detectar configuraciones peligrosas
    if "'unsafe-inline'" in csp_lower:
        warnings.append("CSP allows 'unsafe-inline' - vulnerable to XSS attacks")
    
    if "'unsafe-eval'" in csp_lower:
        warnings.append("CSP allows 'unsafe-eval' - vulnerable to code injection")
        
    if "'unsafe-hashes'" in csp_lower:
        warnings.append("CSP allows 'unsafe-hashes' - potential security risk")
    
    if "* " in csp_value or csp_value.startswith("*") or " *" in csp_value:
        warnings.append("CSP uses wildcard (*) - allows any domain")
        
    # Detectar protocolos inseguros
    if "http://" in csp_lower and "https://" not in csp_lower:
        warnings.append("CSP allows HTTP resources - should prefer HTTPS")
        
    if "data:" in csp_lower:
        warnings.append("CSP allows data: URIs - potential for data exfiltration")
        
    # Verificar directivas críticas
    if "default-src" not in csp_lower and "script-src" not in csp_lower:
        warnings.append("CSP missing script-src directive - scripts may load from anywhere")
        
    if "object-src" not in csp_lower:
        warnings.append("CSP missing object-src directive - consider adding 'object-src 'none''")
    
    # Detectar CDNs conocidos (esto es informativo, no warning)
    known_cdns = ['cdnjs.cloudflare.com', 'unpkg.com', 'jsdelivr.net', 'ajax.googleapis.com']
    for cdn in known_cdns:
        if cdn in csp_lower:
            # No es warning, solo informativo - los CDNs conocidos son generalmente seguros
            pass
    
    return warnings


def _calculate_priority_breakdown(present_headers, security_headers):
    """
    Calcula breakdown por prioridad para mostrar en el dashboard
    """
    priority_stats = {
        'critical': {'present': 0, 'total': 0},
        'high': {'present': 0, 'total': 0}, 
        'medium': {'present': 0, 'total': 0},
        'low': {'present': 0, 'total': 0}
    }
    
    # Contar totales por prioridad
    for header_info in security_headers.values():
        priority = header_info['priority']
        priority_stats[priority]['total'] += 1
    
    # Contar presentes por prioridad
    for header in present_headers:
        priority = header['priority']
        priority_stats[priority]['present'] += 1
    
    return priority_stats

def init_database():
    """Initialize database tables if they don't exist"""
    conn = sqlite3.connect('analysis.db')
    cursor = conn.cursor()

    # Check if database needs migration for new library columns
    try:
        cursor.execute("SELECT description FROM libraries LIMIT 1")
    except sqlite3.OperationalError:
        # Columns don't exist, need to add them
        print("🔄 Migrating database: Adding new library management columns...")
        try:
            cursor.execute("ALTER TABLE libraries ADD COLUMN description TEXT")
            cursor.execute("ALTER TABLE libraries ADD COLUMN latest_safe_version TEXT")
            cursor.execute("ALTER TABLE libraries ADD COLUMN latest_version TEXT")
            cursor.execute("ALTER TABLE libraries ADD COLUMN is_manual INTEGER DEFAULT 0")
            print("✅ Database migration completed successfully!")
        except sqlite3.OperationalError as e:
            print(f"⚠️ Migration warning: {e}")
            # Columns might already exist, continue

    conn.commit()

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS scans (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        url TEXT NOT NULL,
        scan_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        status_code INTEGER,
        title TEXT,
        headers TEXT
    )
    ''')

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS libraries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scan_id INTEGER,
        library_name TEXT NOT NULL,
        version TEXT,
        type TEXT, -- 'js' or 'css'
        source_url TEXT,
        description TEXT, -- Manual description
        latest_safe_version TEXT, -- Latest version without vulnerabilities
        latest_version TEXT, -- Latest available version
        is_manual INTEGER DEFAULT 0, -- 1 if manually added, 0 if auto-detected
        FOREIGN KEY (scan_id) REFERENCES scans (id)
    )
    ''')

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS version_strings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scan_id INTEGER,
        file_url TEXT NOT NULL,
        file_type TEXT, -- 'js' or 'css'
        line_number INTEGER,
        line_content TEXT,
        version_keyword TEXT, -- 'version' or 'versión'
        FOREIGN KEY (scan_id) REFERENCES scans (id)
    )
    ''')

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS file_urls (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scan_id INTEGER,
        file_url TEXT NOT NULL,
        file_type TEXT, -- 'js' or 'css'
        file_size INTEGER, -- in bytes, null if unknown
        status_code INTEGER, -- HTTP response code when fetching file
        FOREIGN KEY (scan_id) REFERENCES scans (id)
    )
    ''')

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL
    )
    ''')

    # Create global libraries catalog table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS global_libraries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        library_name TEXT UNIQUE NOT NULL,
        type TEXT, -- 'js' or 'css'
        latest_safe_version TEXT,
        latest_version TEXT,
        description TEXT,
        vulnerability_info TEXT, -- Description of known vulnerabilities
        source_url TEXT,
        created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')

    # Create projects table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS projects (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        description TEXT,
        contact_email TEXT,
        contact_phone TEXT,
        website TEXT,
        created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        is_active INTEGER DEFAULT 1
    )
    ''')

    # Add project_id column to scans table if it doesn't exist
    try:
        cursor.execute("SELECT project_id FROM scans LIMIT 1")
    except sqlite3.OperationalError:
        cursor.execute("ALTER TABLE scans ADD COLUMN project_id INTEGER")
        print("✅ Added project_id column to scans table")

    # Add role column to users table if it doesn't exist
    try:
        cursor.execute("SELECT role FROM users LIMIT 1")
    except sqlite3.OperationalError:
        cursor.execute("ALTER TABLE users ADD COLUMN role TEXT DEFAULT 'admin'")
        # Migrate existing users to admin role
        cursor.execute("UPDATE users SET role = 'admin' WHERE role IS NULL OR role = ''")
        print("✅ Added role column to users table and migrated existing users to admin")

    # Add reviewed column to scans table if it doesn't exist
    try:
        cursor.execute("SELECT reviewed FROM scans LIMIT 1")
    except sqlite3.OperationalError:
        cursor.execute("ALTER TABLE scans ADD COLUMN reviewed INTEGER DEFAULT 0")
        print("✅ Added reviewed column to scans table")

    # Add global_library_id column to libraries table if it doesn't exist
    try:
        cursor.execute("SELECT global_library_id FROM libraries LIMIT 1")
    except sqlite3.OperationalError:
        cursor.execute("ALTER TABLE libraries ADD COLUMN global_library_id INTEGER REFERENCES global_libraries(id) ON DELETE SET NULL")
        print("✅ Added global_library_id column to libraries table")

    # Add action_history table for audit trail if it doesn't exist
    try:
        cursor.execute("SELECT id FROM action_history LIMIT 1")
    except sqlite3.OperationalError:
        cursor.execute('''
        CREATE TABLE action_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
            user_id INTEGER NOT NULL,
            username VARCHAR(50) NOT NULL,
            user_role VARCHAR(20) NOT NULL,
            action_type VARCHAR(50) NOT NULL,  -- CREATE, UPDATE, DELETE, LOGIN, LOGOUT, UNDO
            target_table VARCHAR(50) NOT NULL, -- scans, libraries, users, etc.
            target_id INTEGER,                 -- ID del registro afectado
            target_description TEXT,           -- Descripción legible
            data_before TEXT,                  -- Estado anterior (JSON)
            data_after TEXT,                   -- Estado posterior (JSON)
            ip_address VARCHAR(45),            -- IPv4/IPv6 del usuario
            user_agent TEXT,                   -- Navegador/proyecto
            success BOOLEAN NOT NULL DEFAULT 1, -- Si la acción fue exitosa
            error_message TEXT,                -- Mensaje de error si falló
            session_id VARCHAR(255),           -- ID de sesión
            notes TEXT,                        -- Notas adicionales
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
        ''')

        # Create indexes for performance
        cursor.execute("CREATE INDEX idx_action_history_timestamp ON action_history(timestamp)")
        cursor.execute("CREATE INDEX idx_action_history_user_id ON action_history(user_id)")
        cursor.execute("CREATE INDEX idx_action_history_action_type ON action_history(action_type)")
        cursor.execute("CREATE INDEX idx_action_history_target_table ON action_history(target_table)")
        cursor.execute("CREATE INDEX idx_action_history_target_id ON action_history(target_id)")
        cursor.execute("CREATE INDEX idx_action_history_session_id ON action_history(session_id)")

        print("✅ Created action_history table with indexes for audit trail")

    conn.commit()
    conn.close()

def compare_versions(version1, version2):
    """
    Compare two semantic version strings.
    Returns: -1 if version1 < version2, 0 if equal, 1 if version1 > version2
    """
    if not version1 or not version2:
        return 0

    try:
        # Normalize versions by removing common prefixes and suffixes
        v1_clean = re.sub(r'[^0-9\.]', '', str(version1))
        v2_clean = re.sub(r'[^0-9\.]', '', str(version2))

        # Split into parts and pad with zeros
        v1_parts = [int(x) for x in v1_clean.split('.') if x.isdigit()]
        v2_parts = [int(x) for x in v2_clean.split('.') if x.isdigit()]

        # Pad to same length
        max_len = max(len(v1_parts), len(v2_parts))
        v1_parts.extend([0] * (max_len - len(v1_parts)))
        v2_parts.extend([0] * (max_len - len(v2_parts)))

        for i in range(max_len):
            if v1_parts[i] < v2_parts[i]:
                return -1
            elif v1_parts[i] > v2_parts[i]:
                return 1

        return 0
    except:
        # Fallback to string comparison
        return -1 if str(version1) < str(version2) else (1 if str(version1) > str(version2) else 0)

def has_vulnerability(current_version, safe_version, latest_version=None, global_safe_version=None):
    """
    Determine if a library version has potential vulnerabilities
    Uses global_safe_version as fallback when safe_version is not available
    """
    if not current_version:
        return False

    # Determine which safe version to use (individual first, then global)
    effective_safe_version = safe_version or global_safe_version

    if not effective_safe_version:
        return False

    # If current version is less than safe version, it's vulnerable
    if compare_versions(current_version, effective_safe_version) < 0:
        return True

    return False

def create_default_admin():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE username = ?", ('admin',))
    user = cursor.fetchone()
    if not user:
        # Generate secure random password
        admin_password = secrets.token_urlsafe(16)
        print(f"Creating default admin user...")
        print(f"\n🔐 IMPORTANT: Save these credentials securely!")
        print(f"Username: admin")
        print(f"Password: {admin_password}\n")
        cursor.execute("INSERT INTO users (username, password) VALUES (?, ?)",
                       ('admin', generate_password_hash(admin_password)))
        conn.commit()

        # Save credentials to a secure file (should be deleted after first use)
        try:
            with open('admin_credentials.txt', 'w') as f:
                f.write(f"Username: admin\nPassword: {admin_password}\n")
            print("Credentials also saved to 'admin_credentials.txt' - DELETE this file after copying the password!")
        except Exception as e:
            print(f"Could not save credentials to file: {e}")
    conn.close()

def get_db_connection():
    """Crea una conexión ultra-optimizada a la base de datos con manejo robusto de concurrencia"""
    import time
    import random

    # Retry en la conexión misma si es necesario
    max_retries = 3
    for attempt in range(max_retries):
        try:
            # Agregar pequeño delay aleatorio para distribuir conexiones
            if attempt > 0:
                time.sleep(random.uniform(0.05, 0.1))

            conn = sqlite3.connect('analysis.db', timeout=60.0)  # Timeout aumentado
            conn.row_factory = sqlite3.Row

            # Configuración ultra-optimizada para concurrencia máxima
            conn.execute('PRAGMA journal_mode=WAL')           # Write-Ahead Logging
            conn.execute('PRAGMA synchronous=NORMAL')         # Balance seguridad/velocidad
            conn.execute('PRAGMA temp_store=memory')          # Temp files en memoria
            conn.execute('PRAGMA busy_timeout=60000')         # 60 segundos timeout
            conn.execute('PRAGMA cache_size=-128000')         # 128MB cache (más grande)
            conn.execute('PRAGMA wal_autocheckpoint=500')     # Checkpoints más frecuentes
            conn.execute('PRAGMA page_size=32768')            # Page size más grande
            conn.execute('PRAGMA mmap_size=268435456')        # 256MB memory mapping
            conn.execute('PRAGMA wal_timeout=30000')          # 30s timeout para WAL
            conn.execute('PRAGMA optimize')                   # Optimizar automáticamente

            return conn

        except sqlite3.OperationalError as e:
            if attempt < max_retries - 1:
                print(f"Error conectando a BD (intento {attempt + 1}): {e}")
                continue
            else:
                raise

def row_to_dict(row):
    """Convert sqlite3.Row to dictionary for compatibility"""
    return dict(row) if hasattr(row, 'keys') else row

def convert_rows_deep(obj):
    """
    Recursively convert all sqlite3.Row objects to dictionaries
    Handles nested structures like lists, dicts, tuples
    """
    import sqlite3

    # More specific detection for sqlite3.Row objects
    if isinstance(obj, sqlite3.Row):
        # This is definitely a sqlite3.Row object
        return dict(obj)
    elif hasattr(obj, 'keys') and not isinstance(obj, dict) and hasattr(obj, '__getitem__'):
        # This might be a Row-like object, convert to dict
        try:
            return dict(obj)
        except:
            pass

    if isinstance(obj, dict):
        # Recursively convert dictionary values
        return {key: convert_rows_deep(value) for key, value in obj.items()}
    elif isinstance(obj, (list, tuple)):
        # Recursively convert list/tuple items
        converted = [convert_rows_deep(item) for item in obj]
        return converted if isinstance(obj, list) else tuple(converted)
    else:
        # Return as-is for primitive types
        return obj

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Decorator to require admin role"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login', next=request.url))

        user_role = session.get('user_role', '')
        if user_role != 'admin':
            flash('Acceso denegado: Se requieren permisos de administrador', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def get_user_role():
    """Get current user's role"""
    return session.get('user_role', 'analyst')

# ========================================
# SISTEMA DE HISTORIAL Y AUDITORÍA
# ========================================

def log_user_action(action_type, target_table, target_id=None, target_description=None,
                   data_before=None, data_after=None, success=True, error_message=None, notes=None):
    """
    Registra una acción en el historial de auditoría con manejo ultra-robusto de concurrencia

    Args:
        action_type: Tipo de acción (CREATE, UPDATE, DELETE, LOGIN, LOGOUT, UNDO)
        target_table: Tabla afectada
        target_id: ID del registro afectado
        target_description: Descripción legible de la acción
        data_before: Estado anterior del registro (dict o JSON)
        data_after: Estado posterior del registro (dict o JSON)
        success: Si la acción fue exitosa
        error_message: Mensaje de error si falló
        notes: Notas adicionales
    """
    # Verificar si el logging está habilitado
    if not ENABLE_ACTION_LOGGING:
        if LOGGING_DEBUG:
            print(f"[DEBUG] Action logging disabled, skipping: {action_type} on {target_table}")
        return

    import time
    import random
    import threading

    # Configuración más agresiva para resolver conflictos
    max_retries = 5
    base_delay = 0.2  # 200ms base más conservador

    for attempt in range(max_retries):
        conn = None
        try:
            # Agregar delay inicial aleatorio para distribuir requests
            if attempt == 0:
                time.sleep(random.uniform(0.01, 0.05))

            conn = get_db_connection()

            # Datos del usuario actual
            user_id = session.get('user_id', 0)
            username = session.get('username', 'Sistema')
            user_role = session.get('user_role', 'system')

            # Datos de la request HTTP
            ip_address = request.remote_addr if request else None
            user_agent = request.headers.get('User-Agent') if request else None
            session_id = session.get('session_id') if session else None

            # Serializar datos JSON
            data_before_json = json.dumps(data_before, default=str) if data_before else None
            data_after_json = json.dumps(data_after, default=str) if data_after else None

            # Configuración ultra-conservadora para resolver deadlocks
            conn.execute('PRAGMA busy_timeout = 20000')  # 20 segundos
            conn.execute('PRAGMA journal_mode = WAL')    # Asegurar WAL mode
            conn.execute('PRAGMA wal_timeout = 10000')   # Timeout para WAL

            # Usar transacción IMMEDIATE para reducir conflicts
            conn.execute('BEGIN IMMEDIATE')

            conn.execute('''
                INSERT INTO action_history (
                    user_id, username, user_role, action_type, target_table,
                    target_id, target_description, data_before, data_after,
                    ip_address, user_agent, success, error_message, session_id, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                user_id, username, user_role, action_type, target_table,
                target_id, target_description, data_before_json, data_after_json,
                ip_address, user_agent, success, error_message, session_id, notes
            ))

            conn.commit()
            return  # Éxito, salir del loop

        except sqlite3.OperationalError as e:
            error_msg = str(e).lower()
            if ("database is locked" in error_msg or "busy" in error_msg) and attempt < max_retries - 1:
                if conn:
                    try:
                        conn.rollback()
                    except:
                        pass

                # Backoff exponencial más agresivo con jitter
                delay = base_delay * (2 ** attempt) + random.uniform(0.1, 0.3)

                # Agregar delay extra en intentos finales
                if attempt >= 2:
                    delay += random.uniform(0.5, 1.0)

                time.sleep(delay)
                if LOGGING_DEBUG:
                    print(f"[DEBUG] Database locked, reintentando en {delay:.2f}s (intento {attempt + 1}/{max_retries}) - Thread: {threading.current_thread().ident} - Action: {action_type} on {target_table}")
                else:
                    print(f"Database locked, reintentando en {delay:.2f}s (intento {attempt + 1}/{max_retries})")
                continue
            else:
                print(f"Error de base de datos en historial (intento {attempt + 1}): {e}")
                if conn:
                    try:
                        conn.rollback()
                    except:
                        pass
                break
        except Exception as e:
            print(f"Error al registrar acción en historial (intento {attempt + 1}): {e}")
            if conn:
                try:
                    conn.rollback()
                except:
                    pass
            break
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass

def log_batch_actions(actions_list):
    """
    Registra múltiples acciones en una sola transacción para mejorar performance

    Args:
        actions_list: Lista de diccionarios con datos de acciones
    """
    if not actions_list:
        return

    import time
    import random

    max_retries = 3
    base_delay = 0.2

    for attempt in range(max_retries):
        conn = None
        try:
            conn = get_db_connection()
            conn.execute('PRAGMA busy_timeout = 15000')  # 15 segundos para lotes

            # Comenzar transacción
            conn.execute('BEGIN IMMEDIATE')

            for action_data in actions_list:
                # Serializar datos JSON
                data_before_json = json.dumps(action_data.get('data_before'), default=str) if action_data.get('data_before') else None
                data_after_json = json.dumps(action_data.get('data_after'), default=str) if action_data.get('data_after') else None

                conn.execute('''
                    INSERT INTO action_history (
                        user_id, username, user_role, action_type, target_table,
                        target_id, target_description, data_before, data_after,
                        ip_address, user_agent, success, error_message, session_id, notes
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    action_data.get('user_id', 0),
                    action_data.get('username', 'Sistema'),
                    action_data.get('user_role', 'system'),
                    action_data.get('action_type'),
                    action_data.get('target_table'),
                    action_data.get('target_id'),
                    action_data.get('target_description'),
                    data_before_json,
                    data_after_json,
                    action_data.get('ip_address'),
                    action_data.get('user_agent'),
                    action_data.get('success', True),
                    action_data.get('error_message'),
                    action_data.get('session_id'),
                    action_data.get('notes')
                ))

            conn.commit()
            print(f"✅ Registradas {len(actions_list)} acciones en lote")
            return  # Éxito

        except sqlite3.OperationalError as e:
            if "database is locked" in str(e).lower() and attempt < max_retries - 1:
                if conn:
                    try:
                        conn.rollback()
                    except:
                        pass
                delay = base_delay * (2 ** attempt) + random.uniform(0, 0.1)
                time.sleep(delay)
                print(f"Database locked en lote, reintentando en {delay:.2f}s (intento {attempt + 1}/{max_retries})")
                continue
            else:
                print(f"Error de base de datos en lote (intento {attempt + 1}): {e}")
                if conn:
                    try:
                        conn.rollback()
                    except:
                        pass
                break
        except Exception as e:
            print(f"Error en logging de lote (intento {attempt + 1}): {e}")
            if conn:
                try:
                    conn.rollback()
                except:
                    pass
            break
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass

def get_record_data(table_name, record_id):
    """
    Obtiene todos los datos de un registro específico

    Args:
        table_name: Nombre de la tabla
        record_id: ID del registro

    Returns:
        dict: Datos del registro o None si no existe
    """
    conn = get_db_connection()
    try:
        cursor = conn.execute(f'SELECT * FROM {table_name} WHERE id = ?', (record_id,))
        record = cursor.fetchone()

        if record:
            # Convertir Row a dict
            return dict(record)
        return None

    except Exception as e:
        print(f"Error al obtener datos del registro {table_name}#{record_id}: {e}")
        return None
    finally:
        conn.close()

def generate_session_id():
    """Genera un ID único de sesión"""
    import uuid
    return str(uuid.uuid4())

def get_next_available_id(table_name):
    """Obtiene el siguiente ID disponible en una tabla"""
    conn = get_db_connection()
    try:
        cursor = conn.execute(f'SELECT MAX(id) + 1 as next_id FROM {table_name}')
        result = cursor.fetchone()
        return result['next_id'] or 1
    finally:
        conn.close()

def record_exists(table_name, record_id):
    """Verifica si un registro existe en una tabla"""
    conn = get_db_connection()
    try:
        cursor = conn.execute(f'SELECT 1 FROM {table_name} WHERE id = ? LIMIT 1', (record_id,))
        return cursor.fetchone() is not None
    finally:
        conn.close()

def log_action(action_type, target_table, get_target_id=None, get_description=None):
    """
    Decorador para registrar acciones automáticamente

    Args:
        action_type: CREATE, UPDATE, DELETE, LOGIN, LOGOUT, UNDO
        target_table: tabla afectada (scans, libraries, users, etc.)
        get_target_id: función lambda para obtener ID del registro
        get_description: función lambda para obtener descripción legible

    Uso:
        @log_action('CREATE', 'scans',
                   get_target_id=lambda: request.form.get('url'),
                   get_description=lambda: f"Nuevo escaneo: {request.form.get('url')}")
        def analyze_url():
            # ... código de la función
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            target_id = None
            data_before = None

            try:
                # Capturar estado anterior para UPDATE/DELETE
                if action_type in ['UPDATE', 'DELETE'] and get_target_id:
                    target_id = get_target_id(*args, **kwargs)
                    if target_id:
                        data_before = get_record_data(target_table, target_id)

                # Ejecutar función original
                result = f(*args, **kwargs)

                # Capturar estado posterior y ID para CREATE/UPDATE
                data_after = None
                if action_type in ['CREATE', 'UPDATE']:
                    if get_target_id:
                        target_id = get_target_id(*args, **kwargs)
                    if target_id:
                        data_after = get_record_data(target_table, target_id)

                # Registrar acción exitosa
                log_user_action(
                    action_type=action_type,
                    target_table=target_table,
                    target_id=target_id,
                    target_description=get_description(*args, **kwargs) if get_description else None,
                    data_before=data_before,
                    data_after=data_after,
                    success=True
                )

                return result

            except Exception as e:
                # Registrar acción fallida
                log_user_action(
                    action_type=action_type,
                    target_table=target_table,
                    target_id=target_id,
                    target_description=get_description(*args, **kwargs) if get_description else None,
                    data_before=data_before,
                    success=False,
                    error_message=str(e)
                )
                raise  # Re-lanzar la excepción

        return decorated_function
    return decorator

def log_action_async(action_type, target_table, get_target_id=None, get_description=None):
    """
    Decorador de logging no-bloqueante para operaciones críticas
    Registra en background sin afectar la operación principal
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            try:
                # Ejecutar función original primero
                result = f(*args, **kwargs)

                # Logging en background (no bloqueante)
                try:
                    target_id = get_target_id(*args, **kwargs) if get_target_id else None
                    description = get_description(*args, **kwargs) if get_description else None

                    # Usar thread para logging no-bloqueante
                    import threading

                    def background_log():
                        try:
                            log_user_action(
                                action_type=action_type,
                                target_table=target_table,
                                target_id=target_id,
                                target_description=description,
                                success=True
                            )
                        except Exception as e:
                            print(f"⚠️ Background logging failed: {e}")

                    # Solo hacer logging si está habilitado
                    if ENABLE_ACTION_LOGGING:
                        thread = threading.Thread(target=background_log, daemon=True)
                        thread.start()

                except Exception as e:
                    print(f"⚠️ Error setting up background logging: {e}")

                return result

            except Exception as e:
                # Registrar error sin afectar el flujo principal
                try:
                    if ENABLE_ACTION_LOGGING:
                        log_user_action(
                            action_type=action_type,
                            target_table=target_table,
                            target_description=get_description(*args, **kwargs) if get_description else None,
                            success=False,
                            error_message=str(e)
                        )
                except:
                    pass  # Ignorar errores de logging

                raise  # Re-lanzar la excepción original

        return decorated_function
    return decorator

@app.route('/login', methods=['GET', 'POST'])
@rate_limit('login', 'Demasiados intentos de login. Espera un minuto antes de intentar nuevamente.')
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
        user = cursor.fetchone()
        conn.close()

        if user and check_password_hash(user['password'], password):
            # Configurar sesión
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['user_role'] = user['role'] if 'role' in user.keys() and user['role'] else 'admin'
            session['session_id'] = generate_session_id()  # Generar ID único de sesión

            # Registrar login exitoso
            log_user_action(
                action_type='LOGIN',
                target_table='users',
                target_id=user['id'],
                target_description=f"Login exitoso de usuario: {username}",
                success=True,
                notes=f"Rol: {session['user_role']}"
            )

            next_page = request.form.get('next') or request.args.get('next')

            # Validar que la URL de redirección sea segura (prevenir open redirect)
            if next_page:
                from urllib.parse import urlparse
                next_url_parsed = urlparse(next_page)

                # Permitir rutas relativas o URLs que apunten al mismo host
                if next_url_parsed.netloc == '' or next_url_parsed.netloc == request.host:
                    return redirect(next_page)

            return redirect(url_for('index'))
        else:
            # Registrar intento de login fallido
            log_user_action(
                action_type='LOGIN',
                target_table='users',
                target_description=f"Intento de login fallido para usuario: {username}",
                success=False,
                error_message='Usuario o contraseña inválidos'
            )
            flash('Usuario o contraseña inválidos', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    # Registrar logout antes de limpiar la sesión
    if 'user_id' in session:
        log_user_action(
            action_type='LOGOUT',
            target_table='users',
            target_id=session['user_id'],
            target_description=f"Logout de usuario: {session.get('username', 'desconocido')}",
            success=True
        )

    session.clear()
    flash('Has cerrado sesión exitosamente.', 'success')
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    # Get page number, project filter, and search from URL parameters
    page = request.args.get('page', 1, type=int)
    project_id_param = request.args.get('project_id', '')
    search_query = request.args.get('search', '').strip()
    per_page = 50
    offset = (page - 1) * per_page

    conn = get_db_connection()

    # Build WHERE clause for filters
    where_conditions = []
    query_params = []

    # Handle project filter
    if project_id_param:
        if project_id_param == 'null':
            # Filter for scans without project
            where_conditions.append("s.project_id IS NULL")
        else:
            # Filter for specific project
            try:
                project_id = int(project_id_param)
                where_conditions.append("s.project_id = ?")
                query_params.append(project_id)
            except ValueError:
                # Invalid project_id, ignore filter
                pass

    if search_query:
        # Search in titles, URLs, and associated libraries
        search_condition = """(
            s.title LIKE ? OR
            s.url LIKE ? OR
            s.id IN (
                SELECT DISTINCT l.scan_id
                FROM libraries l
                WHERE l.library_name LIKE ? OR l.description LIKE ?
            )
        )"""
        where_conditions.append(search_condition)
        search_param = f"%{search_query}%"
        query_params.extend([search_param, search_param, search_param, search_param])

    # Combine WHERE conditions
    where_clause = ""
    if where_conditions:
        where_clause = "WHERE " + " AND ".join(where_conditions)

    # Get summary stats
    # For the stats query, we need to duplicate parameters for each subquery that uses where_clause
    stats_query_params = query_params + query_params + query_params  # Three copies for the three subqueries

    # Build WHERE conditions for subqueries
    libraries_where = ""
    files_where = ""
    if where_conditions:
        # Replace s. with appropriate table alias and build conditions
        libs_conditions = []
        files_conditions = []
        for condition in where_conditions:
            libs_conditions.append(condition.replace('s.', 's2.'))
            files_conditions.append(condition.replace('s.', 's3.'))
        libraries_where = "AND " + " AND ".join(libs_conditions)
        files_where = "AND " + " AND ".join(files_conditions)
    
    stats_query = f'''
        SELECT
            COUNT(DISTINCT s.id) as total_scans,
            (SELECT COUNT(DISTINCT library_name) FROM libraries l JOIN scans s2 ON l.scan_id = s2.id WHERE l.type = 'js' {libraries_where}) as unique_libraries,
            (SELECT COUNT(DISTINCT fu.id) FROM file_urls fu JOIN scans s3 ON fu.scan_id = s3.id WHERE fu.file_type = 'js' {files_where}) as total_files,
            SUM(CASE WHEN s.reviewed = 1 THEN 1 ELSE 0 END) as reviewed_scans,
            SUM(CASE WHEN s.reviewed = 0 THEN 1 ELSE 0 END) as pending_scans,
            (SELECT COUNT(*) FROM projects WHERE is_active = 1) as total_projects,
            (SELECT COUNT(*) FROM global_libraries) as global_libraries_count
        FROM scans s
        {where_clause}
    '''
    stats = conn.execute(stats_query, stats_query_params).fetchone()

    # Calculate vulnerable scans using the same logic as statistics page
    all_scans_query = f'''
        SELECT s.id
        FROM scans s
        {where_clause}
    '''
    all_scans = conn.execute(all_scans_query, query_params).fetchall()

    # Count vulnerable scans using Python logic
    vulnerable_scans_count = 0
    for scan in all_scans:
        # Check if scan has any vulnerable libraries (JavaScript only)
        libs_query = '''
            SELECT library_name, version, latest_safe_version, latest_version
            FROM libraries WHERE scan_id = ? AND type = 'js'
        '''
        libraries = conn.execute(libs_query, (scan['id'],)).fetchall()
        has_vuln = False
        for lib in libraries:
            if has_vulnerability(lib['version'], lib['latest_safe_version'], lib['latest_version']):
                has_vuln = True
                break

        if has_vuln:
            vulnerable_scans_count += 1

    # Add vulnerable scans count to stats
    stats = dict(stats)
    stats['vulnerable_scans'] = vulnerable_scans_count

    # Get projects for filter dropdown
    projects = conn.execute('SELECT id, name FROM projects WHERE is_active = 1 ORDER BY name').fetchall()
    selected_project = None
    if project_id_param and project_id_param != 'null':
        try:
            project_id = int(project_id_param)
            selected_project = conn.execute('SELECT * FROM projects WHERE id = ?', (project_id,)).fetchone()
        except ValueError:
            pass

    # Get recent scans with pagination and project info
    scans_query = f'''
        SELECT s.id, s.url, s.scan_date, s.status_code, s.title, s.project_id, s.reviewed,
               c.name as project_name,
               COUNT(DISTINCT l.id) as library_count,
               COUNT(DISTINCT vs.id) as version_string_count,
               COUNT(DISTINCT fu.id) as file_count,
               COUNT(DISTINCT CASE WHEN fu.status_code IS NOT NULL AND fu.status_code != 200 THEN fu.id END) as error_count
        FROM scans s
        LEFT JOIN projects c ON s.project_id = c.id
        LEFT JOIN libraries l ON s.id = l.scan_id
        LEFT JOIN version_strings vs ON s.id = vs.scan_id
        LEFT JOIN file_urls fu ON s.id = fu.scan_id
        {where_clause}
        GROUP BY s.id
        ORDER BY scan_date DESC
        LIMIT ? OFFSET ?
    '''
    scans_params = query_params + [per_page, offset]
    recent_scans_raw = conn.execute(scans_query, scans_params).fetchall()

    # Calculate vulnerability count for each scan using Python function
    recent_scans = []
    for scan in recent_scans_raw:
        scan_dict = dict(scan)

        # Get libraries for this scan to count vulnerabilities properly
        vuln_count = conn.execute('''
            SELECT COUNT(*) as count
            FROM libraries
            WHERE scan_id = ?
            AND version IS NOT NULL
            AND latest_safe_version IS NOT NULL
            AND version != ''
            AND latest_safe_version != ''
        ''', (scan['id'],)).fetchone()['count']

        # Count vulnerabilities using Python function for accurate comparison
        if vuln_count > 0:
            libraries = conn.execute('''
                SELECT version, latest_safe_version
                FROM libraries
                WHERE scan_id = ?
                AND version IS NOT NULL
                AND latest_safe_version IS NOT NULL
                AND version != ''
                AND latest_safe_version != ''
            ''', (scan['id'],)).fetchall()

            vulnerability_count = sum(1 for lib in libraries
                                    if has_vulnerability(lib['version'], lib['latest_safe_version']))
        else:
            vulnerability_count = 0

        scan_dict['vulnerability_count'] = vulnerability_count
        recent_scans.append(scan_dict)

    # Calculate pagination info
    total_scans = stats['total_scans'] or 0
    total_pages = (total_scans + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages
    prev_num = page - 1 if has_prev else None
    next_num = page + 1 if has_next else None

    # Get top libraries with version information
    top_libraries = conn.execute('''
        WITH library_versions AS (
            SELECT
                library_name,
                type,
                version,
                COUNT(*) as version_count
            FROM libraries
            WHERE version IS NOT NULL AND version != ''
            GROUP BY library_name, type, version
        ),
        most_common_version AS (
            SELECT
                library_name,
                type,
                version as most_common_version,
                ROW_NUMBER() OVER (PARTITION BY library_name, type ORDER BY version_count DESC) as rn
            FROM library_versions
        ),
        library_totals AS (
            SELECT
                library_name,
                type,
                COUNT(*) as usage_count
            FROM libraries
            GROUP BY library_name, type
        )
        SELECT
            lt.library_name,
            lt.type,
            lt.usage_count,
            COALESCE(mcv.most_common_version, 'Unknown') as most_common_version,
            COUNT(DISTINCT l.version) as version_variety
        FROM library_totals lt
        LEFT JOIN most_common_version mcv ON lt.library_name = mcv.library_name
            AND lt.type = mcv.type AND mcv.rn = 1
        LEFT JOIN libraries l ON lt.library_name = l.library_name
            AND lt.type = l.type
        GROUP BY lt.library_name, lt.type, lt.usage_count, mcv.most_common_version
        ORDER BY lt.usage_count DESC
        LIMIT 10
    ''').fetchall()

    conn.close()

    return render_template('index.html',
                         stats=stats,
                         recent_scans=recent_scans,
                         top_libraries=top_libraries,
                         projects=projects,
                         selected_project=selected_project,
                         pagination={
                             'page': page,
                             'per_page': per_page,
                             'total': total_scans,
                             'total_pages': total_pages,
                             'has_prev': has_prev,
                             'has_next': has_next,
                             'prev_num': prev_num,
                             'next_num': next_num
                         })

@app.route('/scan/<int:scan_id>')
@login_required
def scan_detail(scan_id):
    conn = get_db_connection()

    # Get scan details with project information
    scan = conn.execute('''
        SELECT s.*, c.name as project_name
        FROM scans s
        LEFT JOIN projects c ON s.project_id = c.id
        WHERE s.id = ?
    ''', (scan_id,)).fetchone()
    if not scan:
        return "Scan not found", 404

    libraries = conn.execute('''
        SELECT
            l.id, l.library_name, l.version, l.type, l.source_url, l.description,
            l.latest_safe_version, l.latest_version, l.is_manual, l.global_library_id,
            gl.latest_safe_version as gl_latest_safe_version,
            gl.latest_version as gl_latest_version
        FROM libraries l
        LEFT JOIN global_libraries gl ON l.global_library_id = gl.id
        WHERE l.scan_id = ? AND l.type = 'js'
        ORDER BY l.type, l.library_name
    ''', (scan_id,)).fetchall()

    # Get version strings for this scan (JavaScript files only)
    version_strings_raw = conn.execute('''
        SELECT id, file_url, file_type, line_number, line_content, version_keyword
        FROM version_strings
        WHERE scan_id = ? AND file_type = 'js'
        ORDER BY file_url, line_number
    ''', (scan_id,)).fetchall()

    # Group version strings by file_url to avoid duplicates
    version_strings_grouped = {}
    for vs in version_strings_raw:
        file_url = vs['file_url']
        if file_url not in version_strings_grouped:
            version_strings_grouped[file_url] = {
                'id': vs['id'],  # Use first occurrence ID for operations
                'file_url': file_url,
                'file_type': vs['file_type'],
                'lines': [],
                'version_keywords': set(),
                'lines_count': 0,
                'all_ids': []  # Keep track of all IDs for deletion
            }

        version_strings_grouped[file_url]['lines'].append({
            'id': vs['id'],
            'line_number': vs['line_number'],
            'line_content': vs['line_content'],
            'version_keyword': vs['version_keyword']
        })
        version_strings_grouped[file_url]['version_keywords'].add(vs['version_keyword'])
        version_strings_grouped[file_url]['lines_count'] += 1
        version_strings_grouped[file_url]['all_ids'].append(vs['id'])

    # Convert to list and sort by file_url
    version_strings = sorted(version_strings_grouped.values(), key=lambda x: x['file_url'])

    # Get JavaScript file URLs for this scan (CSS excluded)
    file_urls = conn.execute('''
        SELECT id, file_url, file_type, file_size, status_code
        FROM file_urls
        WHERE scan_id = ? AND file_type = 'js'
        ORDER BY file_type, file_url
    ''', (scan_id,)).fetchall()

    # Parse headers
    headers = json.loads(scan['headers']) if scan['headers'] else {}

    # Analyze security headers
    security_analysis = analyze_security_headers(headers)

    # Get all active projects for the edit modal
    projects = conn.execute('''
        SELECT id, name
        FROM projects
        WHERE is_active = 1
        ORDER BY name
    ''').fetchall()

    # Get all global libraries for association dropdown
    global_libraries = conn.execute('SELECT id, library_name, type, latest_safe_version, latest_version FROM global_libraries ORDER BY library_name').fetchall()

    # Get navigation info for scans of the same URL
    url_scans = conn.execute('''
        SELECT id, scan_date
        FROM scans
        WHERE url = ?
        ORDER BY scan_date ASC
    ''', (scan['url'],)).fetchall()

    # Find current position and navigation
    scan_navigation = {
        'current_position': 0,
        'total_scans': len(url_scans),
        'previous_scan_id': None,
        'next_scan_id': None
    }

    for i, url_scan in enumerate(url_scans):
        if url_scan['id'] == scan_id:
            scan_navigation['current_position'] = i + 1
            if i > 0:
                scan_navigation['previous_scan_id'] = url_scans[i - 1]['id']
            if i < len(url_scans) - 1:
                scan_navigation['next_scan_id'] = url_scans[i + 1]['id']
            break

    conn.close()

    return render_template('scan_detail.html',
                         scan=scan,
                         libraries=libraries,
                         version_strings=version_strings,
                         file_urls=file_urls,
                         headers=headers,
                         security_analysis=security_analysis,
                         projects=projects,
                         global_libraries=global_libraries,
                         scan_navigation=scan_navigation)

@app.route('/api/scans')
@login_required
def api_scans():
    conn = get_db_connection()
    scans = conn.execute('''
        SELECT s.id, s.url, s.scan_date, s.status_code, s.title,
               COUNT(l.id) as library_count
        FROM scans s
        LEFT JOIN libraries l ON s.id = l.scan_id
        GROUP BY s.id
        ORDER BY scan_date DESC
    ''').fetchall()
    conn.close()

    return jsonify([dict(scan) for scan in scans])

@app.route('/api/libraries')
@login_required
def api_libraries():
    conn = get_db_connection()
    libraries = conn.execute('''
        SELECT library_name, version, type, source_url, s.url as site_url
        FROM libraries l
        JOIN scans s ON l.scan_id = s.id
        WHERE l.type = 'js'
        ORDER BY library_name, version
    ''').fetchall()
    conn.close()

    return jsonify([dict(lib) for lib in libraries])

@app.route('/api/version-strings')
@login_required
def api_version_strings():
    conn = get_db_connection()
    version_strings = conn.execute('''
        SELECT vs.file_url, vs.file_type, vs.line_number, vs.line_content,
               vs.version_keyword, s.url as site_url, s.scan_date
        FROM version_strings vs
        JOIN scans s ON vs.scan_id = s.id
        WHERE vs.file_type = 'js'
        ORDER BY s.scan_date DESC, vs.file_url, vs.line_number
    ''').fetchall()
    conn.close()

    return jsonify([dict(vs) for vs in version_strings])

@app.route('/api/stats')
@login_required
def api_stats():
    conn = get_db_connection()

    stats = {
        'overview': conn.execute('''
            SELECT
                COUNT(DISTINCT s.id) as total_scans,
                COUNT(CASE WHEN s.status_code = 200 THEN s.id END) as successful_scans,
                COUNT(DISTINCT library_name) as unique_libraries,
                COUNT(vs.id) as total_version_strings
            FROM scans s
            LEFT JOIN libraries l ON s.id = l.scan_id
            LEFT JOIN version_strings vs ON s.id = vs.scan_id
        ''').fetchone(),

        'library_types': conn.execute('''
            SELECT type, COUNT(*) as count
            FROM libraries
            GROUP BY type
        ''').fetchall(),

        'popular_libraries': conn.execute('''
            SELECT library_name, COUNT(*) as usage_count
            FROM libraries
            GROUP BY library_name
            ORDER BY usage_count DESC
            LIMIT 10
        ''').fetchall()
    }

    conn.close()

    return jsonify({
        'overview': dict(stats['overview']),
        'library_types': [dict(row) for row in stats['library_types']],
        'popular_libraries': [dict(row) for row in stats['popular_libraries']]
    })

@app.route('/statistics')
@login_required
def statistics():
    # Get pagination parameters
    page = int(request.args.get('page', 1))
    per_page = 20
    offset = (page - 1) * per_page

    # Get search parameter
    search = request.args.get('search', '').strip()

    conn = get_db_connection()

    # Build where clause for search
    where_clause = "WHERE vulnerability_count > 0"
    search_params = []

    if search:
        where_clause += " AND (s.url LIKE ? OR s.title LIKE ? OR c.name LIKE ?)"
        search_params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])

    # Get all scans for vulnerability counting using Python logic
    all_scans_query = f'''
        SELECT s.id, s.url, s.title, c.name as project_name
        FROM scans s
        LEFT JOIN projects c ON s.project_id = c.id
    '''

    if search:
        all_scans_query += " WHERE (s.url LIKE ? OR s.title LIKE ? OR c.name LIKE ?)"
        all_scans = conn.execute(all_scans_query, [f'%{search}%', f'%{search}%', f'%{search}%']).fetchall()
    else:
        all_scans = conn.execute(all_scans_query).fetchall()

    # Count vulnerable scans using Python logic
    vulnerable_scan_ids = []
    for scan in all_scans:
        # Check if scan has any vulnerable libraries
        libs_query = '''
            SELECT library_name, version, latest_safe_version, latest_version
            FROM libraries WHERE scan_id = ?
        '''
        libraries = conn.execute(libs_query, (scan['id'],)).fetchall()
        has_vuln = False
        for lib in libraries:
            if has_vulnerability(lib['version'], lib['latest_safe_version'], lib['latest_version']):
                has_vuln = True
                break

        if has_vuln:
            vulnerable_scan_ids.append(scan['id'])

    total_vulnerable_scans = len(vulnerable_scan_ids)

    # Calculate pagination
    total_pages = max(1, (total_vulnerable_scans + per_page - 1) // per_page)
    has_prev = page > 1
    has_next = page < total_pages
    prev_num = page - 1 if has_prev else None
    next_num = page + 1 if has_next else None

    # Apply pagination to filtered vulnerable scan IDs
    start_idx = offset
    end_idx = offset + per_page
    paginated_vulnerable_ids = vulnerable_scan_ids[start_idx:end_idx]

    # Get detailed scan information for paginated results
    vulnerable_scans = []
    if paginated_vulnerable_ids:
        placeholders = ','.join('?' for _ in paginated_vulnerable_ids)
        scans_query = f'''
            SELECT s.id, s.url, s.scan_date, s.status_code, s.title, s.project_id, s.reviewed,
                   c.name as project_name,
                   COUNT(DISTINCT l.id) as library_count,
                   COUNT(DISTINCT vs.id) as version_string_count,
                   COUNT(DISTINCT fu.id) as file_count,
                   COUNT(DISTINCT CASE WHEN fu.status_code IS NOT NULL AND fu.status_code != 200 THEN fu.id END) as error_count
            FROM scans s
            LEFT JOIN projects c ON s.project_id = c.id
            LEFT JOIN libraries l ON s.id = l.scan_id
            LEFT JOIN version_strings vs ON s.id = vs.scan_id
            LEFT JOIN file_urls fu ON s.id = fu.scan_id
            WHERE s.id IN ({placeholders})
            GROUP BY s.id
            ORDER BY s.scan_date DESC
        '''
        scans_raw = conn.execute(scans_query, paginated_vulnerable_ids).fetchall()

        # Add vulnerability count to each scan
        for scan in scans_raw:
            scan_dict = dict(scan)
            # Count vulnerabilities using has_vulnerability function
            vuln_count = 0
            libs_query = '''
                SELECT library_name, version, latest_safe_version, latest_version
                FROM libraries WHERE scan_id = ?
            '''
            libraries = conn.execute(libs_query, (scan['id'],)).fetchall()
            for lib in libraries:
                if has_vulnerability(lib['version'], lib['latest_safe_version'], lib['latest_version']):
                    vuln_count += 1
            scan_dict['vulnerability_count'] = vuln_count
            vulnerable_scans.append(scan_dict)

    # Calculate total vulnerabilities across all vulnerable scans
    total_vulnerabilities = 0
    all_scans_query = '''
        SELECT s.id FROM scans s
    '''
    all_scan_ids = [row['id'] for row in conn.execute(all_scans_query).fetchall()]

    for scan_id in vulnerable_scan_ids:
        libs_query = '''
            SELECT library_name, version, latest_safe_version, latest_version
            FROM libraries WHERE scan_id = ?
        '''
        libraries = conn.execute(libs_query, (scan_id,)).fetchall()
        for lib in libraries:
            if has_vulnerability(lib['version'], lib['latest_safe_version'], lib['latest_version']):
                total_vulnerabilities += 1

    # Get total scans for percentage calculation
    total_scans = len(all_scan_ids)

    # Get summary statistics
    summary_stats = {
        'total_vulnerable_scans': total_vulnerable_scans,
        'total_vulnerabilities': total_vulnerabilities,
        'total_scans': total_scans
    }

    # Get projects for potential filtering (future enhancement)
    projects = conn.execute('SELECT id, name FROM projects WHERE is_active = 1 ORDER BY name').fetchall()

    conn.close()

    return render_template('statistics.html',
                         vulnerable_scans=vulnerable_scans,
                         summary_stats=summary_stats,
                         projects=projects,
                         pagination={
                             'page': page,
                             'per_page': per_page,
                             'total': total_vulnerable_scans,
                             'total_pages': total_pages,
                             'has_prev': has_prev,
                             'has_next': has_next,
                             'prev_num': prev_num,
                             'next_num': next_num
                         },
                         search=search)

@app.route('/reset-database', methods=['POST'])
@login_required
def reset_database():
    try:
        # Close any existing connections first
        import gc
        gc.collect()

        # Remove the database file completely
        db_path = 'analysis.db'
        if os.path.exists(db_path):
            os.remove(db_path)

        # Also remove WAL and SHM files if they exist
        wal_path = db_path + '-wal'
        shm_path = db_path + '-shm'

        if os.path.exists(wal_path):
            os.remove(wal_path)
        if os.path.exists(shm_path):
            os.remove(shm_path)

        # Initialize a fresh database
        init_database()

        flash('¡La base de datos ha sido completamente reseteada y recreada exitosamente!', 'success')
        return redirect(url_for('index'))

    except Exception as e:
        flash(f'Error al resetear la base de datos: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/toggle-reviewed/<int:scan_id>', methods=['POST'])
@login_required
@log_action_async('UPDATE', 'scans',
                 get_target_id=lambda scan_id: scan_id,
                 get_description=lambda scan_id: f"Cambio estado revisado del escaneo #{scan_id}")
def toggle_reviewed(scan_id):
    """Toggle reviewed status of a scan"""
    conn = get_db_connection()
    try:
        # Get current reviewed status
        current_status = conn.execute('SELECT reviewed FROM scans WHERE id = ?', (scan_id,)).fetchone()
        if not current_status:
            flash('Escaneo no encontrado', 'error')
            return redirect(request.referrer or url_for('index'))

        # Toggle the status
        new_status = 1 if not current_status['reviewed'] else 0

        # Update the scan
        conn.execute('UPDATE scans SET reviewed = ? WHERE id = ?', (new_status, scan_id))
        conn.commit()

        status_text = 'marcado como revisado' if new_status else 'marcado como no revisado'
        flash(f'Escaneo {status_text} exitosamente', 'success')

    except Exception as e:
        flash(f'Error al actualizar estado de revisión: {str(e)}', 'error')
    finally:
        conn.close()

    return redirect(request.referrer or url_for('index'))

@app.route('/delete-scan/<int:scan_id>', methods=['POST'])
@login_required
@log_action('DELETE', 'scans',
           get_target_id=lambda scan_id: scan_id,
           get_description=lambda scan_id: f"Eliminación del escaneo #{scan_id}")
def delete_scan(scan_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Delete related records first (foreign key constraints)
        cursor.execute('DELETE FROM version_strings WHERE scan_id = ?', (scan_id,))
        cursor.execute('DELETE FROM file_urls WHERE scan_id = ?', (scan_id,))
        cursor.execute('DELETE FROM libraries WHERE scan_id = ?', (scan_id,))
        cursor.execute('DELETE FROM scans WHERE id = ?', (scan_id,))

        conn.commit()
        conn.close()

        flash('¡Escaneo eliminado exitosamente!', 'success')

        # Preserve URL parameters when redirecting
        project_id = request.form.get('return_project_id', '')
        search = request.form.get('return_search', '')
        page = request.form.get('return_page', '')

        # Check if we need to redirect to a specific project page
        if project_id:
            redirect_params = {'project_id': project_id}
            if search:
                redirect_params['search'] = search
            if page:
                redirect_params['page'] = page
            return redirect(url_for('project_detail', **redirect_params))
        
        # Otherwise redirect to index with preserved parameters
        redirect_params = {}
        if search:
            redirect_params['search'] = search
        if page:
            redirect_params['page'] = page

        return redirect(url_for('index', **redirect_params))

    except Exception as e:
        flash(f'Error al eliminar el escaneo: {str(e)}', 'error')
        return redirect(url_for('scan_detail', scan_id=scan_id))

@app.route('/update-scan-project/<int:scan_id>', methods=['POST'])
@login_required
def update_scan_project(scan_id):
    try:
        project_id = request.form.get('project_id', '').strip()

        # Convert empty string to None for project_id
        if not project_id:
            project_id = None

        conn = get_db_connection()
        cursor = conn.cursor()

        # Verify that the scan exists
        scan = cursor.execute('SELECT id, url FROM scans WHERE id = ?', (scan_id,)).fetchone()
        if not scan:
            flash('¡Escaneo no encontrado!', 'error')
            conn.close()
            return redirect(url_for('index'))

        # If project_id is provided, verify that the project exists
        if project_id:
            project = cursor.execute('SELECT id, name FROM projects WHERE id = ? AND is_active = 1', (project_id,)).fetchone()
            if not project:
                flash('¡Proyecto no encontrado o inactivo!', 'error')
                conn.close()
                return redirect(url_for('scan_detail', scan_id=scan_id))

        # Get project name before updating if project_id is provided
        project_name = None
        if project_id:
            project_name = cursor.execute('SELECT name FROM projects WHERE id = ?', (project_id,)).fetchone()['name']

        # Update the scan with the new project_id
        cursor.execute('UPDATE scans SET project_id = ? WHERE id = ?', (project_id, scan_id))
        conn.commit()
        conn.close()

        if project_id and project_name:
            flash(f'¡Escaneo asociado exitosamente al proyecto "{project_name}"!', 'success')
        else:
            flash('¡Escaneo desasociado de proyecto exitosamente!', 'success')

        return redirect(url_for('scan_detail', scan_id=scan_id))

    except Exception as e:
        flash(f'Error al actualizar el proyecto del escaneo: {str(e)}', 'error')
        return redirect(url_for('scan_detail', scan_id=scan_id))

@app.route('/update-scan-project-dashboard/<int:scan_id>', methods=['POST'])
@login_required
def update_scan_project_dashboard(scan_id):
    """Update scan project from dashboard and return to dashboard with preserved parameters"""
    try:
        project_id = request.form.get('project_id', '').strip()

        # Convert empty string to None for project_id
        if not project_id:
            project_id = None

        conn = get_db_connection()
        cursor = conn.cursor()

        # Verify that the scan exists
        scan = cursor.execute('SELECT id, url FROM scans WHERE id = ?', (scan_id,)).fetchone()
        if not scan:
            flash('¡Escaneo no encontrado!', 'error')
            conn.close()
            return redirect(url_for('index'))

        # If project_id is provided, verify that the project exists
        if project_id:
            project = cursor.execute('SELECT id, name FROM projects WHERE id = ? AND is_active = 1', (project_id,)).fetchone()
            if not project:
                flash('¡Proyecto no encontrado o inactivo!', 'error')
                conn.close()
                return redirect(url_for('index'))

        # Get project name before updating if project_id is provided
        project_name = None
        if project_id:
            project_name = cursor.execute('SELECT name FROM projects WHERE id = ?', (project_id,)).fetchone()['name']

        # Update the scan with the new project_id
        cursor.execute('UPDATE scans SET project_id = ? WHERE id = ?', (project_id, scan_id))
        conn.commit()
        conn.close()

        if project_id and project_name:
            flash(f'¡Escaneo asociado exitosamente al proyecto "{project_name}"!', 'success')
        else:
            flash('¡Escaneo desasociado de proyecto exitosamente!', 'success')

        # Preserve URL parameters when redirecting
        return_project_id = request.form.get('return_project_id', '')
        return_search = request.form.get('return_search', '')
        return_page = request.form.get('return_page', '')

        redirect_params = {}
        if return_project_id:
            redirect_params['project_id'] = return_project_id
        if return_search:
            redirect_params['search'] = return_search
        if return_page:
            redirect_params['page'] = return_page

        return redirect(url_for('index', **redirect_params))

    except Exception as e:
        flash(f'Error al actualizar el proyecto del escaneo: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/bulk-update-scan-project', methods=['POST'])
@login_required
def bulk_update_scan_project():
    try:
        project_id = request.form.get('project_id', '').strip()
        scan_ids_str = request.form.get('scan_ids', '').strip()

        # Convert empty string to None for project_id
        if not project_id:
            project_id = None

        # Parse scan IDs
        if not scan_ids_str:
            flash('No se seleccionaron escaneos para actualizar.', 'error')
            return redirect(url_for('index'))

        try:
            scan_ids = [int(id.strip()) for id in scan_ids_str.split(',') if id.strip()]
        except ValueError:
            flash('IDs de escaneo inválidos.', 'error')
            return redirect(url_for('index'))

        if not scan_ids:
            flash('No se seleccionaron escaneos válidos para actualizar.', 'error')
            return redirect(url_for('index'))

        conn = get_db_connection()
        cursor = conn.cursor()

        # If project_id is provided, verify that the project exists
        project_name = None
        if project_id:
            project = cursor.execute('SELECT id, name FROM projects WHERE id = ? AND is_active = 1', (project_id,)).fetchone()
            if not project:
                flash('¡Proyecto no encontrado o inactivo!', 'error')
                conn.close()
                return redirect(url_for('index'))
            project_name = project['name']

        # Update all selected scans
        placeholders = ','.join(['?'] * len(scan_ids))
        query = f'UPDATE scans SET project_id = ? WHERE id IN ({placeholders})'
        cursor.execute(query, [project_id] + scan_ids)

        updated_count = cursor.rowcount
        conn.commit()
        conn.close()

        if project_id and project_name:
            flash(f'¡{updated_count} escaneos asociados exitosamente al proyecto "{project_name}"!', 'success')
        else:
            flash(f'¡{updated_count} escaneos desasociados de proyecto exitosamente!', 'success')

        return redirect(url_for('index'))

    except Exception as e:
        flash(f'Error al actualizar proyectos de escaneos: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/re-scan/<int:scan_id>', methods=['POST'])
@login_required
def re_scan_url(scan_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Get original scan details
        original_scan = cursor.execute('''
            SELECT url, project_id
            FROM scans
            WHERE id = ?
        ''', (scan_id,)).fetchone()

        if not original_scan:
            flash('¡Escaneo original no encontrado!', 'error')
            return redirect(url_for('index'))

        conn.close()

        # Re-analyze the URL using existing function
        result = analyze_single_url(original_scan['url'], project_id=original_scan['project_id'])

        if result['success']:
            flash(f'¡Re-escaneo completado! Se encontraron {result["libraries_count"]} librerías, {result["files_count"]} archivos, y {result["version_strings_count"]} cadenas de versión.', 'success')
            return redirect(url_for('scan_detail', scan_id=result['scan_id']))
        else:
            flash(f'Re-escaneo fallido: {result["error"]}', 'error')
            if result.get('scan_id'):
                return redirect(url_for('scan_detail', scan_id=result['scan_id']))
            else:
                return redirect(url_for('scan_detail', scan_id=scan_id))

    except Exception as e:
        flash(f'Error durante el re-escaneo: {str(e)}', 'error')
        return redirect(url_for('scan_detail', scan_id=scan_id))

@app.route('/url-history/<int:scan_id>')
@login_required
def url_history(scan_id):
    try:
        conn = get_db_connection()

        # Get the URL from the original scan
        original_scan = conn.execute('SELECT url FROM scans WHERE id = ?', (scan_id,)).fetchone()
        if not original_scan:
            flash('¡Escaneo no encontrado!', 'error')
            return redirect(url_for('index'))

        url = original_scan['url']

        # Get all scans for this URL with additional details
        scans = conn.execute('''
            SELECT s.id, s.url, s.scan_date, s.status_code, s.title, s.project_id,
                   p.name as project_name,
                   COUNT(DISTINCT l.id) as library_count,
                   COUNT(DISTINCT vs.id) as version_string_count,
                   COUNT(DISTINCT fu.id) as file_count
            FROM scans s
            LEFT JOIN projects p ON s.project_id = p.id AND p.is_active = 1
            LEFT JOIN libraries l ON s.id = l.scan_id
            LEFT JOIN version_strings vs ON s.id = vs.scan_id
            LEFT JOIN file_urls fu ON s.id = fu.scan_id
            WHERE s.url = ?
            GROUP BY s.id, s.url, s.scan_date, s.status_code, s.title, s.project_id, p.name
            ORDER BY s.scan_date DESC
        ''', (url,)).fetchall()

        # Get latest safe version and latest version for each library across all scans
        library_summary = conn.execute('''
            SELECT
                l.library_name,
                l.type,
                COUNT(DISTINCT s.id) as scan_count,
                MIN(s.scan_date) as first_detected,
                MAX(s.scan_date) as last_detected,
                GROUP_CONCAT(DISTINCT l.version) as versions_found
            FROM libraries l
            JOIN scans s ON l.scan_id = s.id
            WHERE s.url = ?
            GROUP BY l.library_name, l.type
            ORDER BY l.library_name
        ''', (url,)).fetchall()

        conn.close()

        return render_template('url_history.html',
                             scans=scans,
                             library_summary=library_summary,
                             url=url,
                             current_scan_id=scan_id,
                             total_scans=len(scans))

    except Exception as e:
        flash(f'Error al obtener historial: {str(e)}', 'error')
        return redirect(url_for('scan_detail', scan_id=scan_id))

@app.route('/compare-scans/<int:scan_id1>/<int:scan_id2>')
@login_required
def compare_scans(scan_id1, scan_id2):
    """Compare two scans and show differences"""
    try:
        conn = get_db_connection()

        # Get both scans
        scan1_raw = conn.execute('''
            SELECT s.*, c.name as project_name
            FROM scans s
            LEFT JOIN projects c ON s.project_id = c.id
            WHERE s.id = ?
        ''', (scan_id1,)).fetchone()

        scan2_raw = conn.execute('''
            SELECT s.*, c.name as project_name
            FROM scans s
            LEFT JOIN projects c ON s.project_id = c.id
            WHERE s.id = ?
        ''', (scan_id2,)).fetchone()

        if not scan1_raw or not scan2_raw:
            flash('Uno o ambos escaneos no existen', 'error')
            return redirect(url_for('index'))

        # Verify both scans are from the same URL
        if scan1_raw['url'] != scan2_raw['url']:
            flash('Los escaneos deben ser de la misma URL para comparar', 'error')
            return redirect(url_for('scan_detail', scan_id=scan_id1))

        # Order scans by date: scan1 = older (anterior), scan2 = newer (actual)
        if scan1_raw['scan_date'] > scan2_raw['scan_date']:
            # scan1_raw is newer, so swap them
            scan1 = scan2_raw  # older scan (anterior)
            scan2 = scan1_raw  # newer scan (actual)
            scan_id1, scan_id2 = scan_id2, scan_id1  # update IDs for library queries
        else:
            # scan1_raw is older, keep as is
            scan1 = scan1_raw  # older scan (anterior)
            scan2 = scan2_raw  # newer scan (actual)

        # Get libraries for both scans
        libraries1 = conn.execute('''
            SELECT library_name, type, version, description,
                   latest_safe_version, latest_version, is_manual
            FROM libraries
            WHERE scan_id = ?
            ORDER BY library_name
        ''', (scan_id1,)).fetchall()

        libraries2 = conn.execute('''
            SELECT library_name, type, version, description,
                   latest_safe_version, latest_version, is_manual
            FROM libraries
            WHERE scan_id = ?
            ORDER BY library_name
        ''', (scan_id2,)).fetchall()

        # Create dictionaries for easier comparison
        libs1_dict = {(lib['library_name'], lib['type']): lib for lib in libraries1}
        libs2_dict = {(lib['library_name'], lib['type']): lib for lib in libraries2}

        # Compare libraries
        all_lib_keys = set(libs1_dict.keys()) | set(libs2_dict.keys())

        libraries_comparison = {
            'added': [],
            'removed': [],
            'updated': [],
            'unchanged': []
        }

        for lib_key in all_lib_keys:
            lib1 = libs1_dict.get(lib_key)
            lib2 = libs2_dict.get(lib_key)

            if lib1 and not lib2:
                # Library was removed
                libraries_comparison['removed'].append({
                    'library': lib1,
                    'change_type': 'removed'
                })
            elif not lib1 and lib2:
                # Library was added
                libraries_comparison['added'].append({
                    'library': lib2,
                    'change_type': 'added'
                })
            elif lib1 and lib2:
                # Library exists in both - check for version changes
                if lib1['version'] != lib2['version']:
                    libraries_comparison['updated'].append({
                        'library_name': lib1['library_name'],
                        'type': lib1['type'],
                        'old_version': lib1['version'],
                        'new_version': lib2['version'],
                        'description': lib2['description'],
                        'change_type': 'updated'
                    })
                else:
                    libraries_comparison['unchanged'].append({
                        'library': lib1,
                        'change_type': 'unchanged'
                    })

        # Get file counts
        files1_count = conn.execute('SELECT COUNT(*) as cnt FROM file_urls WHERE scan_id = ? AND file_type = "js"', (scan_id1,)).fetchone()['cnt']
        files2_count = conn.execute('SELECT COUNT(*) as cnt FROM file_urls WHERE scan_id = ? AND file_type = "js"', (scan_id2,)).fetchone()['cnt']

        # Get version strings counts
        versions1_count = conn.execute('SELECT COUNT(*) as cnt FROM version_strings WHERE scan_id = ?', (scan_id1,)).fetchone()['cnt']
        versions2_count = conn.execute('SELECT COUNT(*) as cnt FROM version_strings WHERE scan_id = ?', (scan_id2,)).fetchone()['cnt']

        # Calculate security scores if headers exist
        security1_score = 0
        security2_score = 0
        headers1_analysis = None
        headers2_analysis = None

        if scan1['headers']:
            try:
                headers1 = json.loads(scan1['headers'])
                security1_score = calculate_security_score(headers1)
                headers1_analysis = analyze_security_headers(headers1)
            except:
                pass

        if scan2['headers']:
            try:
                headers2 = json.loads(scan2['headers'])
                security2_score = calculate_security_score(headers2)
                headers2_analysis = analyze_security_headers(headers2)
            except:
                pass

        # Calculate metrics differences
        metrics_diff = {
            'libraries': len(libraries2) - len(libraries1),
            'files': files2_count - files1_count,
            'versions': versions2_count - versions1_count,
            'security': security2_score - security1_score
        }

        # Create summary
        total_changes = (len(libraries_comparison['added']) +
                        len(libraries_comparison['removed']) +
                        len(libraries_comparison['updated']))

        # Calculate improvement score (simplified)
        improvement_score = 50  # Base score
        if metrics_diff['security'] > 0:
            improvement_score += 20
        elif metrics_diff['security'] < 0:
            improvement_score -= 20

        if len(libraries_comparison['added']) > len(libraries_comparison['removed']):
            improvement_score += 10
        elif len(libraries_comparison['removed']) > len(libraries_comparison['added']):
            improvement_score -= 10

        # Check for version updates (usually good)
        if len(libraries_comparison['updated']) > 0:
            improvement_score += 15

        improvement_score = max(0, min(100, improvement_score))  # Clamp between 0-100

        comparison_data = {
            'scan1': scan1,
            'scan2': scan2,
            'libraries1': libraries1,
            'libraries2': libraries2,
            'libraries_comparison': libraries_comparison,
            'metrics_diff': metrics_diff,
            'security1_score': security1_score,
            'security2_score': security2_score,
            'headers1_analysis': headers1_analysis,
            'headers2_analysis': headers2_analysis,
            'files1_count': files1_count,
            'files2_count': files2_count,
            'versions1_count': versions1_count,
            'versions2_count': versions2_count,
            'total_changes': total_changes,
            'improvement_score': improvement_score
        }

        conn.close()
        return render_template('scan_comparison.html', **comparison_data)

    except Exception as e:
        flash(f'Error al comparar escaneos: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/delete-version-string/<version_string_ids>', methods=['POST'])
@login_required
def delete_version_string(version_string_ids):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Parse comma-separated IDs
        try:
            ids = [int(id_str.strip()) for id_str in version_string_ids.split(',') if id_str.strip()]
        except ValueError:
            flash('IDs de cadenas de versión inválidos', 'error')
            return redirect(url_for('index'))

        if not ids:
            flash('No se proporcionaron IDs válidos', 'error')
            return redirect(url_for('index'))

        # Get the scan_id before deleting to redirect back to scan details
        version_string = cursor.execute('SELECT scan_id FROM version_strings WHERE id = ? LIMIT 1', (ids[0],)).fetchone()

        if not version_string:
            flash('¡Cadenas de versión no encontradas!', 'error')
            return redirect(url_for('index'))

        scan_id = version_string['scan_id']

        # Delete all version strings with the provided IDs
        placeholders = ','.join('?' * len(ids))
        deleted_count = cursor.execute(f'DELETE FROM version_strings WHERE id IN ({placeholders})', ids).rowcount

        conn.commit()
        conn.close()

        if deleted_count == 1:
            flash('¡Cadena de versión eliminada exitosamente!', 'success')
        else:
            flash(f'¡{deleted_count} cadenas de versión eliminadas exitosamente!', 'success')
        return redirect(url_for('scan_detail', scan_id=scan_id))

    except Exception as e:
        flash(f'Error al eliminar cadena de versión: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/batch-delete-version-strings', methods=['POST'])
@login_required
def batch_delete_version_strings():
    try:
        version_string_ids = request.form.getlist('version_string_ids[]')

        if not version_string_ids:
            flash('¡No se seleccionaron cadenas de versión para eliminar!', 'error')
            return redirect(url_for('index'))

        conn = get_db_connection()
        cursor = conn.cursor()

        # Get scan_id from the first version string to know where to redirect
        cursor.execute('SELECT scan_id FROM version_strings WHERE id = ?', (version_string_ids[0],))
        scan_row = cursor.fetchone()
        scan_id = scan_row[0] if scan_row else None

        if not scan_id:
            flash('¡Cadenas de versión no encontradas!', 'error')
            conn.close()
            return redirect(url_for('index'))

        # Delete selected version strings (secure SQL with proper parameterization)
        placeholders = ','.join(['?'] * len(version_string_ids))
        cursor.execute(f'DELETE FROM version_strings WHERE id IN ({placeholders})', tuple(version_string_ids))

        conn.commit()
        conn.close()

        flash(f'¡Se eliminaron exitosamente {len(version_string_ids)} cadenas de versión!', 'success')
        return redirect(url_for('scan_detail', scan_id=scan_id))

    except Exception as e:
        flash(f'Error al eliminar cadenas de versión: {str(e)}', 'error')
        # Try to redirect to index if we can't get scan_id
        try:
            return redirect(url_for('scan_detail', scan_id=scan_id))
        except:
            return redirect(url_for('index'))

@app.route('/batch-delete-file-urls', methods=['POST'])
@login_required
def batch_delete_file_urls():
    try:
        file_url_ids = request.form.getlist('file_url_ids[]')

        if not file_url_ids:
            flash('¡No se seleccionaron URLs de archivos para eliminar!', 'error')
            return redirect(url_for('index'))

        conn = get_db_connection()
        cursor = conn.cursor()

        # Get scan_id from the first file URL to know where to redirect
        cursor.execute('SELECT scan_id FROM file_urls WHERE id = ?', (file_url_ids[0],))
        scan_row = cursor.fetchone()
        scan_id = scan_row[0] if scan_row else None

        if not scan_id:
            flash('¡URLs de archivos no encontradas!', 'error')
            conn.close()
            return redirect(url_for('index'))

        # Delete selected file URLs (secure SQL with proper parameterization)
        placeholders = ','.join(['?'] * len(file_url_ids))
        cursor.execute(f'DELETE FROM file_urls WHERE id IN ({placeholders})', tuple(file_url_ids))

        conn.commit()
        conn.close()

        flash(f'¡Se eliminaron exitosamente {len(file_url_ids)} URLs de archivos!', 'success')
        return redirect(url_for('scan_detail', scan_id=scan_id))

    except Exception as e:
        flash(f'Error al eliminar URLs de archivos: {str(e)}', 'error')
        # Try to redirect to index if we can't get scan_id
        try:
            return redirect(url_for('scan_detail', scan_id=scan_id))
        except:
            return redirect(url_for('index'))

@app.route('/delete-file-url/<int:file_url_id>', methods=['POST'])
@login_required
def delete_individual_file_url(file_url_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Get the scan_id and file URL before deleting
        file_url_record = cursor.execute('SELECT scan_id, file_url FROM file_urls WHERE id = ?', (file_url_id,)).fetchone()
        if not file_url_record:
            flash('¡URL de archivo no encontrada!', 'error')
            return redirect(url_for('index'))

        scan_id = file_url_record['scan_id']
        file_url = file_url_record['file_url']

        cursor.execute('DELETE FROM file_urls WHERE id = ?', (file_url_id,))

        conn.commit()
        conn.close()

        # Truncate URL for display if it's too long
        display_url = file_url[:50] + '...' if len(file_url) > 50 else file_url
        flash(f'URL de archivo "{display_url}" eliminada exitosamente!', 'success')
        return redirect(url_for('scan_detail', scan_id=scan_id))

    except Exception as e:
        flash(f'Error al eliminar URL de archivo: {str(e)}', 'error')
        return redirect(request.referrer or url_for('index'))

@app.route('/add-manual-library', methods=['POST'])
@login_required
def add_manual_library():
    try:
        scan_id = request.form.get('scan_id')
        library_name = request.form.get('library_name', '').strip()
        library_type = request.form.get('library_type')
        version = request.form.get('version', '').strip()
        source_url = request.form.get('source_url', '').strip()
        description = request.form.get('description', '').strip()
        latest_safe_version = request.form.get('latest_safe_version', '').strip()
        latest_version = request.form.get('latest_version', '').strip()
        global_library_id = request.form.get('global_library_id')

        if not library_name:
            flash('¡El nombre de la librería es requerido!', 'error')
            return redirect(url_for('scan_detail', scan_id=scan_id) + '#libraries')

        if not library_type or library_type not in ['js', 'css']:
            flash('¡Se requiere un tipo de librería válido (js o css)!', 'error')
            return redirect(url_for('scan_detail', scan_id=scan_id) + '#libraries')

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO libraries (scan_id, library_name, version, type, source_url,
                                 description, latest_safe_version, latest_version, is_manual, global_library_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
        ''', (scan_id, library_name, version or None, library_type, source_url or None,
              description or None, latest_safe_version or None, latest_version or None,
              global_library_id if global_library_id and global_library_id.isdigit() else None))

        conn.commit()
        conn.close()

        flash(f'Librería manual "{library_name}" agregada exitosamente!', 'success')
        return redirect(url_for('scan_detail', scan_id=scan_id) + '#libraries')

    except Exception as e:
        flash(f'Error al agregar librería manual: {str(e)}', 'error')
        return redirect(url_for('scan_detail', scan_id=scan_id) + '#libraries')

@app.route('/edit-library/<int:library_id>', methods=['POST'])
@login_required
def edit_library(library_id):
    try:
        library_name = request.form.get('library_name', '').strip()
        library_type = request.form.get('library_type')
        version = request.form.get('version', '').strip()
        source_url = request.form.get('source_url', '').strip()
        description = request.form.get('description', '').strip()
        latest_safe_version = request.form.get('latest_safe_version', '').strip()
        latest_version = request.form.get('latest_version', '').strip()
        global_library_id = request.form.get('global_library_id')

        conn = get_db_connection()
        cursor = conn.cursor()

        # Get the scan_id before updating
        library = cursor.execute('SELECT scan_id FROM libraries WHERE id = ?', (library_id,)).fetchone()
        if not library:
            flash('¡Librería no encontrada!', 'error')
            return redirect(url_for('index'))

        scan_id = library['scan_id']

        # If a global library is selected, get its data to override name and versions
        if global_library_id and global_library_id.isdigit():
            global_lib = cursor.execute('''
                SELECT library_name, latest_safe_version, latest_version, description, type
                FROM global_libraries WHERE id = ?
            ''', (int(global_library_id),)).fetchone()

            if global_lib:
                # Override with global library data (keep version and source_url from form)
                library_name = global_lib['library_name']
                latest_safe_version = global_lib['latest_safe_version'] or ''
                latest_version = global_lib['latest_version'] or ''
                if not description.strip():  # Only use global description if form description is empty
                    description = global_lib['description'] or ''
                # Use global library type if it's set
                if global_lib['type'] and not library_type:
                    library_type = global_lib['type']

        if not library_name:
            flash('¡El nombre de la librería es requerido!', 'error')
            return redirect(request.referrer or url_for('index'))

        if not library_type or library_type not in ['js', 'css']:
            flash('¡Se requiere un tipo de librería válido (js o css)!', 'error')
            return redirect(request.referrer or url_for('index'))

        cursor.execute('''
            UPDATE libraries
            SET library_name = ?, version = ?, type = ?, source_url = ?,
                description = ?, latest_safe_version = ?, latest_version = ?,
                global_library_id = ?
            WHERE id = ?
        ''', (library_name, version or None, library_type, source_url or None,
              description or None, latest_safe_version or None, latest_version or None,
              global_library_id if global_library_id and global_library_id.isdigit() else None, library_id))

        conn.commit()
        conn.close()

        flash(f'Librería "{library_name}" actualizada exitosamente!', 'success')
        return redirect(url_for('scan_detail', scan_id=scan_id) + '#libraries')

    except Exception as e:
        flash(f'Error al actualizar librería: {str(e)}', 'error')
        return redirect(request.referrer or url_for('index'))

@app.route('/delete-library/<int:library_id>', methods=['POST'])
@login_required
def delete_library(library_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Get the scan_id and library name before deleting
        library = cursor.execute('SELECT scan_id, library_name FROM libraries WHERE id = ?', (library_id,)).fetchone()
        if not library:
            flash('¡Librería no encontrada!', 'error')
            return redirect(url_for('index'))

        scan_id = library['scan_id']
        library_name = library['library_name']

        cursor.execute('DELETE FROM libraries WHERE id = ?', (library_id,))

        conn.commit()
        conn.close()

        flash(f'Librería "{library_name}" eliminada exitosamente!', 'success')
        return redirect(url_for('scan_detail', scan_id=scan_id) + '#libraries')

    except Exception as e:
        flash(f'Error al eliminar librería: {str(e)}', 'error')
        return redirect(request.referrer or url_for('index'))

def get_scan_export_data(scan_id):
    """Get all scan data for export"""
    conn = get_db_connection()

    # Get scan info
    scan = conn.execute('SELECT * FROM scans WHERE id = ?', (scan_id,)).fetchone()
    if not scan:
        conn.close()
        return None

    # Get libraries (JavaScript only)
    libraries = conn.execute('''
        SELECT id, library_name, version, type, source_url, description,
               latest_safe_version, latest_version, is_manual
        FROM libraries
        WHERE scan_id = ? AND type = 'js'
        ORDER BY type, library_name
    ''', (scan_id,)).fetchall()

    # Get version strings (JavaScript files only)
    version_strings = conn.execute('''
        SELECT id, file_url, file_type, line_number, line_content, version_keyword
        FROM version_strings
        WHERE scan_id = ? AND file_type = 'js'
        ORDER BY file_url, line_number
    ''', (scan_id,)).fetchall()

    # Get file URLs (JavaScript files only)
    file_urls = conn.execute('''
        SELECT id, file_url, file_type, file_size, status_code
        FROM file_urls
        WHERE scan_id = ? AND file_type = 'js'
        ORDER BY file_type, file_url
    ''', (scan_id,)).fetchall()

    # Parse headers and security analysis
    headers = json.loads(scan['headers']) if scan['headers'] else {}
    security_analysis = analyze_security_headers(headers)

    conn.close()

    # Convert all sqlite3.Row objects to dictionaries before returning
    result = {
        'scan': scan,
        'libraries': libraries,
        'version_strings': version_strings,
        'file_urls': file_urls,
        'headers': headers,
        'security_analysis': security_analysis
    }

    # Deep convert all Row objects to ensure JSON serialization compatibility
    return convert_rows_deep(result)

def get_project_consolidated_data(project_id):
    """
    Get consolidated data for project report:
    1. Filter only scans with reviewed = 1  
    2. Deduplicate by URL (take most recent)
    3. Aggregate all libraries and files
    4. Consolidate security headers
    """
    conn = get_db_connection()
    
    # Get project info
    project = conn.execute('SELECT * FROM projects WHERE id = ? AND is_active = 1', (project_id,)).fetchone()
    if not project:
        conn.close()
        return None
    
    # Get latest reviewed scans for each URL in project (deduplication logic)
    latest_scans_query = '''
        WITH latest_scans AS (
            SELECT url, MAX(scan_date) as latest_date
            FROM scans 
            WHERE project_id = ? AND reviewed = 1
            GROUP BY url
        )
        SELECT s.*, ls.url as deduplicated_url
        FROM scans s
        INNER JOIN latest_scans ls ON s.url = ls.url AND s.scan_date = ls.latest_date
        WHERE s.project_id = ? AND s.reviewed = 1
        ORDER BY s.scan_date DESC
    '''
    
    scans = conn.execute(latest_scans_query, (project_id, project_id)).fetchall()
    
    if not scans:
        conn.close()
        return {
            'project': convert_rows_deep(project),
            'scans': [],
            'consolidated_libraries': [],
            'consolidated_file_urls': [], 
            'consolidated_headers': {},
            'consolidated_security_analysis': {},
            'urls_data': [],
            'project_stats': {
                'total_urls': 0,
                'total_scans': 0,
                'total_libraries': 0,
                'total_vulnerabilities': 0,
                'total_files': 0
            }
        }
    
    # Extract scan IDs for further queries
    scan_ids = [scan['id'] for scan in scans]
    scan_ids_placeholder = ','.join('?' * len(scan_ids))
    
    # Get all libraries from deduplicated scans
    libraries_query = f'''
        SELECT l.*, s.url, s.scan_date
        FROM libraries l
        INNER JOIN scans s ON l.scan_id = s.id
        WHERE l.scan_id IN ({scan_ids_placeholder}) AND l.type = 'js'
        ORDER BY s.url, l.library_name
    '''
    all_libraries = conn.execute(libraries_query, scan_ids).fetchall()
    
    # Get all file URLs from deduplicated scans
    file_urls_query = f'''
        SELECT f.*, s.url as scan_url
        FROM file_urls f
        INNER JOIN scans s ON f.scan_id = s.id  
        WHERE f.scan_id IN ({scan_ids_placeholder}) AND f.file_type = 'js'
        ORDER BY s.url, f.file_url
    '''
    all_file_urls = conn.execute(file_urls_query, scan_ids).fetchall()
    
    # Get all version strings from deduplicated scans
    version_strings_query = f'''
        SELECT v.*, s.url as scan_url
        FROM version_strings v
        INNER JOIN scans s ON v.scan_id = s.id
        WHERE v.scan_id IN ({scan_ids_placeholder}) AND v.file_type = 'js'
        ORDER BY s.url, v.file_url, v.line_number
    '''
    all_version_strings = conn.execute(version_strings_query, scan_ids).fetchall()
    
    conn.close()
    
    # Process and deduplicate libraries across URLs
    consolidated_libraries = deduplicate_libraries(all_libraries)
    
    # Consolidate security headers from all scans
    consolidated_headers, consolidated_security_analysis = consolidate_security_headers(scans)
    
    # Create URL-specific data for detailed view
    urls_data = []
    for scan in scans:
        url_libraries = [lib for lib in all_libraries if lib['scan_id'] == scan['id']]
        url_files = [f for f in all_file_urls if f['scan_id'] == scan['id']]
        
        # Parse headers for this specific URL
        scan_headers = json.loads(scan['headers']) if scan['headers'] else {}
        url_security_analysis = analyze_security_headers(scan_headers)
        
        urls_data.append({
            'scan': convert_rows_deep(scan),
            'libraries': convert_rows_deep(url_libraries),
            'file_urls': convert_rows_deep(url_files),
            'headers': scan_headers,
            'security_analysis': url_security_analysis
        })
    
    # Calculate project statistics
    project_stats = calculate_project_stats(scans, consolidated_libraries, all_file_urls)
    
    result = {
        'project': convert_rows_deep(project),
        'scans': convert_rows_deep(scans),
        'consolidated_libraries': consolidated_libraries,
        'consolidated_file_urls': convert_rows_deep(all_file_urls),
        'consolidated_version_strings': convert_rows_deep(all_version_strings),
        'consolidated_headers': consolidated_headers,
        'consolidated_security_analysis': consolidated_security_analysis,
        'urls_data': urls_data,
        'project_stats': project_stats
    }
    
    return result

def deduplicate_libraries(all_libraries):
    """
    Deduplicate libraries across URLs:
    - Group by library_name and version
    - Keep track of which URLs use each library
    - Maintain vulnerability information
    """
    libraries_dict = {}
    
    for lib in all_libraries:
        key = f"{lib['library_name']}_{lib['version'] or 'unknown'}"
        
        if key not in libraries_dict:
            lib_dict = dict(lib)
            lib_dict['used_in_urls'] = [lib['url']]
            lib_dict['scan_count'] = 1
            libraries_dict[key] = lib_dict
        else:
            # Add URL to the list if not already present
            if lib['url'] not in libraries_dict[key]['used_in_urls']:
                libraries_dict[key]['used_in_urls'].append(lib['url'])
                libraries_dict[key]['scan_count'] += 1
    
    # Convert back to list and sort
    consolidated = list(libraries_dict.values())
    consolidated.sort(key=lambda x: (x['library_name'], x['version'] or ''))
    
    return consolidated

def consolidate_security_headers(scans):
    """
    Consolidate security headers from multiple scans:
    - Analyze headers present/missing across all URLs
    - Calculate overall security score
    - Generate consolidated recommendations
    """
    all_headers = {}
    security_analyses = []
    
    # Process each scan's headers
    for scan in scans:
        scan_headers = json.loads(scan['headers']) if scan['headers'] else {}
        all_headers[scan['url']] = scan_headers
        
        # Get security analysis for this scan
        security_analysis = analyze_security_headers(scan_headers)
        security_analyses.append({
            'url': scan['url'],
            'analysis': security_analysis
        })
    
    # Consolidate security analysis across all URLs
    all_security_headers = {
        'Strict-Transport-Security': {'name': 'Strict-Transport-Security', 'description': 'Fuerza conexiones HTTPS seguras', 'recommendation': 'max-age=31536000; includeSubDomains'},
        'Content-Security-Policy': {'name': 'Content-Security-Policy', 'description': 'Previene ataques XSS y de inyección de código', 'recommendation': "default-src 'self'"},
        'X-Frame-Options': {'name': 'X-Frame-Options', 'description': 'Previene ataques de clickjacking', 'recommendation': 'DENY'},
        'X-Content-Type-Options': {'name': 'X-Content-Type-Options', 'description': 'Previene ataques de MIME sniffing', 'recommendation': 'nosniff'},
        'X-XSS-Protection': {'name': 'X-XSS-Protection', 'description': 'Activa protección XSS del navegador', 'recommendation': '1; mode=block'},
        'Referrer-Policy': {'name': 'Referrer-Policy', 'description': 'Controla información del referrer', 'recommendation': 'strict-origin-when-cross-origin'},
        'Permissions-Policy': {'name': 'Permissions-Policy', 'description': 'Controla características del navegador', 'recommendation': 'geolocation=(), microphone=(), camera=()'}
    }
    
    present_headers = []
    missing_headers = []
    
    # Determine which headers are consistently present/missing
    for header_name, header_info in all_security_headers.items():
        urls_with_header = []
        urls_without_header = []
        header_values = []
        
        for url, headers in all_headers.items():
            if header_name in headers:
                urls_with_header.append(url)
                header_values.append(headers[header_name])
            else:
                urls_without_header.append(url)
        
        if urls_with_header:
            # Header is present in at least one URL
            present_headers.append({
                'name': header_name,
                'description': header_info['description'],
                'value': header_values[0] if header_values else '',  # Use first occurrence
                'present_in_urls': urls_with_header,
                'missing_in_urls': urls_without_header
            })
        else:
            # Header is missing in all URLs
            missing_headers.append({
                'name': header_name,
                'description': header_info['description'],
                'recommendation': header_info['recommendation'],
                'missing_in_urls': list(all_headers.keys())
            })
    
    # Calculate consolidated security score
    total_headers = len(all_security_headers)
    present_count = len(present_headers)
    security_score = int((present_count / total_headers) * 100) if total_headers > 0 else 0
    
    consolidated_security_analysis = {
        'present': present_headers,
        'missing': missing_headers,
        'security_score': security_score,
        'total_urls': len(all_headers)
    }
    
    return all_headers, consolidated_security_analysis

def calculate_project_stats(scans, consolidated_libraries, all_file_urls):
    """Calculate consolidated project statistics"""
    
    # Count vulnerabilities using global library data
    vulnerable_count = 0
    for lib in consolidated_libraries:
        # Check if this library has vulnerability using existing logic
        effective_safe = get_effective_safe_version(
            lib.get('latest_safe_version'), 
            lib.get('gl_latest_safe_version')
        )
        if check_vulnerability_with_global(
            lib.get('version'),
            lib.get('latest_safe_version'), 
            lib.get('gl_latest_safe_version')
        ):
            vulnerable_count += 1
    
    stats = {
        'total_urls': len(scans),
        'total_scans': len(scans),
        'total_libraries': len(consolidated_libraries),
        'total_vulnerabilities': vulnerable_count,
        'total_files': len(all_file_urls),
        'urls_analyzed': [scan['url'] for scan in scans]
    }
    
    return stats

@app.route('/report/enhanced/<int:scan_id>')
@login_required
def enhanced_report(scan_id):
    """Display enhanced HTML report with professional styling"""
    try:
        data = get_scan_export_data(scan_id)
        if not data:
            flash('Escaneo no encontrado', 'error')
            return redirect(url_for('index'))

        # All data is now already converted to dictionaries by get_scan_export_data
        # Pre-serialize libraries to JSON for JavaScript
        import json
        libraries_json = json.dumps(data['libraries'])

        return render_template('enhanced_report.html',
                             scan=data['scan'],
                             libraries=data['libraries'],
                             libraries_json=libraries_json,
                             file_urls=data['file_urls'],
                             version_strings=data['version_strings'],
                             security_analysis=data['security_analysis'],
                             headers=data['headers'])

    except Exception as e:
        print(f"Enhanced report error: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error al generar reporte: {str(e)}', 'error')
        return redirect(url_for('scan_detail', scan_id=scan_id))

@app.route('/report/project/<int:project_id>')
@login_required
def project_consolidated_report(project_id):
    """Display consolidated HTML report for project with all reviewed scans"""
    try:
        data = get_project_consolidated_data(project_id)
        if not data:
            flash('Proyecto no encontrado', 'error')
            return redirect(url_for('projects'))
        
        # Check if project has any reviewed scans
        if not data['scans']:
            flash('No hay escaneos revisados en este proyecto para generar el reporte', 'warning')
            return redirect(url_for('project_detail', project_id=project_id))
        
        # Pre-serialize data to JSON for JavaScript
        import json
        consolidated_libraries_json = json.dumps(data['consolidated_libraries'])
        urls_data_json = json.dumps(data['urls_data'])
        project_stats_json = json.dumps(data['project_stats'])
        
        return render_template('project_consolidated_report.html',
                             project=data['project'],
                             scans=data['scans'],
                             consolidated_libraries=data['consolidated_libraries'],
                             consolidated_libraries_json=consolidated_libraries_json,
                             consolidated_file_urls=data['consolidated_file_urls'],
                             consolidated_version_strings=data['consolidated_version_strings'],
                             consolidated_headers=data['consolidated_headers'],
                             consolidated_security_analysis=data['consolidated_security_analysis'],
                             urls_data=data['urls_data'],
                             urls_data_json=urls_data_json,
                             project_stats=data['project_stats'],
                             project_stats_json=project_stats_json)
    
    except Exception as e:
        print(f"Project consolidated report error: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error al generar reporte consolidado: {str(e)}', 'error')
        return redirect(url_for('project_detail', project_id=project_id))

@app.route('/export/pdf/<int:scan_id>')
@login_required
@rate_limit('export')
def export_pdf(scan_id):
    """Generate enhanced PDF report with professional styling"""
    try:
        data = get_scan_export_data(scan_id)
        if not data:
            flash('Escaneo no encontrado', 'error')
            return redirect(url_for('index'))

        # Import the enhanced PDF generator
        try:
            from pdf_report_enhanced import create_enhanced_pdf_report
            buffer = create_enhanced_pdf_report(data)
        except ImportError:
            # Fallback to basic PDF generation if enhanced module fails
            buffer = create_basic_pdf_report(data)

        filename = f"security_analysis_report_{scan_id}_{get_chile_time().strftime('%Y%m%d_%H%M%S')}.pdf"

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )

    except Exception as e:
        flash(f'Error al generar PDF: {str(e)}', 'error')
        return redirect(url_for('scan_detail', scan_id=scan_id))

def create_basic_pdf_report(data):
    """Fallback basic PDF generation"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=18,
        spaceAfter=20
    )
    story.append(Paragraph(f"Security Analysis Report - {data['scan']['url']}", title_style))
    story.append(Spacer(1, 20))

    # Scan Information
    story.append(Paragraph("Scan Information", styles['Heading2']))
    scan_info = [
        ['URL', data['scan']['url']],
        ['Title', data['scan']['title'] or 'No title'],
        ['Status Code', str(data['scan']['status_code'] or 'Error')],
        ['Scan Date', data['scan']['scan_date']],
        ['Libraries Found', str(len(data['libraries']))],
        ['Files Found', str(len(data['file_urls']))],
        ['Security Score', f"{data['security_analysis']['security_score']}%"]
    ]

    scan_table = Table(scan_info, colWidths=[2*inch, 4*inch])
    scan_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(scan_table)
    story.append(Spacer(1, 20))

    # Security Analysis
    story.append(Paragraph("Security Analysis", styles['Heading2']))
    story.append(Paragraph(f"Security Score: {data['security_analysis']['security_score']}%", styles['Normal']))

    if data['security_analysis']['present']:
        story.append(Paragraph("✅ Present Security Headers:", styles['Heading3']))
        for header in data['security_analysis']['present']:
            story.append(Paragraph(f"• {header['name']}: {header['value'][:100]}{'...' if len(header['value']) > 100 else ''}", styles['Normal']))

    if data['security_analysis']['missing']:
        story.append(Paragraph("🔴 Missing Security Headers:", styles['Heading3']))
        for header in data['security_analysis']['missing']:
            story.append(Paragraph(f"• {header['name']}: {header['recommendation']}", styles['Normal']))

    story.append(Spacer(1, 20))

    # Libraries with enhanced vulnerability display
    if data['libraries']:
        story.append(Paragraph("Detected Libraries", styles['Heading2']))
        lib_data = [['Library', 'Version', 'Type', 'Status', 'Recommendation']]

        for lib in data['libraries']:
            # Convert sqlite3.Row to dict for compatibility
            lib_dict = dict(lib) if hasattr(lib, 'keys') else lib

            # Enhanced vulnerability check
            has_vuln = (lib_dict.get('latest_safe_version') and lib_dict.get('version') and
                       lib_dict['version'] != lib_dict['latest_safe_version'] and
                       lib_dict['version'] < lib_dict['latest_safe_version'])

            if has_vuln:
                status = "🔴 VULNERABLE"
                recommendation = f"Update to {lib_dict['latest_safe_version']}"
            elif lib_dict.get('latest_safe_version'):
                status = "✅ SAFE"
                recommendation = "No action needed"
            else:
                status = "⚠️ UNKNOWN"
                recommendation = "Manual review"

            lib_data.append([
                lib_dict['library_name'],
                lib_dict.get('version', 'Unknown'),
                lib_dict['type'].upper(),
                status,
                recommendation
            ])

        lib_table = Table(lib_data, colWidths=[1.5*inch, 1*inch, 0.5*inch, 1*inch, 2*inch])
        lib_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(lib_table)

    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer

@app.route('/export/csv/<int:scan_id>')
@login_required
@rate_limit('export')
def export_csv(scan_id):
    try:
        data = get_scan_export_data(scan_id)
        if not data:
            flash('Escaneo no encontrado', 'error')
            return redirect(url_for('index'))

        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)

        # Scan Information
        writer.writerow(['=== SCAN INFORMATION ==='])
        writer.writerow(['URL', data['scan']['url']])
        writer.writerow(['Title', data['scan']['title'] or 'No title'])
        writer.writerow(['Status Code', data['scan']['status_code'] or 'Error'])
        writer.writerow(['Scan Date', data['scan']['scan_date']])
        writer.writerow(['Libraries Found', len(data['libraries'])])
        writer.writerow(['Files Found', len(data['file_urls'])])
        writer.writerow(['Version Strings Found', len(data['version_strings'])])
        writer.writerow([])

        # Security Analysis
        writer.writerow(['=== SECURITY ANALYSIS ==='])
        writer.writerow(['Security Score', f"{data['security_analysis']['security_score']}%"])
        writer.writerow([])

        if data['security_analysis']['present']:
            writer.writerow(['Present Security Headers'])
            writer.writerow(['Header Name', 'Header Value', 'Description'])
            for header in data['security_analysis']['present']:
                writer.writerow([header['name'], header['value'], header['description']])
            writer.writerow([])

        if data['security_analysis']['missing']:
            writer.writerow(['Missing Security Headers'])
            writer.writerow(['Header Name', 'Description', 'Recommendation'])
            for header in data['security_analysis']['missing']:
                writer.writerow([header['name'], header['description'], header['recommendation']])
            writer.writerow([])

        # Libraries
        if data['libraries']:
            writer.writerow(['=== DETECTED LIBRARIES ==='])
            writer.writerow(['Library Name', 'Current Version', 'Type', 'Source URL', 'Description',
                           'Latest Safe Version', 'Latest Version', 'Manual Entry', 'Has Vulnerability'])

            for lib in data['libraries']:
                # Check vulnerability with the same logic as template
                has_vuln = (lib['latest_safe_version'] and lib['version'] and
                           lib['version'] != lib['latest_safe_version'] and
                           lib['version'] < lib['latest_safe_version'])

                writer.writerow([
                    lib['library_name'],
                    lib['version'] or 'Unknown',
                    lib['type'].upper(),
                    lib['source_url'] or '',
                    lib['description'] or '',
                    lib['latest_safe_version'] or '',
                    lib['latest_version'] or '',
                    'Yes' if lib['is_manual'] else 'No',
                    'YES - VULNERABLE' if has_vuln else 'No'
                ])
            writer.writerow([])

        # File URLs
        if data['file_urls']:
            writer.writerow(['=== JS/CSS FILES FOUND ==='])
            writer.writerow(['File URL', 'Type', 'File Size (bytes)', 'Status Code'])

            for file_url in data['file_urls']:
                writer.writerow([
                    file_url['file_url'],
                    file_url['file_type'].upper(),
                    file_url['file_size'] or 'Unknown',
                    file_url['status_code'] or 'Error'
                ])
            writer.writerow([])

        # Version Strings
        if data['version_strings']:
            writer.writerow(['=== VERSION STRINGS IN FILES ==='])
            writer.writerow(['File URL', 'File Type', 'Line Number', 'Version Keyword', 'Line Content'])

            for vs in data['version_strings']:
                writer.writerow([
                    vs['file_url'],
                    vs['file_type'].upper(),
                    vs['line_number'],
                    vs['version_keyword'],
                    vs['line_content']
                ])

        # HTTP Headers
        if data['headers']:
            writer.writerow([])
            writer.writerow(['=== HTTP RESPONSE HEADERS ==='])
            writer.writerow(['Header Name', 'Header Value'])

            for header_name, header_value in data['headers'].items():
                writer.writerow([header_name, header_value])

        output.seek(0)
        filename = f"scan_report_{scan_id}_{get_chile_time().strftime('%Y%m%d_%H%M%S')}.csv"

        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'text/csv'

        return response

    except Exception as e:
        flash(f'Error al generar CSV: {str(e)}', 'error')
        return redirect(url_for('scan_detail', scan_id=scan_id))

@app.route('/export/excel/<int:scan_id>')
@login_required
@rate_limit('export')
def export_excel(scan_id):
    try:
        data = get_scan_export_data(scan_id)
        if not data:
            flash('Escaneo no encontrado', 'error')
            return redirect(url_for('index'))

        # Create Excel workbook in memory
        output = io.BytesIO()
        workbook = openpyxl.Workbook()

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_font = Font(bold=True, size=14)

        # Remove default sheet and create custom sheets
        workbook.remove(workbook.active)

        # Sheet 1: Scan Overview
        overview_sheet = workbook.create_sheet("Scan Overview")
        overview_sheet.append(["Scan Information Report"])
        overview_sheet["A1"].font = Font(bold=True, size=16)
        overview_sheet.append([])

        overview_data = [
            ["URL", data['scan']['url']],
            ["Title", data['scan']['title'] or 'No title'],
            ["Status Code", data['scan']['status_code'] or 'Error'],
            ["Scan Date", data['scan']['scan_date']],
            ["Libraries Found", len(data['libraries'])],
            ["Files Found", len(data['file_urls'])],
            ["Version Strings Found", len(data['version_strings'])],
            ["Security Score", f"{data['security_analysis']['security_score']}%"]
        ]

        for row in overview_data:
            overview_sheet.append(row)

        # Format overview sheet
        for cell in overview_sheet['A']:
            if cell.value and cell.row > 2:
                cell.font = Font(bold=True)

        # Sheet 2: Libraries
        if data['libraries']:
            lib_sheet = workbook.create_sheet("Libraries")
            lib_sheet.append(["Detected Libraries"])
            lib_sheet["A1"].font = section_font
            lib_sheet.append([])

            headers = ["Library Name", "Current Version", "Type", "Source URL", "Description",
                      "Latest Safe Version", "Latest Version", "Manual Entry", "Has Vulnerability"]
            lib_sheet.append(headers)

            # Format headers
            for col, header in enumerate(headers, 1):
                cell = lib_sheet.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill

            for lib in data['libraries']:
                # Check vulnerability with the same logic as template (updated with 3 conditions)
                has_vuln = "YES - VULNERABLE" if (lib['latest_safe_version'] and lib['version'] and
                                                 lib['version'] != lib['latest_safe_version'] and
                                                 lib['version'] < lib['latest_safe_version']) else "No"

                lib_sheet.append([
                    lib['library_name'],
                    lib['version'] or 'Unknown',
                    lib['type'].upper(),
                    lib['source_url'] or '',
                    lib['description'] or '',
                    lib['latest_safe_version'] or '',
                    lib['latest_version'] or '',
                    'Yes' if lib['is_manual'] else 'No',
                    has_vuln
                ])

            # Auto-adjust column widths
            for column in lib_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                lib_sheet.column_dimensions[column_letter].width = adjusted_width

        # Sheet 3: Security Analysis
        security_sheet = workbook.create_sheet("Security Analysis")
        security_sheet.append(["Security Analysis Report"])
        security_sheet["A1"].font = section_font
        security_sheet.append([])
        security_sheet.append(["Security Score", f"{data['security_analysis']['security_score']}%"])
        security_sheet.append([])

        if data['security_analysis']['present']:
            security_sheet.append(["Present Security Headers"])
            security_sheet[f"A{security_sheet.max_row}"].font = Font(bold=True)
            headers = ["Header Name", "Header Value", "Description"]
            security_sheet.append(headers)

            for col, header in enumerate(headers, 1):
                cell = security_sheet.cell(row=security_sheet.max_row, column=col)
                cell.font = header_font
                cell.fill = header_fill

            for header in data['security_analysis']['present']:
                security_sheet.append([header['name'], header['value'], header['description']])

            security_sheet.append([])

        if data['security_analysis']['missing']:
            security_sheet.append(["Missing Security Headers"])
            security_sheet[f"A{security_sheet.max_row}"].font = Font(bold=True)
            headers = ["Header Name", "Description", "Recommendation"]
            security_sheet.append(headers)

            for col, header in enumerate(headers, 1):
                cell = security_sheet.cell(row=security_sheet.max_row, column=col)
                cell.font = header_font
                cell.fill = header_fill

            for header in data['security_analysis']['missing']:
                security_sheet.append([header['name'], header['description'], header['recommendation']])

        # Sheet 4: File URLs
        if data['file_urls']:
            files_sheet = workbook.create_sheet("JS CSS Files")
            files_sheet.append(["JavaScript & CSS Files Found"])
            files_sheet["A1"].font = section_font
            files_sheet.append([])

            headers = ["File URL", "Type", "File Size (bytes)", "Status Code"]
            files_sheet.append(headers)

            for col, header in enumerate(headers, 1):
                cell = files_sheet.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill

            for file_url in data['file_urls']:
                files_sheet.append([
                    file_url['file_url'],
                    file_url['file_type'].upper(),
                    file_url['file_size'] or 'Unknown',
                    file_url['status_code'] or 'Error'
                ])

            # Auto-adjust column widths
            for column in files_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)
                files_sheet.column_dimensions[column_letter].width = adjusted_width

        # Sheet 5: Version Strings
        if data['version_strings']:
            vs_sheet = workbook.create_sheet("Version Strings")
            vs_sheet.append(["Version Strings in JS/CSS Files"])
            vs_sheet["A1"].font = section_font
            vs_sheet.append([])

            headers = ["File URL", "File Type", "Line Number", "Version Keyword", "Line Content"]
            vs_sheet.append(headers)

            for col, header in enumerate(headers, 1):
                cell = vs_sheet.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill

            for vs in data['version_strings']:
                vs_sheet.append([
                    vs['file_url'],
                    vs['file_type'].upper(),
                    vs['line_number'],
                    vs['version_keyword'],
                    vs['line_content']
                ])

            # Auto-adjust column widths
            for column in vs_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 100)
                vs_sheet.column_dimensions[column_letter].width = adjusted_width

        # Sheet 6: HTTP Headers
        if data['headers']:
            headers_sheet = workbook.create_sheet("HTTP Headers")
            headers_sheet.append(["HTTP Response Headers"])
            headers_sheet["A1"].font = section_font
            headers_sheet.append([])

            headers = ["Header Name", "Header Value"]
            headers_sheet.append(headers)

            for col, header in enumerate(headers, 1):
                cell = headers_sheet.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill

            for header_name, header_value in data['headers'].items():
                headers_sheet.append([header_name, header_value])

            # Auto-adjust column widths
            for column in headers_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)
                headers_sheet.column_dimensions[column_letter].width = adjusted_width

        # Save workbook to buffer
        workbook.save(output)
        output.seek(0)

        filename = f"scan_report_{scan_id}_{get_chile_time().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        flash(f'Error al generar Excel: {str(e)}', 'error')
        return redirect(url_for('scan_detail', scan_id=scan_id))

# Import analyzer functionality
def detect_js_libraries(soup, base_url):
    libraries = []

    # jQuery detection
    jquery_scripts = soup.find_all('script', src=re.compile(r'jquery', re.I))
    for script in jquery_scripts:
        src = script.get('src', '')
        version_match = re.search(r'jquery[-.]?(\d+\.\d+\.\d+)', src, re.I)
        if version_match:
            libraries.append({
                'name': 'jQuery',
                'version': version_match.group(1),
                'type': 'js',
                'source': urljoin(base_url, src)
            })

    # React detection
    react_scripts = soup.find_all('script', src=re.compile(r'react', re.I))
    for script in react_scripts:
        src = script.get('src', '')
        version_match = re.search(r'react[-.]?(\d+\.\d+\.\d+)', src, re.I)
        if version_match:
            libraries.append({
                'name': 'React',
                'version': version_match.group(1),
                'type': 'js',
                'source': urljoin(base_url, src)
            })

    # Vue.js detection
    vue_scripts = soup.find_all('script', src=re.compile(r'vue', re.I))
    for script in vue_scripts:
        src = script.get('src', '')
        version_match = re.search(r'vue[-.]?(\d+\.\d+\.\d+)', src, re.I)
        if version_match:
            libraries.append({
                'name': 'Vue.js',
                'version': version_match.group(1),
                'type': 'js',
                'source': urljoin(base_url, src)
            })

    # Bootstrap JS detection
    bootstrap_scripts = soup.find_all('script', src=re.compile(r'bootstrap', re.I))
    for script in bootstrap_scripts:
        src = script.get('src', '')
        version_match = re.search(r'bootstrap[-.]?(\d+\.\d+\.\d+)', src, re.I)
        if version_match:
            libraries.append({
                'name': 'Bootstrap JS',
                'version': version_match.group(1),
                'type': 'js',
                'source': urljoin(base_url, src)
            })

    # Angular detection
    angular_scripts = soup.find_all('script', src=re.compile(r'angular', re.I))
    for script in angular_scripts:
        src = script.get('src', '')
        version_match = re.search(r'angular[-.]?(\d+\.\d+\.\d+)', src, re.I)
        if version_match:
            libraries.append({
                'name': 'Angular',
                'version': version_match.group(1),
                'type': 'js',
                'source': urljoin(base_url, src)
            })

    return libraries

def detect_css_libraries(soup, base_url):
    libraries = []

    # Bootstrap CSS detection
    bootstrap_links = soup.find_all('link', href=re.compile(r'bootstrap', re.I))
    for link in bootstrap_links:
        href = link.get('href', '')
        version_match = re.search(r'bootstrap[-.]?(\d+\.\d+\.\d+)', href, re.I)
        if version_match:
            libraries.append({
                'name': 'Bootstrap CSS',
                'version': version_match.group(1),
                'type': 'css',
                'source': urljoin(base_url, href)
            })

    # Font Awesome detection
    fa_links = soup.find_all('link', href=re.compile(r'font-?awesome', re.I))
    for link in fa_links:
        href = link.get('href', '')
        version_match = re.search(r'font-?awesome[-.]?(\d+\.\d+\.\d+)', href, re.I)
        if version_match:
            libraries.append({
                'name': 'Font Awesome',
                'version': version_match.group(1),
                'type': 'css',
                'source': urljoin(base_url, href)
            })

    return libraries

def scan_file_for_versions(file_url, file_type, scan_id):
    """
    Enhanced version scanning with multiple patterns and automatic library detection
    """
    version_strings = []
    detected_libraries = []

    # Diccionarios para evitar duplicados por URL fuente
    # Solo mantenemos la PRIMERA biblioteca y cadena de versión detectada por archivo
    first_library_per_source = {}
    first_version_string_per_source = {}

    # Patrones de versión solicitados por el usuario
    VERSION_PATTERNS = [
        # Patrones v/V con números
        (r'\bv\.?\s*(\d+(?:\.\d+)*)\b', 'v_pattern'),
        (r'\bV\.?\s*(\d+(?:\.\d+)*)\b', 'V_pattern'),

        # Versiones con formato x.x.x
        (r'\b(\d+\.\d+\.\d+)\b', 'semver'),
        (r'["\'](\d+\.\d+\.\d+)["\']', 'quoted_semver'),
        (r'\s(\d+\.\d+\.\d+)\s', 'spaced_semver'),

        # Patrones version con =
        (r'\bversion\s*=\s*["\']?(\d+\.\d+\.\d+)["\']?', 'version_equals'),
        (r'\bVersion\s*=\s*["\']?(\d+\.\d+\.\d+)["\']?', 'Version_equals'),

        # Patrones version con :
        (r'\bversion\s*:\s*["\']?(\d+\.\d+\.\d+)["\']?', 'version_colon'),
        (r'\bVersion\s*:\s*["\']?(\d+\.\d+\.\d+)["\']?', 'Version_colon'),

        # Patrones adicionales útiles
        (r'/\*.*?v\.?\s*(\d+\.\d+\.\d+).*?\*/', 'comment_version'),
        (r'//.*?v\.?\s*(\d+\.\d+\.\d+)', 'line_comment_version'),
        (r'\brelease[:\s]+["\']?(\d+\.\d+\.\d+)["\']?', 'release'),
        (r'\bbuild[:\s]+["\']?(\d+\.\d+\.\d+)["\']?', 'build'),
        (r'@version\s+(\d+\.\d+\.\d+)', 'jsdoc_version'),
        (r'-(\d+\.\d+\.\d+)\.(?:min\.)?(?:js|css)', 'filename_version'),
        (r'["\']version["\']\s*:\s*["\'](\d+\.\d+\.\d+)["\']', 'json_version'),
    ]

    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

        response = requests.get(file_url, headers=headers, timeout=10)
        if response.status_code == 200:
            content = response.text
            lines = content.split('\n')

            for line_num, line in enumerate(lines, 1):
                # SOLO LA PRIMERA CADENA DE VERSIÓN POR ARCHIVO
                if file_url not in first_version_string_per_source:

                    # MANTENER funcionalidad existente para compatibilidad
                    if re.search(r'version', line, re.I):
                        first_version_string_per_source[file_url] = {
                            'scan_id': scan_id,
                            'file_url': file_url,
                            'file_type': file_type,
                            'line_number': line_num,
                            'line_content': line.strip()[:200],
                            'version_keyword': 'version'
                        }
                        continue

                    if re.search(r'versión', line, re.I):
                        first_version_string_per_source[file_url] = {
                            'scan_id': scan_id,
                            'file_url': file_url,
                            'file_type': file_type,
                            'line_number': line_num,
                            'line_content': line.strip()[:200],
                            'version_keyword': 'versión'
                        }
                        continue

                    # NUEVOS PATRONES: Buscar versiones específicas
                    for pattern, pattern_type in VERSION_PATTERNS:
                        matches = re.finditer(pattern, line, re.I)
                        for match in matches:
                            version_number = match.group(1)

                            # Agregar PRIMERA cadena de versión por archivo
                            first_version_string_per_source[file_url] = {
                                'scan_id': scan_id,
                                'file_url': file_url,
                                'file_type': file_type,
                                'line_number': line_num,
                                'line_content': line.strip()[:200],
                                'version_keyword': pattern_type
                            }
                            break  # Solo el primer match por patrón

                        # Si ya encontramos una cadena de versión, salir del bucle de patrones
                        if file_url in first_version_string_per_source:
                            break

                # Detectar biblioteca automáticamente - SOLO LA PRIMERA POR URL
                if file_url not in first_library_per_source:
                    for pattern, pattern_type in VERSION_PATTERNS:
                        matches = re.finditer(pattern, line, re.I)
                        for match in matches:
                            version_number = match.group(1)
                            library_name = extract_library_name_from_context(line, file_url, version_number)
                            if library_name:
                                first_library_per_source[file_url] = {
                                    'name': library_name,
                                    'version': version_number,
                                    'type': file_type,
                                    'source': file_url,
                                    'detection_method': 'version_pattern',
                                    'confidence': 0.6
                                }
                                break
                        if file_url in first_library_per_source:
                            break

    except Exception as e:
        print(f"  ✗ Error scanning {file_url}: {str(e)}")

    # 🚀 FASE 2: DETECCIÓN AVANZADA POR CONTENIDO
    if CONTENT_DETECTION_AVAILABLE and 'content' in locals() and content and file_url not in first_library_per_source:
        try:
            content_detections = detect_libraries_by_content(content, file_type)
            if content_detections:
                # Tomar la detección con mayor confianza
                best_detection = max(content_detections, key=lambda x: x['confidence'])
                
                first_library_per_source[file_url] = {
                    'name': best_detection['library_name'].title(),
                    'version': best_detection.get('version', 'unknown'),
                    'type': file_type,
                    'source': file_url,
                    'detection_method': 'content_analysis',
                    'confidence': best_detection.get('confidence', 0.8),
                    'analysis_details': best_detection.get('details', []),
                    'matches': best_detection.get('matches', 0)
                }
                
                # Agregar versión string si se detectó versión
                version = best_detection.get('version', 'unknown')
                if version != 'unknown' and file_url not in first_version_string_per_source:
                    first_version_string_per_source[file_url] = {
                        'scan_id': scan_id,
                        'file_url': file_url,
                        'file_type': file_type,
                        'line_number': 1,  # Línea estimada
                        'line_content': f'Library detected: {best_detection["library_name"]} v{version}',
                        'version_keyword': f'{best_detection["library_name"]}_content_analysis'
                    }
                
                print(f"  🎯 Content analysis detected: {best_detection['library_name']} v{version} (confidence: {best_detection['confidence']:.1f}, matches: {best_detection.get('matches', 0)})")
                
        except Exception as e:
            print(f"  ⚠️ Error in content analysis for {file_url}: {str(e)}")

    # Convertir diccionarios a listas - solo UNA entrada por archivo fuente
    version_strings = list(first_version_string_per_source.values())
    detected_libraries = list(first_library_per_source.values())
    return version_strings, detected_libraries




def extract_library_name_from_context(line, file_url, version):
    """
    Intenta extraer el nombre de la biblioteca del contexto
    """
    line_lower = line.lower()
    url_lower = file_url.lower()

    # Bibliotecas conocidas en comentarios/líneas
    LIBRARY_PATTERNS = [
        ('jquery', 'jQuery'),
        ('bootstrap', 'Bootstrap'),
        ('react', 'React'),
        ('vue', 'Vue.js'),
        ('angular', 'Angular'),
        ('lodash', 'Lodash'),
        ('moment', 'Moment.js'),
        ('chart', 'Chart.js'),
        ('d3', 'D3.js'),
        ('three', 'Three.js'),
        ('axios', 'Axios'),
        ('underscore', 'Underscore.js'),
        ('backbone', 'Backbone.js'),
        ('ember', 'Ember.js'),
        ('knockout', 'Knockout.js'),
        ('handlebars', 'Handlebars.js'),
        ('mustache', 'Mustache.js'),
        ('font-awesome', 'Font Awesome'),
        ('fontawesome', 'Font Awesome'),
        ('material', 'Material UI'),
        ('semantic', 'Semantic UI'),
        ('foundation', 'Foundation'),
        ('bulma', 'Bulma'),
        ('tailwind', 'Tailwind CSS'),
    ]

    # Buscar en la línea
    for pattern, name in LIBRARY_PATTERNS:
        if pattern in line_lower or pattern in url_lower:
            return name

    # Buscar en la URL del archivo
    filename = file_url.split('/')[-1].lower()
    for pattern, name in LIBRARY_PATTERNS:
        if pattern in filename:
            return name

    # Si no se encuentra, retornar nombre genérico
    return f"Biblioteca desconocida ({filename.split('.')[0]})" if '.' in filename else "Biblioteca desconocida"

def get_all_js_css_files(soup, base_url):
    files = []

    # Get all script files
    scripts = soup.find_all('script', src=True)
    for script in scripts:
        src = script.get('src')
        if src:
            full_url = urljoin(base_url, src)
            files.append({'url': full_url, 'type': 'js'})

    # Get all CSS files
    links = soup.find_all('link', {'rel': 'stylesheet', 'href': True})
    for link in links:
        href = link.get('href')
        if href:
            full_url = urljoin(base_url, href)
            files.append({'url': full_url, 'type': 'css'})

    return files

def store_file_urls_with_info(files, scan_id, cursor):
    """Store file URLs using an existing cursor/connection"""
    for file_info in files:
        file_url = file_info['url']
        file_type = file_info['type']

        # Try to get file information
        file_size = None
        status_code = None

        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/36'
            }

            # Make a HEAD request to get file info without downloading content
            response = requests.head(file_url, headers=headers, timeout=3, allow_redirects=True)
            status_code = response.status_code

            # Get file size from headers if available
            content_length = response.headers.get('content-length')
            if content_length:
                file_size = int(content_length)

        except Exception as e:
            print(f"  ! Could not get info for {file_url}: {str(e)}")
            status_code = 0

        # Store file URL information
        cursor.execute('''
        INSERT INTO file_urls (scan_id, file_url, file_type, file_size, status_code)
        VALUES (?, ?, ?, ?, ?)
        ''', (scan_id, file_url, file_type, file_size, status_code))

def is_safe_url(url):
    """
    Validate URL to prevent SSRF attacks
    Blocks private IP ranges, localhost, and other potentially dangerous URLs
    """
    try:
        parsed = urlparse(url)
        hostname = parsed.hostname

        if not hostname:
            return False

        # Block localhost and loopback
        if hostname.lower() in ['localhost', '127.0.0.1', '::1']:
            return False

        # Try to resolve hostname to IP and check if it's private
        try:
            import socket
            ip = socket.gethostbyname(hostname)
            import ipaddress
            ip_obj = ipaddress.ip_address(ip)

            # Block private IP ranges
            if ip_obj.is_private or ip_obj.is_loopback or ip_obj.is_link_local:
                return False

        except (socket.gaierror, ipaddress.AddressValueError, ValueError):
            # If we can't resolve, it might be suspicious
            return False

        # Block non-HTTP(S) schemes
        if parsed.scheme not in ['http', 'https']:
            return False

        return True

    except Exception:
        return False

def analyze_single_url_no_logging(url, project_id=None):
    """Versión optimizada sin logging automático para análisis masivos"""
    conn = None
    try:
        # Validate URL to prevent SSRF attacks
        if not is_safe_url(url):
            raise ValueError("URL is not allowed for security reasons (private IP, localhost, or invalid scheme)")

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

        response = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(response.content, 'html.parser')

        # Get page title
        title_tag = soup.find('title')
        title = title_tag.text.strip() if title_tag else 'No title'

        # Store scan info
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute('''
        INSERT INTO scans (url, status_code, title, headers, project_id)
        VALUES (?, ?, ?, ?, ?)
        ''', (url, response.status_code, title, json.dumps(dict(response.headers)), project_id))

        scan_id = cursor.lastrowid

        # NO logging automático aquí para evitar conflictos en análisis masivos

        # Detect libraries
        js_libraries = detect_js_libraries(soup, url)
        css_libraries = detect_css_libraries(soup, url)

        all_libraries = js_libraries + css_libraries

        # Store libraries (only if source_url is unique)
        for lib in all_libraries:
            source_url = lib.get('source')
            if not library_source_exists(cursor, scan_id, source_url):
                cursor.execute('''
                INSERT INTO libraries (scan_id, library_name, version, type, source_url)
                VALUES (?, ?, ?, ?, ?)
                ''', (scan_id, lib['name'], lib['version'], lib['type'], source_url))
                print(f"  → Stored library: {lib['name']} from {source_url or 'No source'}")
            else:
                print(f"  → Skipped duplicate library: {lib['name']} (source already exists: {source_url})")

        # Get all JS and CSS files
        js_css_files = get_all_js_css_files(soup, url)

        # Store all file URLs with additional info using the same connection
        store_file_urls_with_info(js_css_files, scan_id, cursor)

        # Scan files for version strings and detect libraries
        all_version_strings = []
        all_detected_libraries = []
        for file_info in js_css_files:
            version_strings, detected_libraries = scan_file_for_versions(file_info['url'], file_info['type'], scan_id)
            all_version_strings.extend(version_strings)
            all_detected_libraries.extend(detected_libraries)

        # Store version strings
        for vs in all_version_strings:
            cursor.execute('''
            INSERT INTO version_strings (scan_id, file_url, file_type, line_number, line_content, version_keyword)
            VALUES (?, ?, ?, ?, ?, ?)
            ''', (vs['scan_id'], vs['file_url'], vs['file_type'], vs['line_number'], vs['line_content'], vs['version_keyword']))

        # Store automatically detected libraries (only if source_url is unique)
        for lib in all_detected_libraries:
            source_url = lib.get('source')
            if not library_source_exists(cursor, scan_id, source_url):
                cursor.execute('''
                INSERT INTO libraries (scan_id, library_name, version, type, source_url, description, is_manual)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (scan_id, lib['name'], lib['version'], lib['type'], source_url,
                      f"Detectada automáticamente por patrón de versión ({lib['detection_method']})", 0))
                print(f"  → Stored auto-detected library: {lib['name']} from {source_url or 'No source'}")
            else:
                print(f"  → Skipped duplicate auto-detected library: {lib['name']} (source already exists: {source_url})")

        conn.commit()

        return {
            'success': True,
            'scan_id': scan_id,
            'libraries_count': len(all_libraries) + len(all_detected_libraries),
            'files_count': len(js_css_files),
            'version_strings_count': len(all_version_strings)
        }

    except Exception as e:
        # Store failed scan
        try:
            if not conn:
                conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('''
            INSERT INTO scans (url, status_code, title, headers)
            VALUES (?, ?, ?, ?)
            ''', (url, 0, f"Error: {str(e)}", "{}"))
            scan_id = cursor.lastrowid
            conn.commit()
        except:
            scan_id = None

        return {
            'success': False,
            'error': str(e),
            'scan_id': scan_id
        }
    finally:
        if conn:
            conn.close()

def library_source_exists(cursor, scan_id, source_url):
    """
    Verificar si ya existe una biblioteca con la misma source_url para este scan
    Retorna True si existe, False si no existe
    """
    if not source_url or source_url.strip() == '':
        return False
    
    result = cursor.execute(
        'SELECT id FROM libraries WHERE scan_id = ? AND source_url = ?', 
        (scan_id, source_url.strip())
    ).fetchone()
    return result is not None

def analyze_single_url(url, project_id=None):
    conn = None
    try:
        # Validate URL to prevent SSRF attacks
        if not is_safe_url(url):
            raise ValueError("URL is not allowed for security reasons (private IP, localhost, or invalid scheme)")

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

        response = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(response.content, 'html.parser')

        # Get page title
        title_tag = soup.find('title')
        title = title_tag.text.strip() if title_tag else 'No title'

        # Store scan info
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute('''
        INSERT INTO scans (url, status_code, title, headers, project_id)
        VALUES (?, ?, ?, ?, ?)
        ''', (url, response.status_code, title, json.dumps(dict(response.headers)), project_id))

        scan_id = cursor.lastrowid

        # Log the scan creation con conexión separada para evitar deadlocks
        try:
            log_user_action(
                action_type='CREATE',
                target_table='scans',
                target_id=scan_id,
                target_description=f"Nuevo análisis de URL: {url}",
                success=True,
                notes=f"Status: {response.status_code}, Título: {title[:50]}{'...' if len(title) > 50 else ''}"
            )
        except Exception as log_error:
            # No fallar el escaneo por problemas de logging
            print(f"⚠️ Error en logging (no crítico): {log_error}")
            if LOGGING_DEBUG:
                import traceback
                traceback.print_exc()

        # Detect libraries
        js_libraries = detect_js_libraries(soup, url)
        css_libraries = detect_css_libraries(soup, url)

        all_libraries = js_libraries + css_libraries

        # 🚀 FASE 2: Store libraries with vulnerability and CDN analysis
        cdn_libraries = []
        outdated_cdn_count = 0
        
        for lib in all_libraries:
            source_url = lib.get('source')
            if not library_source_exists(cursor, scan_id, source_url):
                
                
                # 🌐 ANÁLISIS CDN 
                cdn_analysis = None
                if CDN_ANALYZER_AVAILABLE and source_url:
                    cdn_analysis = analyze_cdn_url(source_url)
                    if cdn_analysis:
                        cdn_libraries.append(cdn_analysis)
                        if cdn_analysis.get('is_outdated', False):
                            outdated_cdn_count += 1
                
                cursor.execute('''
                INSERT INTO libraries (scan_id, library_name, version, type, source_url)
                VALUES (?, ?, ?, ?, ?)
                ''', (scan_id, lib['name'], lib['version'], lib['type'], source_url))
                
                # Log con información de CDN
                cdn_indicator = ""
                
                if cdn_analysis:
                    if cdn_analysis.get('is_outdated', False):
                        cdn_indicator = f" 📦 {cdn_analysis.get('cdn_name', 'Unknown')} (OUTDATED)"
                    else:
                        cdn_indicator = f" 📦 {cdn_analysis.get('cdn_name', 'Unknown')}"
                
                print(f"  → Stored library: {lib['name']} v{lib['version']}{cdn_indicator}")
            else:
                print(f"  → Skipped duplicate library: {lib['name']} (source already exists: {source_url})")
        
        # 🌐 RESUMEN ANÁLISIS CDN
        if CDN_ANALYZER_AVAILABLE and cdn_libraries:
            print(f"  🌐 CDN Analysis: {len(cdn_libraries)} libraries from CDN")
            if outdated_cdn_count > 0:
                print(f"    ⚠️ {outdated_cdn_count} outdated CDN libraries detected")

        # Get all JS and CSS files
        js_css_files = get_all_js_css_files(soup, url)

        # Store all file URLs with additional info using the same connection
        store_file_urls_with_info(js_css_files, scan_id, cursor)

        # Scan files for version strings and detect libraries
        all_version_strings = []
        all_detected_libraries = []
        for file_info in js_css_files:
            version_strings, detected_libraries = scan_file_for_versions(file_info['url'], file_info['type'], scan_id)
            all_version_strings.extend(version_strings)
            all_detected_libraries.extend(detected_libraries)

        # Store version strings
        for vs in all_version_strings:
            cursor.execute('''
            INSERT INTO version_strings (scan_id, file_url, file_type, line_number, line_content, version_keyword)
            VALUES (?, ?, ?, ?, ?, ?)
            ''', (vs['scan_id'], vs['file_url'], vs['file_type'], vs['line_number'], vs['line_content'], vs['version_keyword']))

        # Store automatically detected libraries (only if source_url is unique)
        for lib in all_detected_libraries:
            source_url = lib.get('source')
            if not library_source_exists(cursor, scan_id, source_url):
                cursor.execute('''
                INSERT INTO libraries (scan_id, library_name, version, type, source_url, description, is_manual)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (scan_id, lib['name'], lib['version'], lib['type'], source_url,
                      f"Detectada automáticamente por patrón de versión ({lib['detection_method']})", 0))
                print(f"  → Stored auto-detected library: {lib['name']} from {source_url or 'No source'}")
            else:
                print(f"  → Skipped duplicate auto-detected library: {lib['name']} (source already exists: {source_url})")

        conn.commit()

        return {
            'success': True,
            'scan_id': scan_id,
            'libraries_count': len(all_libraries) + len(all_detected_libraries),
            'files_count': len(js_css_files),
            'version_strings_count': len(all_version_strings)
        }

    except Exception as e:
        # Store failed scan
        try:
            if not conn:
                conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('''
            INSERT INTO scans (url, status_code, title, headers)
            VALUES (?, ?, ?, ?)
            ''', (url, 0, f"Error: {str(e)}", "{}"))
            scan_id = cursor.lastrowid
            conn.commit()
        except:
            scan_id = None

        return {
            'success': False,
            'error': str(e),
            'scan_id': scan_id
        }
    finally:
        if conn:
            conn.close()

@app.route('/analyze-url', methods=['POST'])
@login_required
@rate_limit('analysis', 'Límite de análisis alcanzado. Espera un momento antes de analizar otra URL.')
def analyze_url_route():
    url = request.form.get('url', '').strip()
    project_id = request.form.get('project_id', '').strip()

    # Convert empty string to None for project_id
    if not project_id:
        project_id = None

    if not url:
        flash('Por favor ingresa una URL válida', 'error')
        return redirect(url_for('index'))

    # Add protocol if missing
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url

    try:
        result = analyze_single_url(url, project_id=project_id)

        if result['success']:
            flash(f'¡Análisis completado! Se encontraron {result["libraries_count"]} librerías, {result["files_count"]} archivos, y {result["version_strings_count"]} cadenas de versión.', 'success')
            return redirect(url_for('scan_detail', scan_id=result['scan_id']))
        else:
            flash(f'Análisis fallido: {result["error"]}', 'error')
            return redirect(url_for('scan_detail', scan_id=result['scan_id']))

    except Exception as e:
        flash(f'Error inesperado: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/batch-analyze', methods=['POST'])
@login_required
@rate_limit('batch_analysis', 'Límite de análisis en lotes alcanzado. Espera un momento.')
def batch_analyze_route():
    urls_text = request.form.get('urls', '').strip()
    project_id = request.form.get('project_id', '').strip()

    # Convert empty string to None for project_id
    if not project_id:
        project_id = None

    if not urls_text:
        flash('Por favor ingresa al menos una URL', 'error')
        return redirect(url_for('index'))

    # Parse URLs from text area
    urls = []
    for line in urls_text.split('\n'):
        url = line.strip()
        if url and not url.startswith('#'):  # Skip empty lines and comments
            # Add protocol if missing
            if not url.startswith(('http://', 'https://')):
                url = 'https://' + url
            urls.append(url)

    if not urls:
        flash('No se encontraron URLs válidas', 'error')
        return redirect(url_for('index'))

    # Perform batch analysis
    results = {
        'total_urls': len(urls),
        'successful': 0,
        'failed': 0,
        'total_libraries': 0,
        'total_files': 0,
        'total_version_strings': 0,
        'scan_ids': []
    }

    # Colectar acciones de logging para procesamiento en lotes
    batch_actions = []

    try:
        for i, url in enumerate(urls, 1):
            print(f"[{i}/{len(urls)}] Analyzing: {url}")

            # Usar análisis sin logging automático para evitar conflictos de base de datos
            result = analyze_single_url_no_logging(url, project_id=project_id)

            if result['success']:
                results['successful'] += 1
                results['total_libraries'] += result['libraries_count']
                results['total_files'] += result['files_count']
                results['total_version_strings'] += result['version_strings_count']
                results['scan_ids'].append(result['scan_id'])
                print(f"  ✓ Success: {result['libraries_count']} libs, {result['files_count']} files, {result['version_strings_count']} versions")

                # Preparar acción de logging para lote
                batch_actions.append({
                    'user_id': session.get('user_id', 0),
                    'username': session.get('username', 'Sistema'),
                    'user_role': session.get('user_role', 'system'),
                    'action_type': 'CREATE',
                    'target_table': 'scans',
                    'target_id': result['scan_id'],
                    'target_description': f"Análisis masivo: {url}",
                    'success': True,
                    'ip_address': request.remote_addr,
                    'user_agent': request.headers.get('User-Agent'),
                    'session_id': session.get('session_id'),
                    'notes': f"Análisis masivo ({i}/{len(urls)}) - Libs: {result['libraries_count']}, Files: {result['files_count']}"
                })
            else:
                results['failed'] += 1
                if result['scan_id']:
                    results['scan_ids'].append(result['scan_id'])
                print(f"  ✗ Failed: {result['error']}")

                # Preparar acción de logging de error para lote
                batch_actions.append({
                    'user_id': session.get('user_id', 0),
                    'username': session.get('username', 'Sistema'),
                    'user_role': session.get('user_role', 'system'),
                    'action_type': 'CREATE',
                    'target_table': 'scans',
                    'target_id': result.get('scan_id'),
                    'target_description': f"Análisis masivo fallido: {url}",
                    'success': False,
                    'error_message': result['error'],
                    'ip_address': request.remote_addr,
                    'user_agent': request.headers.get('User-Agent'),
                    'session_id': session.get('session_id'),
                    'notes': f"Análisis masivo ({i}/{len(urls)}) - Error"
                })

            # Small delay to prevent overwhelming servers
            time.sleep(0.5)

        # Registrar todas las acciones en lote al final
        if batch_actions:
            print(f"📝 Registrando {len(batch_actions)} acciones en lote...")
            log_batch_actions(batch_actions)

        # Create summary message
        if results['successful'] > 0:
            flash(f'¡Análisis masivo completado! Se analizaron exitosamente {results["successful"]}/{results["total_urls"]} URLs. ' +
                  f'Se encontraron {results["total_libraries"]} librerías, {results["total_files"]} archivos, ' +
                  f'y {results["total_version_strings"]} cadenas de versión.', 'success')
        else:
            flash(f'Análisis masivo completado con errores. {results["failed"]} URLs fallaron al analizar.', 'error')

    except Exception as e:
        flash(f'Análisis masivo fallido: {str(e)}', 'error')

    return redirect(url_for('index'))


# Global Libraries Management Routes
@app.route('/global-libraries')
@login_required
def global_libraries():
    conn = get_db_connection()

    # Get search parameter
    search_query = request.args.get('search', '').strip()

    # Build query with search filter including manual libraries count
    if search_query:
        libraries = conn.execute('''
            SELECT gl.id, gl.library_name, gl.type, gl.latest_safe_version, gl.latest_version,
                   gl.description, gl.vulnerability_info, gl.source_url, gl.created_date, gl.updated_date,
                   COUNT(l.id) as manual_count
            FROM global_libraries gl
            LEFT JOIN libraries l ON gl.id = l.global_library_id AND l.is_manual = 1
            WHERE gl.library_name LIKE ? OR gl.description LIKE ?
            GROUP BY gl.id, gl.library_name, gl.type, gl.latest_safe_version, gl.latest_version,
                     gl.description, gl.vulnerability_info, gl.source_url, gl.created_date, gl.updated_date
            ORDER BY gl.library_name
        ''', (f'%{search_query}%', f'%{search_query}%')).fetchall()
    else:
        libraries = conn.execute('''
            SELECT gl.id, gl.library_name, gl.type, gl.latest_safe_version, gl.latest_version,
                   gl.description, gl.vulnerability_info, gl.source_url, gl.created_date, gl.updated_date,
                   COUNT(l.id) as manual_count
            FROM global_libraries gl
            LEFT JOIN libraries l ON gl.id = l.global_library_id AND l.is_manual = 1
            GROUP BY gl.id, gl.library_name, gl.type, gl.latest_safe_version, gl.latest_version,
                     gl.description, gl.vulnerability_info, gl.source_url, gl.created_date, gl.updated_date
            ORDER BY gl.library_name
        ''').fetchall()

    # Get top libraries with version information from actual scans
    top_libraries = conn.execute('''
        WITH library_versions AS (
            SELECT
                library_name,
                type,
                version,
                COUNT(*) as version_count
            FROM libraries
            WHERE version IS NOT NULL AND version != ''
            GROUP BY library_name, type, version
        ),
        most_common_version AS (
            SELECT
                library_name,
                type,
                version as most_common_version,
                ROW_NUMBER() OVER (PARTITION BY library_name, type ORDER BY version_count DESC) as rn
            FROM library_versions
        ),
        library_totals AS (
            SELECT
                library_name,
                type,
                COUNT(*) as usage_count
            FROM libraries
            GROUP BY library_name, type
        )
        SELECT
            lt.library_name,
            lt.type,
            lt.usage_count,
            COALESCE(mcv.most_common_version, 'Unknown') as most_common_version,
            COUNT(DISTINCT l.version) as version_variety
        FROM library_totals lt
        LEFT JOIN most_common_version mcv ON lt.library_name = mcv.library_name
            AND lt.type = mcv.type AND mcv.rn = 1
        LEFT JOIN libraries l ON lt.library_name = l.library_name
            AND lt.type = l.type
        GROUP BY lt.library_name, lt.type, lt.usage_count, mcv.most_common_version
        ORDER BY lt.usage_count DESC
        LIMIT 15
    ''').fetchall()

    conn.close()
    return render_template('global_libraries.html', libraries=libraries, top_libraries=top_libraries)

@app.route('/api/global-libraries')
@login_required
def api_global_libraries():
    conn = get_db_connection()
    libraries = conn.execute('''
        SELECT id, library_name, type, latest_safe_version, latest_version,
               description, vulnerability_info, source_url
        FROM global_libraries
        ORDER BY library_name
    ''').fetchall()
    conn.close()
    return jsonify([dict(lib) for lib in libraries])

@app.route('/add-global-library', methods=['POST'])
@login_required
def add_global_library():
    try:
        library_name = request.form.get('library_name', '').strip()
        library_type = request.form.get('library_type')
        latest_safe_version = request.form.get('latest_safe_version', '').strip()
        latest_version = request.form.get('latest_version', '').strip()
        description = request.form.get('description', '').strip()
        vulnerability_info = request.form.get('vulnerability_info', '').strip()
        source_url = request.form.get('source_url', '').strip()

        if not library_name:
            flash('El nombre de la librería es requerido', 'error')
            return redirect(url_for('global_libraries'))

        if not library_type or library_type not in ['js', 'css']:
            flash('Se requiere un tipo de librería válido (js o css)', 'error')
            return redirect(url_for('global_libraries'))

        conn = get_db_connection()
        cursor = conn.cursor()

        # Check if library already exists
        cursor.execute("SELECT id FROM global_libraries WHERE library_name = ?", (library_name,))
        if cursor.fetchone():
            flash(f'La librería "{library_name}" ya existe en el catálogo', 'error')
            conn.close()
            return redirect(url_for('global_libraries'))

        cursor.execute('''
            INSERT INTO global_libraries
            (library_name, type, latest_safe_version, latest_version, description, vulnerability_info, source_url)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (library_name, library_type, latest_safe_version, latest_version,
              description, vulnerability_info, source_url))

        conn.commit()
        conn.close()
        flash(f'Librería "{library_name}" agregada exitosamente al catálogo', 'success')

    except Exception as e:
        flash(f'Error al agregar librería: {str(e)}', 'error')

    return redirect(url_for('global_libraries'))

@app.route('/edit-global-library/<int:library_id>', methods=['POST'])
@login_required
def edit_global_library(library_id):
    try:
        library_name = request.form.get('library_name', '').strip()
        library_type = request.form.get('library_type')
        latest_safe_version = request.form.get('latest_safe_version', '').strip()
        latest_version = request.form.get('latest_version', '').strip()
        description = request.form.get('description', '').strip()
        vulnerability_info = request.form.get('vulnerability_info', '').strip()
        source_url = request.form.get('source_url', '').strip()

        if not library_name:
            flash('El nombre de la librería es requerido', 'error')
            return redirect(url_for('global_libraries'))

        if not library_type or library_type not in ['js', 'css']:
            flash('Se requiere un tipo de librería válido (js o css)', 'error')
            return redirect(url_for('global_libraries'))

        conn = get_db_connection()
        cursor = conn.cursor()

        # Check if another library with same name exists
        cursor.execute("SELECT id FROM global_libraries WHERE library_name = ? AND id != ?",
                      (library_name, library_id))
        if cursor.fetchone():
            flash(f'Ya existe otra librería con el nombre "{library_name}"', 'error')
            conn.close()
            return redirect(url_for('global_libraries'))

        cursor.execute('''
            UPDATE global_libraries
            SET library_name = ?, type = ?, latest_safe_version = ?, latest_version = ?,
                description = ?, vulnerability_info = ?, source_url = ?,
                updated_date = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (library_name, library_type, latest_safe_version, latest_version,
              description, vulnerability_info, source_url, library_id))

        if cursor.rowcount == 0:
            flash('Librería no encontrada', 'error')
        else:
            flash(f'Librería "{library_name}" actualizada exitosamente', 'success')

        conn.commit()
        conn.close()

    except Exception as e:
        flash(f'Error al actualizar librería: {str(e)}', 'error')

    return redirect(url_for('global_libraries'))

@app.route('/delete-global-library/<int:library_id>', methods=['POST'])
@login_required
def delete_global_library(library_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Get library name before deletion for flash message
        cursor.execute("SELECT library_name FROM global_libraries WHERE id = ?", (library_id,))
        library = cursor.fetchone()

        if not library:
            flash('Librería no encontrada', 'error')
            conn.close()
            return redirect(url_for('global_libraries'))

        cursor.execute("DELETE FROM global_libraries WHERE id = ?", (library_id,))
        conn.commit()
        conn.close()

        flash(f'Librería "{library["library_name"]}" eliminada exitosamente', 'success')

    except Exception as e:
        flash(f'Error al eliminar librería: {str(e)}', 'error')

    return redirect(url_for('global_libraries'))

@app.route('/asociar-bibliotecas')
@login_required
def asociar_bibliotecas():
    conn = get_db_connection()
    unassociated_libs = conn.execute('''
        SELECT l.id, l.library_name, l.version, l.scan_id, s.url
        FROM libraries l
        JOIN scans s ON l.scan_id = s.id
        WHERE l.is_manual = 1 AND l.global_library_id IS NULL
        ORDER BY s.scan_date DESC, l.library_name
    ''').fetchall()

    global_libraries = conn.execute('SELECT id, library_name, type FROM global_libraries ORDER BY library_name').fetchall()
    conn.close()

    return render_template('asociar_bibliotecas.html',
                           unassociated_libs=unassociated_libs,
                           global_libraries=global_libraries)

@app.route('/global-library/<int:global_lib_id>/manual-libraries')
@login_required
def global_library_manual_libraries(global_lib_id):
    """Show all manual libraries associated with a specific global library"""
    conn = get_db_connection()

    # Get global library info
    global_lib = conn.execute('''
        SELECT id, library_name, type, latest_safe_version, latest_version,
               description, vulnerability_info
        FROM global_libraries
        WHERE id = ?
    ''', (global_lib_id,)).fetchone()

    if not global_lib:
        flash('Biblioteca global no encontrada', 'error')
        return redirect(url_for('global_libraries'))

    # Get all manual libraries associated with this global library
    manual_libraries = conn.execute('''
        SELECT l.id, l.library_name, l.version, l.type, l.source_url,
               l.description, l.latest_safe_version, l.latest_version,
               s.id as scan_id, s.url, s.title, s.scan_date,
               c.name as project_name, c.id as project_id
        FROM libraries l
        INNER JOIN scans s ON l.scan_id = s.id
        LEFT JOIN projects c ON s.project_id = c.id
        WHERE l.global_library_id = ? AND l.is_manual = 1
        ORDER BY s.scan_date DESC, l.library_name
    ''', (global_lib_id,)).fetchall()

    # Add vulnerability status to each library
    manual_libs_with_vuln = []
    for lib in manual_libraries:
        lib_dict = dict(lib)
        # Check vulnerability using current version vs safe version (preferring library-specific, then global)
        current_version = lib['version']
        safe_version = lib['latest_safe_version'] or global_lib['latest_safe_version']
        lib_dict['is_vulnerable'] = has_vulnerability(current_version, safe_version)
        manual_libs_with_vuln.append(lib_dict)

    conn.close()

    return render_template('manual_libraries_by_global.html',
                           global_library=global_lib,
                           manual_libraries=manual_libs_with_vuln)

@app.route('/associate-library/<int:library_id>', methods=['POST'])
@login_required
def associate_library(library_id):
    global_library_id = request.form.get('global_library_id')

    if not global_library_id or not global_library_id.isdigit():
        flash('Seleccione una biblioteca global válida.', 'error')
        return redirect(url_for('asociar_bibliotecas'))

    try:
        conn = get_db_connection()
        # Check if library exists and is manual
        lib = conn.execute('SELECT id FROM libraries WHERE id = ? AND is_manual = 1', (library_id,)).fetchone()
        if not lib:
            flash('La biblioteca manual no existe.', 'error')
            conn.close()
            return redirect(url_for('asociar_bibliotecas'))

        conn.execute('UPDATE libraries SET global_library_id = ? WHERE id = ?', (global_library_id, library_id))
        conn.commit()
        conn.close()
        flash('Biblioteca asociada exitosamente.', 'success')
    except Exception as e:
        flash(f'Error al asociar la biblioteca: {str(e)}', 'error')

    return redirect(url_for('asociar_bibliotecas'))

# Project Management Routes
@app.route('/projects')
@login_required
def projects():
    conn = get_db_connection()
    projects = conn.execute('''
        SELECT c.*,
               COUNT(s.id) as scan_count,
               MAX(s.scan_date) as last_scan,
               COUNT(CASE WHEN s.id IS NOT NULL AND (s.reviewed = 0 OR s.reviewed IS NULL) THEN 1 END) as pending_scans,
               COUNT(CASE WHEN s.id IS NOT NULL AND s.reviewed = 1 THEN 1 END) as completed_scans
        FROM projects c
        LEFT JOIN scans s ON c.id = s.project_id
        WHERE c.is_active = 1
        GROUP BY c.id, c.name, c.description, c.contact_email, c.contact_phone, c.website, c.created_date, c.updated_date, c.is_active
        ORDER BY c.name
    ''').fetchall()
    conn.close()
    return render_template('projects.html', projects=projects)

@app.route('/add-project', methods=['POST'])
@login_required
def add_project():
    try:
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        contact_email = request.form.get('contact_email', '').strip()
        contact_phone = request.form.get('contact_phone', '').strip()
        website = request.form.get('website', '').strip()

        if not name:
            flash('El nombre del proyecto es requerido', 'error')
            return redirect(url_for('projects'))

        conn = get_db_connection()
        cursor = conn.cursor()

        # Check if project already exists
        cursor.execute("SELECT id FROM projects WHERE name = ? AND is_active = 1", (name,))
        if cursor.fetchone():
            flash(f'Ya existe un proyecto con el nombre "{name}"', 'error')
            conn.close()
            return redirect(url_for('projects'))

        cursor.execute('''
            INSERT INTO projects (name, description, contact_email, contact_phone, website)
            VALUES (?, ?, ?, ?, ?)
        ''', (name, description, contact_email, contact_phone, website))

        conn.commit()
        conn.close()
        flash(f'Proyecto "{name}" agregado exitosamente', 'success')

    except Exception as e:
        flash(f'Error al agregar proyecto: {str(e)}', 'error')

    return redirect(url_for('projects'))

@app.route('/edit-project/<int:project_id>', methods=['POST'])
@login_required
def edit_project(project_id):
    try:
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        contact_email = request.form.get('contact_email', '').strip()
        contact_phone = request.form.get('contact_phone', '').strip()
        website = request.form.get('website', '').strip()

        if not name:
            flash('El nombre del proyecto es requerido', 'error')
            return redirect(url_for('projects'))

        conn = get_db_connection()
        cursor = conn.cursor()

        # Check if another project with same name exists
        cursor.execute("SELECT id FROM projects WHERE name = ? AND id != ? AND is_active = 1",
                      (name, project_id))
        if cursor.fetchone():
            flash(f'Ya existe otro proyecto con el nombre "{name}"', 'error')
            conn.close()
            return redirect(url_for('projects'))

        cursor.execute('''
            UPDATE projects
            SET name = ?, description = ?, contact_email = ?, contact_phone = ?,
                website = ?, updated_date = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (name, description, contact_email, contact_phone, website, project_id))

        if cursor.rowcount == 0:
            flash('Proyecto no encontrado', 'error')
        else:
            flash(f'Proyecto "{name}" actualizado exitosamente', 'success')

        conn.commit()
        conn.close()

    except Exception as e:
        flash(f'Error al actualizar proyecto: {str(e)}', 'error')

    return redirect(url_for('projects'))

@app.route('/delete-project/<int:project_id>', methods=['POST'])
@login_required
def delete_project(project_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Get project name before deletion for flash message
        cursor.execute("SELECT name FROM projects WHERE id = ?", (project_id,))
        project = cursor.fetchone()

        if not project:
            flash('Proyecto no encontrado', 'error')
            conn.close()
            return redirect(url_for('projects'))

        # Soft delete - mark as inactive instead of deleting
        cursor.execute("UPDATE projects SET is_active = 0 WHERE id = ?", (project_id,))
        conn.commit()
        conn.close()

        flash(f'Proyecto "{project["name"]}" desactivado exitosamente', 'success')

    except Exception as e:
        flash(f'Error al eliminar proyecto: {str(e)}', 'error')

    return redirect(url_for('projects'))

@app.route('/project/<int:project_id>')
@login_required
def project_detail(project_id):
    conn = get_db_connection()

    # Get project info
    project = conn.execute('SELECT * FROM projects WHERE id = ? AND is_active = 1', (project_id,)).fetchone()
    if not project:
        flash('Proyecto no encontrado', 'error')
        conn.close()
        return redirect(url_for('projects'))

    # Get search and status filter parameters
    search_query = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '').strip()

    # Build WHERE clause for project, search, and status filters
    where_conditions = ["s.project_id = ?"]
    query_params = [project_id]

    if search_query:
        # Search in titles, URLs, and associated libraries
        search_condition = """(
            s.title LIKE ? OR
            s.url LIKE ? OR
            s.id IN (
                SELECT DISTINCT l.scan_id
                FROM libraries l
                WHERE l.library_name LIKE ? OR l.description LIKE ?
            )
        )"""
        where_conditions.append(search_condition)
        search_param = f"%{search_query}%"
        query_params.extend([search_param, search_param, search_param, search_param])

    if status_filter:
        if status_filter == 'revisado':
            where_conditions.append("s.reviewed = 1")
        elif status_filter == 'pendiente':
            where_conditions.append("s.reviewed = 0")

    where_clause = "WHERE " + " AND ".join(where_conditions)

    # Get project scans with detailed statistics (matching dashboard counters)
    scans_query = f'''
        SELECT s.*,
               COUNT(DISTINCT l.id) as library_count,
               COUNT(DISTINCT vs.id) as version_string_count,
               COUNT(DISTINCT fu.id) as file_count,
               COUNT(DISTINCT CASE WHEN fu.status_code IS NOT NULL AND fu.status_code != 200 THEN fu.id END) as error_count
        FROM scans s
        LEFT JOIN libraries l ON s.id = l.scan_id
        LEFT JOIN version_strings vs ON s.id = vs.scan_id
        LEFT JOIN file_urls fu ON s.id = fu.scan_id
        {where_clause}
        GROUP BY s.id
        ORDER BY s.scan_date DESC
    '''
    scans_raw = conn.execute(scans_query, query_params).fetchall()

    # Calculate vulnerability_count using has_vulnerability function for each scan
    scans = []
    for scan in scans_raw:
        scan_dict = dict(scan)

        # Get libraries for this scan to count vulnerabilities correctly
        libraries_query = '''
            SELECT l.version, COALESCE(l.latest_safe_version, gl.latest_safe_version) as safe_version
            FROM libraries l
            LEFT JOIN global_libraries gl ON l.library_name = gl.library_name AND l.type = gl.type
            WHERE l.scan_id = ?
            AND COALESCE(l.latest_safe_version, gl.latest_safe_version) IS NOT NULL
            AND l.version IS NOT NULL
            AND l.version != ''
        '''
        libraries = conn.execute(libraries_query, (scan['id'],)).fetchall()

        vulnerability_count = sum(1 for lib in libraries
                                if has_vulnerability(lib['version'], lib['safe_version']))

        scan_dict['vulnerability_count'] = vulnerability_count
        scans.append(scan_dict)

    # Get project statistics (matching dashboard structure) with search filter
    stats_query = f'''
        SELECT
            COUNT(DISTINCT s.id) as total_scans,
            COUNT(DISTINCT l.id) as total_libraries,
            COUNT(DISTINCT vs.id) as total_version_strings,
            COUNT(DISTINCT fu.id) as total_files,
            COUNT(DISTINCT CASE WHEN fu.status_code IS NOT NULL AND fu.status_code != 200 THEN fu.id END) as total_errors,
            MAX(s.scan_date) as last_scan,
            MIN(s.scan_date) as first_scan
        FROM scans s
        LEFT JOIN libraries l ON s.id = l.scan_id
        LEFT JOIN version_strings vs ON s.id = vs.scan_id
        LEFT JOIN file_urls fu ON s.id = fu.scan_id
        {where_clause}
    '''
    stats_raw = conn.execute(stats_query, query_params).fetchone()

    # Calculate total_vulnerabilities using has_vulnerability function
    total_vulnerabilities_query = f'''
        SELECT l.version, COALESCE(l.latest_safe_version, gl.latest_safe_version) as safe_version
        FROM scans s
        LEFT JOIN libraries l ON s.id = l.scan_id
        LEFT JOIN global_libraries gl ON l.library_name = gl.library_name AND l.type = gl.type
        {where_clause}
        AND COALESCE(l.latest_safe_version, gl.latest_safe_version) IS NOT NULL
        AND l.version IS NOT NULL
        AND l.version != ''
    '''
    all_libraries = conn.execute(total_vulnerabilities_query, query_params).fetchall()

    total_vulnerabilities = sum(1 for lib in all_libraries
                              if has_vulnerability(lib['version'], lib['safe_version']))

    # Convert stats to dict and add calculated vulnerabilities
    stats = dict(stats_raw)
    stats['total_vulnerabilities'] = total_vulnerabilities

    # Get all projects for the edit modal dropdown
    projects = conn.execute('SELECT id, name FROM projects WHERE is_active = 1 ORDER BY name').fetchall()

    conn.close()
    return render_template('project_detail.html', project=project, scans=scans, stats=stats, projects=projects)

@app.route('/export-project-data/<int:project_id>/<format>')
@login_required
def export_project_data(project_id, format):
    """Export specific project data with all scans and libraries in the specified format"""
    try:
        conn = get_db_connection()

        # Get project info
        project = conn.execute('SELECT * FROM projects WHERE id = ? AND is_active = 1', (project_id,)).fetchone()
        if not project:
            flash('Proyecto no encontrado', 'error')
            conn.close()
            return redirect(url_for('projects'))

        # Get all scans for this project with libraries
        scans_data = conn.execute('''
            SELECT s.*,
                   GROUP_CONCAT(DISTINCT l.library_name || '|' || COALESCE(l.version, '') || '|' || l.type) as libraries_info
            FROM scans s
            LEFT JOIN libraries l ON s.id = l.scan_id
            WHERE s.project_id = ?
            GROUP BY s.id
            ORDER BY s.scan_date DESC
        ''', (project_id,)).fetchall()

        # Get statistics
        stats = conn.execute('''
            SELECT
                COUNT(DISTINCT s.id) as total_scans,
                COUNT(DISTINCT l.library_name) as unique_libraries,
                COUNT(DISTINCT fu.id) as total_files,
                SUM(CASE WHEN s.reviewed = 1 THEN 1 ELSE 0 END) as reviewed_scans,
                SUM(CASE WHEN s.reviewed = 0 THEN 1 ELSE 0 END) as pending_scans
            FROM scans s
            LEFT JOIN libraries l ON s.id = l.scan_id
            LEFT JOIN file_urls fu ON s.id = fu.scan_id
            WHERE s.project_id = ?
        ''', (project_id,)).fetchone()

        conn.close()

        if format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)

            # Write project info section
            writer.writerow(['# INFORMACIÓN DEL PROYECTO'])
            writer.writerow(['Nombre', 'Descripción', 'Email', 'Teléfono', 'Sitio Web', 'Fecha Creación'])
            writer.writerow([
                project['name'],
                project['description'] or '',
                project['contact_email'] or '',
                project['contact_phone'] or '',
                project['website'] or '',
                project['created_date'][:10] if project['created_date'] else ''
            ])
            writer.writerow([])  # Empty row for separation

            # Write statistics section
            writer.writerow(['# ESTADÍSTICAS'])
            writer.writerow(['Total Escaneos', 'Bibliotecas Únicas', 'Total Archivos', 'Revisados', 'Pendientes'])
            writer.writerow([
                stats['total_scans'] or 0,
                stats['unique_libraries'] or 0,
                stats['total_files'] or 0,
                stats['reviewed_scans'] or 0,
                stats['pending_scans'] or 0
            ])
            writer.writerow([])  # Empty row for separation

            # Write scans section
            writer.writerow(['# ESCANEOS REALIZADOS'])
            writer.writerow(['URL', 'Título', 'Fecha', 'Estado HTTP', 'Revisado', 'Bibliotecas Detectadas'])

            for scan in scans_data:
                libraries = ''
                if scan['libraries_info']:
                    lib_list = []
                    for lib_info in scan['libraries_info'].split(','):
                        parts = lib_info.split('|')
                        if len(parts) >= 2:
                            lib_name = parts[0]
                            lib_version = parts[1] if parts[1] else 'N/A'
                            lib_list.append(f"{lib_name} ({lib_version})")
                    libraries = '; '.join(lib_list)

                writer.writerow([
                    scan['url'],
                    scan['title'] or '',
                    scan['scan_date'][:16] if scan['scan_date'] else '',
                    scan['status_code'] or 'Error',
                    'Sí' if scan['reviewed'] == 1 else 'No',
                    libraries or 'Sin bibliotecas'
                ])

            response = make_response(output.getvalue())
            filename = f"project_{project['name'].replace(' ', '_')}_data.csv"
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            response.headers["Content-type"] = "text/csv"
            return response

        elif format == 'json':
            # Prepare JSON data
            export_data = {
                'project': {
                    'id': project['id'],
                    'name': project['name'],
                    'description': project['description'],
                    'contact_email': project['contact_email'],
                    'contact_phone': project['contact_phone'],
                    'website': project['website'],
                    'created_date': project['created_date']
                },
                'statistics': {
                    'total_scans': stats['total_scans'] or 0,
                    'unique_libraries': stats['unique_libraries'] or 0,
                    'total_files': stats['total_files'] or 0,
                    'reviewed_scans': stats['reviewed_scans'] or 0,
                    'pending_scans': stats['pending_scans'] or 0
                },
                'scans': []
            }

            for scan in scans_data:
                scan_obj = {
                    'id': scan['id'],
                    'url': scan['url'],
                    'title': scan['title'],
                    'scan_date': scan['scan_date'],
                    'status_code': scan['status_code'],
                    'reviewed': bool(scan['reviewed']),
                    'libraries': []
                }

                if scan['libraries_info']:
                    for lib_info in scan['libraries_info'].split(','):
                        parts = lib_info.split('|')
                        if len(parts) >= 3:
                            scan_obj['libraries'].append({
                                'name': parts[0],
                                'version': parts[1] if parts[1] else None,
                                'type': parts[2]
                            })

                export_data['scans'].append(scan_obj)

            response = make_response(json.dumps(export_data, indent=2, ensure_ascii=False))
            filename = f"project_{project['name'].replace(' ', '_')}_data.json"
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            response.headers["Content-type"] = "application/json"
            return response
        else:
            flash('Formato de exportación no válido', 'error')
            return redirect(url_for('project_detail', project_id=project_id))

    except Exception as e:
        flash(f'Error al exportar datos del proyecto: {str(e)}', 'error')
        return redirect(url_for('project_detail', project_id=project_id))

@app.route('/import-project-data/<int:project_id>', methods=['POST'])
@login_required
def import_project_data(project_id):
    """Import scan data from CSV or JSON file for a specific project"""
    try:
        # Verify project exists
        conn = get_db_connection()
        project = conn.execute('SELECT * FROM projects WHERE id = ? AND is_active = 1', (project_id,)).fetchone()
        if not project:
            flash('Proyecto no encontrado', 'error')
            conn.close()
            return redirect(url_for('projects'))

        if 'file' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('project_detail', project_id=project_id))

        file = request.files['file']
        if file.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('project_detail', project_id=project_id))

        if not file.filename.lower().endswith(('.csv', '.json')):
            flash('Solo se permiten archivos CSV o JSON', 'error')
            return redirect(url_for('project_detail', project_id=project_id))

        skip_duplicates = request.form.get('skip_duplicates') == 'on'
        cursor = conn.cursor()
        imported_scans = 0
        skipped_urls = []

        if file.filename.lower().endswith('.csv'):
            # Parse CSV file
            content = file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(content))

            for row in csv_reader:
                url = row.get('URL', '').strip()
                if not url:
                    continue

                # Check for duplicate URL if skip_duplicates is enabled
                if skip_duplicates:
                    existing = cursor.execute(
                        'SELECT id FROM scans WHERE url = ? AND project_id = ?',
                        (url, project_id)
                    ).fetchone()
                    if existing:
                        skipped_urls.append(url)
                        continue

                # Insert scan
                title = row.get('Título', '').strip() or None
                scan_date = row.get('Fecha', '').strip() or get_chile_time().isoformat()
                status_code = int(row.get('Estado HTTP', 0)) if row.get('Estado HTTP', '').isdigit() else None
                reviewed = 1 if row.get('Revisado', '').lower() in ['sí', 'si', 'yes', '1', 'true'] else 0

                cursor.execute('''
                    INSERT INTO scans (url, title, scan_date, status_code, reviewed, project_id, headers)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (url, title, scan_date, status_code, reviewed, project_id, '{}'))

                scan_id = cursor.lastrowid
                imported_scans += 1

                # Parse and insert libraries if present
                libraries_str = row.get('Bibliotecas Detectadas', '').strip()
                if libraries_str and libraries_str != 'Sin bibliotecas':
                    # Parse format: "jQuery (3.6.0); Bootstrap (5.1.3)"
                    for lib_str in libraries_str.split(';'):
                        lib_str = lib_str.strip()
                        if '(' in lib_str and ')' in lib_str:
                            lib_name = lib_str[:lib_str.index('(')].strip()
                            lib_version = lib_str[lib_str.index('(')+1:lib_str.index(')')].strip()
                            if lib_version == 'N/A':
                                lib_version = None

                            cursor.execute('''
                                INSERT INTO libraries (scan_id, library_name, version, type, is_manual)
                                VALUES (?, ?, ?, 'js', 1)
                            ''', (scan_id, lib_name, lib_version))

        elif file.filename.lower().endswith('.json'):
            # Parse JSON file
            content = file.read().decode('utf-8')
            data = json.loads(content)

            scans_list = data.get('scans', [])
            for scan_data in scans_list:
                url = scan_data.get('url', '').strip()
                if not url:
                    continue

                # Check for duplicate URL if skip_duplicates is enabled
                if skip_duplicates:
                    existing = cursor.execute(
                        'SELECT id FROM scans WHERE url = ? AND project_id = ?',
                        (url, project_id)
                    ).fetchone()
                    if existing:
                        skipped_urls.append(url)
                        continue

                # Insert scan
                title = scan_data.get('title') or None
                scan_date = scan_data.get('scan_date') or get_chile_time().isoformat()
                status_code = scan_data.get('status_code')
                reviewed = 1 if scan_data.get('reviewed') else 0

                cursor.execute('''
                    INSERT INTO scans (url, title, scan_date, status_code, reviewed, project_id, headers)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (url, title, scan_date, status_code, reviewed, project_id, '{}'))

                scan_id = cursor.lastrowid
                imported_scans += 1

                # Insert libraries if present
                for lib in scan_data.get('libraries', []):
                    lib_name = lib.get('name')
                    if lib_name:
                        cursor.execute('''
                            INSERT INTO libraries (scan_id, library_name, version, type, is_manual)
                            VALUES (?, ?, ?, ?, 1)
                        ''', (scan_id, lib_name, lib.get('version'), lib.get('type', 'js')))

        conn.commit()
        conn.close()

        # Prepare success message
        if imported_scans > 0:
            msg = f'Se importaron {imported_scans} escaneo(s) exitosamente'
            if skipped_urls:
                msg += f'. Se omitieron {len(skipped_urls)} URL(s) duplicadas'
            flash(msg, 'success')
        else:
            if skipped_urls:
                flash(f'No se importaron nuevos escaneos. {len(skipped_urls)} URL(s) ya existían', 'warning')
            else:
                flash('No se encontraron escaneos válidos en el archivo', 'warning')

        return redirect(url_for('project_detail', project_id=project_id))

    except json.JSONDecodeError:
        flash('Error al parsear el archivo JSON. Verifica el formato', 'error')
        return redirect(url_for('project_detail', project_id=project_id))
    except Exception as e:
        flash(f'Error al importar datos: {str(e)}', 'error')
        return redirect(url_for('project_detail', project_id=project_id))

@app.route('/export-projects/<format>')
@login_required
def export_projects(format):
    """Export all projects and their associated URLs in the specified format"""
    try:
        conn = get_db_connection()

        # Get all projects with their scan counts
        projects = conn.execute('''
            SELECT c.*,
                   COUNT(DISTINCT s.id) as scan_count
            FROM projects c
            LEFT JOIN scans s ON c.id = s.project_id
            WHERE c.is_active = 1
            GROUP BY c.id
            ORDER BY c.name
        ''').fetchall()

        if format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)

            # Write headers for projects section
            writer.writerow(['# PROYECTOS'])
            writer.writerow(['Nombre', 'Descripción', 'Email', 'Teléfono', 'Sitio Web', 'Fecha Creación', 'Total URLs'])

            # Write project data
            for project in projects:
                writer.writerow([
                    project['name'],
                    project['description'] or '',
                    project['contact_email'] or '',
                    project['contact_phone'] or '',
                    project['website'] or '',
                    project['created_date'][:10] if project['created_date'] else '',
                    project['scan_count']
                ])

            writer.writerow([])  # Empty row separator
            writer.writerow(['# URLS ASOCIADAS'])
            writer.writerow(['Proyecto', 'URL', 'Título', 'Fecha Escaneo', 'Estado', 'Librerías'])

            # Get all scans for active projects
            scans = conn.execute('''
                SELECT c.name as project_name, s.url, s.title, s.scan_date, s.status_code,
                       COUNT(DISTINCT l.id) as library_count
                FROM scans s
                JOIN projects c ON s.project_id = c.id
                LEFT JOIN libraries l ON s.id = l.scan_id
                WHERE c.is_active = 1
                GROUP BY s.id
                ORDER BY c.name, s.scan_date DESC
            ''').fetchall()

            for scan in scans:
                writer.writerow([
                    scan['project_name'],
                    scan['url'],
                    scan['title'] or '',
                    scan['scan_date'][:10] if scan['scan_date'] else '',
                    scan['status_code'] or 'Error',
                    scan['library_count']
                ])

            conn.close()

            # Create response
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename=projects_export_{get_chile_time().strftime("%Y%m%d_%H%M%S")}.csv'}
            )

        elif format == 'json':
            # Build JSON structure
            export_data = {
                'export_date': get_chile_time().isoformat(),
                'total_projects': len(projects),
                'projects': []
            }

            for project in projects:
                # Get scans for this project
                project_scans = conn.execute('''
                    SELECT s.*, COUNT(DISTINCT CASE WHEN l.type = 'js' THEN l.id END) as libraries_count,
                           COUNT(DISTINCT CASE WHEN f.file_type = 'js' THEN f.id END) as files_count
                    FROM scans s
                    LEFT JOIN libraries l ON s.id = l.scan_id
                    LEFT JOIN file_urls f ON s.id = f.scan_id
                    WHERE s.project_id = ?
                    GROUP BY s.id
                    ORDER BY s.scan_date DESC
                ''', (project['id'],)).fetchall()

                project_data = {
                    'name': project['name'],
                    'description': project['description'],
                    'contact_email': project['contact_email'],
                    'contact_phone': project['contact_phone'],
                    'website': project['website'],
                    'created_date': project['created_date'],
                    'scans': []
                }

                for scan in project_scans:
                    project_data['scans'].append({
                        'url': scan['url'],
                        'title': scan['title'],
                        'scan_date': scan['scan_date'],
                        'status_code': scan['status_code'],
                        'libraries_count': scan['libraries_count'],
                        'files_count': scan['files_count']
                    })

                export_data['projects'].append(project_data)

            conn.close()

            return Response(
                json.dumps(export_data, indent=2, ensure_ascii=False),
                mimetype='application/json',
                headers={'Content-Disposition': f'attachment; filename=projects_export_{get_chile_time().strftime("%Y%m%d_%H%M%S")}.json'}
            )

        elif format == 'xlsx':
            # Create Excel workbook with one sheet per project
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Header styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Subheader styles
            subheader_font = Font(bold=True)
            subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

            for project in projects:
                # Create sheet for each project (limit sheet name to 31 chars)
                sheet_name = project['name'][:31] if len(project['name']) > 31 else project['name']
                # Remove invalid characters for Excel sheet names
                for char in ['\\', '/', '*', '[', ']', ':', '?']:
                    sheet_name = sheet_name.replace(char, '_')

                ws = wb.create_sheet(title=sheet_name)

                # Project Information Section
                ws['A1'] = 'INFORMACIÓN DEL PROYECTO'
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:F1')

                # Project details
                project_info = [
                    ['Campo', 'Valor'],
                    ['Nombre', project['name']],
                    ['Descripción', project['description'] or ''],
                    ['Email', project['contact_email'] or ''],
                    ['Teléfono', project['contact_phone'] or ''],
                    ['Sitio Web', project['website'] or ''],
                    ['Proyecto desde', project['created_date'][:10] if project['created_date'] else ''],
                    ['Total de escaneos', project['scan_count']]
                ]

                for row_idx, row_data in enumerate(project_info, start=3):
                    for col_idx, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        if row_idx == 3:  # Header row
                            cell.font = subheader_font
                            cell.fill = subheader_fill

                # URLs Section (starting from row 13)
                ws['A13'] = 'URLS ESCANEADAS'
                ws['A13'].font = Font(bold=True, size=14)
                ws.merge_cells('A13:G13')

                # Get scans for this project
                project_scans = conn.execute('''
                    SELECT s.*, COUNT(DISTINCT CASE WHEN l.type = 'js' THEN l.id END) as libraries_count,
                           COUNT(DISTINCT CASE WHEN f.file_type = 'js' THEN f.id END) as files_count
                    FROM scans s
                    LEFT JOIN libraries l ON s.id = l.scan_id
                    LEFT JOIN file_urls f ON s.id = f.scan_id
                    WHERE s.project_id = ?
                    GROUP BY s.id
                    ORDER BY s.scan_date DESC
                ''', (project['id'],)).fetchall()

                # Headers for scans table
                scan_headers = ['URL', 'Título', 'Fecha Escaneo', 'Estado', 'Librerías', 'Archivos', 'Score Seguridad']
                for col_idx, header in enumerate(scan_headers, start=1):
                    cell = ws.cell(row=15, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # Scan data
                for row_idx, scan in enumerate(project_scans, start=16):
                    ws.cell(row=row_idx, column=1, value=scan['url'])
                    ws.cell(row=row_idx, column=2, value=scan['title'] or '')
                    ws.cell(row=row_idx, column=3, value=scan['scan_date'][:16] if scan['scan_date'] else '')
                    ws.cell(row=row_idx, column=4, value=scan['status_code'] or 'Error')
                    ws.cell(row=row_idx, column=5, value=scan['libraries_count'])
                    ws.cell(row=row_idx, column=6, value=scan['files_count'])

                    # Try to get security score from headers
                    if scan['headers']:
                        try:
                            headers_data = json.loads(scan['headers'])
                            security_score = calculate_security_score(headers_data)
                            ws.cell(row=row_idx, column=7, value=f"{security_score}%")
                        except:
                            ws.cell(row=row_idx, column=7, value='N/A')
                    else:
                        ws.cell(row=row_idx, column=7, value='N/A')

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Create summary sheet
            summary = wb.create_sheet(title="Resumen", index=0)
            summary['A1'] = 'RESUMEN DE EXPORTACIÓN'
            summary['A1'].font = Font(bold=True, size=14)
            summary.merge_cells('A1:D1')

            summary_data = [
                ['Fecha de exportación', format_chile_time(fmt='%Y-%m-%d %H:%M:%S')],
                ['Total de proyectos', len(projects)],
                ['Total de URLs escaneadas', conn.execute('SELECT COUNT(*) as cnt FROM scans WHERE project_id IS NOT NULL').fetchone()['cnt']],
                ['Proyecto con más escaneos', '']
            ]

            # Find project with most scans
            if projects:
                max_project = max(projects, key=lambda x: x['scan_count'])
                summary_data[3][1] = f"{max_project['name']} ({max_project['scan_count']} escaneos)"

            for row_idx, row_data in enumerate(summary_data, start=3):
                for col_idx, value in enumerate(row_data, start=1):
                    summary.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx == 1:
                        summary.cell(row=row_idx, column=col_idx).font = Font(bold=True)

            # Auto-adjust summary columns
            for column in summary.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                summary.column_dimensions[column_letter].width = adjusted_width

            conn.close()

            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename=projects_export_{get_chile_time().strftime("%Y%m%d_%H%M%S")}.xlsx'}
            )

        else:
            flash('Formato de exportación no válido', 'error')
            return redirect(url_for('projects'))

    except Exception as e:
        flash(f'Error al exportar proyectos: {str(e)}', 'error')
        return redirect(url_for('projects'))

@app.route('/import-projects', methods=['POST'])
@login_required
def import_projects():
    """Import projects from CSV or JSON file"""
    try:
        if 'file' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('projects'))

        file = request.files['file']
        if file.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('projects'))

        if not file.filename.lower().endswith(('.csv', '.json')):
            flash('Solo se permiten archivos CSV o JSON', 'error')
            return redirect(url_for('projects'))

        conn = get_db_connection()
        cursor = conn.cursor()
        imported_projects = 0
        imported_urls = 0
        errors = []

        if file.filename.lower().endswith('.csv'):
            # Parse CSV file
            content = file.read().decode('utf-8')
            lines = content.strip().split('\n')

            # Find sections
            project_section_start = -1
            url_section_start = -1

            for i, line in enumerate(lines):
                if '# PROYECTOS' in line:
                    project_section_start = i + 2  # Skip header and column names
                elif '# URLS ASOCIADAS' in line:
                    url_section_start = i + 2  # Skip header and column names
                    break

            if project_section_start > 0:
                # Parse projects
                csv_reader = csv.DictReader(io.StringIO('\n'.join(lines[project_section_start-1:url_section_start-3])))

                for row in csv_reader:
                    try:
                        project_name = row.get('Nombre', '').strip()
                        if not project_name:
                            continue

                        # Check if project already exists
                        cursor.execute("SELECT id FROM projects WHERE name = ?", (project_name,))
                        existing = cursor.fetchone()

                        if existing:
                            errors.append(f'Proyecto "{project_name}" ya existe')
                            continue

                        cursor.execute('''
                            INSERT INTO projects (name, description, contact_email, contact_phone, website)
                            VALUES (?, ?, ?, ?, ?)
                        ''', (
                            project_name,
                            row.get('Descripción', ''),
                            row.get('Email', ''),
                            row.get('Teléfono', ''),
                            row.get('Sitio Web', '')
                        ))

                        imported_projects += 1

                    except Exception as e:
                        errors.append(f'Error importando proyecto: {str(e)}')

            conn.commit()

        elif file.filename.lower().endswith('.json'):
            # Parse JSON file
            content = file.read().decode('utf-8')
            data = json.loads(content)

            if 'projects' not in data:
                flash('Formato de archivo JSON inválido', 'error')
                return redirect(url_for('projects'))

            for project_data in data['projects']:
                try:
                    project_name = project_data.get('name', '').strip()
                    if not project_name:
                        continue

                    # Check if project already exists
                    cursor.execute("SELECT id FROM projects WHERE name = ?", (project_name,))
                    existing = cursor.fetchone()

                    if existing:
                        errors.append(f'Proyecto "{project_name}" ya existe')
                        continue

                    cursor.execute('''
                        INSERT INTO projects (name, description, contact_email, contact_phone, website)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (
                        project_name,
                        project_data.get('description', ''),
                        project_data.get('contact_email', ''),
                        project_data.get('contact_phone', ''),
                        project_data.get('website', '')
                    ))

                    imported_projects += 1

                except Exception as e:
                    errors.append(f'Error importando proyecto {project_name}: {str(e)}')

            conn.commit()

        conn.close()

        # Prepare success message
        success_msg = f'Importación completada: {imported_projects} proyecto(s) importado(s)'
        if errors:
            success_msg += f', {len(errors)} error(es)'
            for error in errors[:5]:  # Show first 5 errors
                flash(error, 'warning')

        flash(success_msg, 'success')
        return redirect(url_for('projects'))

    except json.JSONDecodeError:
        flash('Error al parsear archivo JSON', 'error')
        return redirect(url_for('projects'))
    except Exception as e:
        flash(f'Error al importar proyectos: {str(e)}', 'error')
        return redirect(url_for('projects'))

def calculate_security_score(headers):
    """Calculate security score based on headers"""
    score = 0
    max_score = 7

    security_headers = [
        'Strict-Transport-Security',
        'Content-Security-Policy',
        'X-Frame-Options',
        'X-Content-Type-Options',
        'X-XSS-Protection',
        'Referrer-Policy',
        'Permissions-Policy'
    ]

    for header in security_headers:
        if header in headers:
            score += 1

    return int((score / max_score) * 100)

@app.route('/users')
@login_required
@admin_required
def users():
    conn = get_db_connection()
    users = conn.execute('SELECT id, username, role FROM users ORDER BY username').fetchall()
    conn.close()
    return render_template('users.html', users=users)

@app.route('/add_user', methods=['POST'])
@login_required
@admin_required
def add_user():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    role = request.form.get('role', 'analyst').strip()

    if not username or not password:
        flash('Usuario y contraseña son requeridos', 'error')
        return redirect(url_for('users'))

    if role not in ['admin', 'analyst']:
        role = 'analyst'  # Default to analyst if invalid role

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE username = ?", (username,))
    user = cursor.fetchone()
    if user:
        flash('El usuario ya existe', 'error')
        conn.close()
        return redirect(url_for('users'))

    cursor.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                   (username, generate_password_hash(password), role))
    conn.commit()
    conn.close()
    flash(f'Usuario "{username}" ({role}) agregado exitosamente', 'success')
    return redirect(url_for('users'))

@app.route('/change_password/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def change_password(user_id):
    new_password = request.form.get('new_password', '').strip()

    if not new_password:
        flash('La nueva contraseña no puede estar vacía', 'error')
        return redirect(url_for('users'))

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET password = ? WHERE id = ?",
                   (generate_password_hash(new_password), user_id))
    conn.commit()
    conn.close()
    flash('Contraseña actualizada exitosamente', 'success')
    return redirect(url_for('users'))

@app.route('/ayuda')
@login_required
def ayuda():
    """Página de ayuda y documentación"""
    return render_template('ayuda.html', current_page='ayuda')

@app.route('/historial')
@login_required
def historial():
    """Página principal de historial global de acciones"""
    page = request.args.get('page', 1, type=int)
    per_page = 50

    # Filtros
    user_filter = request.args.get('user', '').strip()
    action_filter = request.args.get('action', '').strip()
    table_filter = request.args.get('table', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    search = request.args.get('search', '').strip()

    conn = get_db_connection()

    # Construir query con filtros
    where_conditions = []
    params = []

    if user_filter:
        where_conditions.append("(username LIKE ? OR user_role LIKE ?)")
        params.extend([f'%{user_filter}%', f'%{user_filter}%'])

    if action_filter:
        where_conditions.append("action_type = ?")
        params.append(action_filter)

    if table_filter:
        where_conditions.append("target_table = ?")
        params.append(table_filter)

    if date_from:
        where_conditions.append("DATE(timestamp) >= ?")
        params.append(date_from)

    if date_to:
        where_conditions.append("DATE(timestamp) <= ?")
        params.append(date_to)

    if search:
        where_conditions.append("(target_description LIKE ? OR notes LIKE ? OR error_message LIKE ?)")
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])

    where_clause = " AND ".join(where_conditions) if where_conditions else "1=1"

    # Obtener total de registros para paginación
    count_query = f"SELECT COUNT(*) as total FROM action_history WHERE {where_clause}"
    total_records = conn.execute(count_query, params).fetchone()['total']

    # Calcular paginación
    total_pages = (total_records + per_page - 1) // per_page
    offset = (page - 1) * per_page
    has_prev = page > 1
    has_next = page < total_pages
    prev_num = page - 1 if has_prev else None
    next_num = page + 1 if has_next else None

    # Obtener registros paginados
    records_query = f'''
        SELECT * FROM action_history
        WHERE {where_clause}
        ORDER BY timestamp DESC
        LIMIT ? OFFSET ?
    '''
    records = conn.execute(records_query, params + [per_page, offset]).fetchall()

    # Datos para filtros
    users_query = "SELECT DISTINCT username, user_role FROM action_history ORDER BY username"
    users = conn.execute(users_query).fetchall()

    action_types_query = "SELECT DISTINCT action_type FROM action_history ORDER BY action_type"
    action_types = [row['action_type'] for row in conn.execute(action_types_query).fetchall()]

    table_names_query = "SELECT DISTINCT target_table FROM action_history ORDER BY target_table"
    table_names = [row['target_table'] for row in conn.execute(table_names_query).fetchall()]

    conn.close()

    return render_template('historial.html',
        records=[dict(record) for record in records],
        pagination={
            'page': page,
            'per_page': per_page,
            'total': total_records,
            'total_pages': total_pages,
            'has_prev': has_prev,
            'has_next': has_next,
            'prev_num': prev_num,
            'next_num': next_num
        },
        users=[dict(user) for user in users],
        action_types=action_types,
        table_names=table_names,
        filters={
            'user': user_filter,
            'action': action_filter,
            'table': table_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search
        },
        current_page='historial'
    )

@app.route('/historial/details/<int:action_id>')
@login_required
def historial_details(action_id):
    """Obtiene los detalles de una acción específica"""
    conn = get_db_connection()

    try:
        record = conn.execute(
            'SELECT * FROM action_history WHERE id = ?',
            (action_id,)
        ).fetchone()

        if not record:
            return jsonify({'error': 'Registro no encontrado'}), 404

        return jsonify(dict(record))

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/test-logging')
@login_required
def test_logging():
    """Ruta para probar el sistema de logging"""
    if not ENABLE_ACTION_LOGGING:
        return """
        <h2>🔧 Sistema de Logging DESHABILITADO</h2>
        <p>El logging está deshabilitado para evitar conflictos de base de datos.</p>
        <p><strong>Para habilitar:</strong></p>
        <pre>python3 toggle_logging.py on
source logging_config.sh
# Reiniciar servidor</pre>
        <a href="/">← Volver al inicio</a>
        """

    try:
        # Test logging simple
        log_user_action(
            action_type='TEST',
            target_table='system',
            target_description='Test de logging manual',
            success=True,
            notes='Prueba desde /test-logging'
        )

        return """
        <h2>✅ Test de Logging EXITOSO</h2>
        <p>El logging está funcionando correctamente.</p>
        <p><a href="/historial">Ver historial →</a></p>
        <p><a href="/">← Volver al inicio</a></p>
        """

    except Exception as e:
        return f"""
        <h2>❌ Test de Logging FALLIDO</h2>
        <p><strong>Error:</strong> {e}</p>
        <p>Recomendación: Mantener logging deshabilitado.</p>
        <pre>python3 toggle_logging.py off</pre>
        <a href="/">← Volver al inicio</a>
        """

@app.route('/historial/undo/<int:action_id>', methods=['POST'])
@login_required
def undo_action(action_id):
    """Deshace una acción específica del historial"""
    conn = get_db_connection()

    try:
        # Obtener registro del historial
        history_record = conn.execute(
            'SELECT * FROM action_history WHERE id = ?',
            (action_id,)
        ).fetchone()

        if not history_record:
            return jsonify({'success': False, 'message': 'Registro de historial no encontrado'}), 404

        # Verificar que se puede deshacer
        if history_record['action_type'] not in ['UPDATE', 'DELETE']:
            return jsonify({'success': False, 'message': 'Esta acción no se puede deshacer'}), 400

        if not history_record['data_before']:
            return jsonify({'success': False, 'message': 'No hay datos previos para restaurar'}), 400

        # Verificar permisos
        if not check_undo_permission(history_record['action_type'], history_record['target_table']):
            return jsonify({'success': False, 'message': 'No tienes permisos para deshacer esta acción'}), 403

        try:
            data_before = json.loads(history_record['data_before'])

            if history_record['action_type'] == 'DELETE':
                # Restaurar registro eliminado
                restore_deleted_record(history_record['target_table'], data_before)
                action_description = f"Restaurado registro eliminado de {history_record['target_table']} #{data_before.get('id', 'N/A')}"

            elif history_record['action_type'] == 'UPDATE':
                # Revertir cambios
                revert_update(history_record['target_table'],
                             history_record['target_id'], data_before)
                action_description = f"Revertidos cambios en {history_record['target_table']} #{history_record['target_id']}"

            # Registrar la acción de deshacer
            log_user_action(
                action_type='UNDO',
                target_table='action_history',
                target_id=action_id,
                target_description=f"Deshecha acción: {action_description}",
                data_before={'original_action_id': action_id, 'original_action_type': history_record['action_type']},
                data_after={'undone_by': session['username'], 'undone_at': get_chile_time().isoformat()},
                notes=f"Usuario {session['username']} deshizo acción {history_record['action_type']} del {history_record['timestamp']}"
            )

            return jsonify({
                'success': True,
                'message': f'Acción deshecha exitosamente: {action_description}'
            })

        except json.JSONDecodeError:
            return jsonify({'success': False, 'message': 'Error al procesar datos del historial'}), 500
        except Exception as e:
            # Registrar error en el historial
            log_user_action(
                action_type='UNDO',
                target_table='action_history',
                target_id=action_id,
                target_description=f"Error al deshacer acción #{action_id}",
                success=False,
                error_message=str(e)
            )
            return jsonify({'success': False, 'message': f'Error al deshacer acción: {str(e)}'}), 500

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error interno: {str(e)}'}), 500
    finally:
        conn.close()

def restore_deleted_record(table_name, data):
    """Restaura un registro eliminado"""
    conn = get_db_connection()

    try:
        # Manejar conflicto de ID
        original_id = data.get('id')
        if original_id and record_exists(table_name, original_id):
            # Generar nuevo ID
            new_id = get_next_available_id(table_name)
            data['id'] = new_id
            print(f"⚠️ ID {original_id} ya existe, usando nuevo ID {new_id}")

        # Remover campos automáticos que no deben insertarse
        data_copy = data.copy()

        # Construir query de inserción
        columns = ', '.join(data_copy.keys())
        placeholders = ', '.join(['?' for _ in data_copy])
        values = list(data_copy.values())

        conn.execute(f'INSERT INTO {table_name} ({columns}) VALUES ({placeholders})', values)
        conn.commit()

        print(f"✅ Registro restaurado en {table_name} con ID {data_copy.get('id')}")

    except Exception as e:
        print(f"❌ Error al restaurar registro en {table_name}: {e}")
        raise
    finally:
        conn.close()

def revert_update(table_name, record_id, previous_data):
    """Revierte un registro a su estado anterior"""
    conn = get_db_connection()

    try:
        # Verificar que el registro existe
        if not record_exists(table_name, record_id):
            raise Exception(f"El registro {table_name}#{record_id} ya no existe")

        # Remover el ID de los datos a actualizar
        update_data = {k: v for k, v in previous_data.items() if k != 'id'}

        if not update_data:
            raise Exception("No hay datos para revertir")

        # Construir query de actualización
        set_clause = ', '.join([f'{col} = ?' for col in update_data.keys()])
        values = list(update_data.values()) + [record_id]

        conn.execute(f'UPDATE {table_name} SET {set_clause} WHERE id = ?', values)
        conn.commit()

        print(f"✅ Registro revertido en {table_name}#{record_id}")

    except Exception as e:
        print(f"❌ Error al revertir registro en {table_name}#{record_id}: {e}")
        raise
    finally:
        conn.close()

def check_undo_permission(action_type, target_table):
    """Verifica permisos para deshacer acciones específicas"""
    user_role = session.get('user_role')

    # Acciones críticas solo para admins
    critical_actions = ['DELETE']
    critical_tables = ['users']
    admin_only_combinations = [
        ('UPDATE', 'users'),
        ('DELETE', 'scans'),
        ('DELETE', 'libraries')
    ]

    # Verificar si es acción crítica
    if action_type in critical_actions:
        return user_role == 'admin'

    # Verificar tabla crítica
    if target_table in critical_tables:
        return user_role == 'admin'

    # Verificar combinaciones específicas
    if (action_type, target_table) in admin_only_combinations:
        return user_role == 'admin'

    # Otras acciones permitidas para todos los usuarios autenticados
    return True

@app.route('/change_own_password', methods=['POST'])
@login_required
def change_own_password():
    current_password = request.form.get('current_password', '').strip()
    new_password = request.form.get('new_password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()

    if not current_password or not new_password or not confirm_password:
        flash('Todos los campos son requeridos', 'error')
        return redirect(request.referrer or url_for('index'))

    if new_password != confirm_password:
        flash('Las nuevas contraseñas no coinciden', 'error')
        return redirect(request.referrer or url_for('index'))

    if len(new_password) < 6:
        flash('La nueva contraseña debe tener al menos 6 caracteres', 'error')
        return redirect(request.referrer or url_for('index'))

    conn = get_db_connection()
    cursor = conn.cursor()

    # Get current user info
    cursor.execute("SELECT password FROM users WHERE id = ?", (session['user_id'],))
    user = cursor.fetchone()

    if not user or not check_password_hash(user['password'], current_password):
        flash('La contraseña actual es incorrecta', 'error')
        conn.close()
        return redirect(request.referrer or url_for('index'))

    # Update password
    cursor.execute("UPDATE users SET password = ? WHERE id = ?",
                   (generate_password_hash(new_password), session['user_id']))
    conn.commit()
    conn.close()

    flash('Tu contraseña ha sido actualizada exitosamente', 'success')
    return redirect(request.referrer or url_for('index'))

@app.route('/change_role/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def change_role(user_id):
    new_role = request.form.get('new_role', '').strip()

    if new_role not in ['admin', 'analyst']:
        flash('Rol inválido. Debe ser "admin" o "analyst"', 'error')
        return redirect(url_for('users'))

    conn = get_db_connection()
    cursor = conn.cursor()

    # Get current user info
    cursor.execute("SELECT username, role FROM users WHERE id = ?", (user_id,))
    user = cursor.fetchone()

    if not user:
        flash('Usuario no encontrado', 'error')
        conn.close()
        return redirect(url_for('users'))

    # Prevent changing role of default admin user 'gabo'
    if user['username'] == 'gabo' and new_role != 'admin':
        flash('No se puede cambiar el rol del usuario administrador por defecto', 'error')
        conn.close()
        return redirect(url_for('users'))

    # Update role
    cursor.execute("UPDATE users SET role = ? WHERE id = ?", (new_role, user_id))
    conn.commit()
    conn.close()

    role_name = 'Administrador' if new_role == 'admin' else 'Analista'
    flash(f'Rol actualizado exitosamente a {role_name}', 'success')
    return redirect(url_for('users'))

@app.route('/delete_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT username FROM users WHERE id = ?", (user_id,))
    user = cursor.fetchone()
    if user and user['username'] == 'gabo':
        flash('No se puede eliminar el usuario administrador por defecto', 'error')
        conn.close()
        return redirect(url_for('users'))

    cursor.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    flash('Usuario eliminado exitosamente', 'success')
    return redirect(url_for('users'))

@app.route('/export/db')
@login_required
def export_db():
    try:
        # Use absolute path and validate file exists to prevent path traversal
        import os
        db_path = os.path.abspath('analysis.db')

        if not os.path.exists(db_path) or not os.path.isfile(db_path):
            flash('Archivo de base de datos no encontrado', 'error')
            return redirect(url_for('index'))

        return send_file(db_path, as_attachment=True, download_name='analysis_backup.db')
    except Exception as e:
        flash(f'Error al exportar base de datos: {str(e)}', 'error')
        return redirect(url_for('index'))

# Global Libraries Import/Export Routes
@app.route('/export-global-libraries/<format>')
@login_required
def export_global_libraries(format):
    try:
        conn = get_db_connection()
        libraries = conn.execute('''
            SELECT library_name, type, latest_safe_version, latest_version,
                   description, vulnerability_info, source_url
            FROM global_libraries
            ORDER BY library_name
        ''').fetchall()
        conn.close()

        if format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)

            # Write headers
            writer.writerow([
                'Nombre', 'Tipo', 'Versión Segura', 'Última Versión',
                'Descripción', 'Vulnerabilidades', 'URL Fuente'
            ])

            # Write data
            for lib in libraries:
                writer.writerow([
                    lib['library_name'], lib['type'], lib['latest_safe_version'] or '',
                    lib['latest_version'] or '', lib['description'] or '',
                    lib['vulnerability_info'] or '', lib['source_url'] or ''
                ])

            response = make_response(output.getvalue())
            response.headers["Content-Disposition"] = "attachment; filename=global_libraries.csv"
            response.headers["Content-type"] = "text/csv"
            return response

        elif format == 'json':
            libraries_data = []
            for lib in libraries:
                libraries_data.append({
                    'library_name': lib['library_name'],
                    'type': lib['type'],
                    'latest_safe_version': lib['latest_safe_version'],
                    'latest_version': lib['latest_version'],
                    'description': lib['description'],
                    'vulnerability_info': lib['vulnerability_info'],
                    'source_url': lib['source_url']
                })

            response = make_response(json.dumps(libraries_data, indent=2))
            response.headers["Content-Disposition"] = "attachment; filename=global_libraries.json"
            response.headers["Content-type"] = "application/json"
            return response
        else:
            flash('Formato de exportación no válido', 'error')
            return redirect(url_for('global_libraries'))

    except Exception as e:
        flash(f'Error al exportar catálogo: {str(e)}', 'error')
        return redirect(url_for('global_libraries'))

@app.route('/import-global-libraries', methods=['POST'])
@login_required
def import_global_libraries():
    try:
        if 'file' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('global_libraries'))

        file = request.files['file']
        if file.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('global_libraries'))

        if not file.filename.lower().endswith(('.csv', '.json')):
            flash('Solo se permiten archivos CSV o JSON', 'error')
            return redirect(url_for('global_libraries'))

        conn = get_db_connection()
        cursor = conn.cursor()
        imported_count = 0
        errors = []

        if file.filename.lower().endswith('.csv'):
            # Parse CSV file
            content = file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(content))

            for row in csv_reader:
                try:
                    library_name = row.get('Nombre', '').strip()
                    if not library_name:
                        continue

                    library_type = row.get('Tipo', '').strip().lower()
                    if library_type not in ['js', 'css']:
                        library_type = 'js'  # default

                    # Check if library already exists
                    cursor.execute("SELECT id FROM global_libraries WHERE library_name = ?", (library_name,))
                    if cursor.fetchone():
                        errors.append(f'Librería "{library_name}" ya existe')
                        continue

                    cursor.execute('''
                        INSERT INTO global_libraries
                        (library_name, type, latest_safe_version, latest_version, description, vulnerability_info, source_url)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        library_name, library_type,
                        row.get('Versión Segura', '').strip() or None,
                        row.get('Última Versión', '').strip() or None,
                        row.get('Descripción', '').strip() or None,
                        row.get('Vulnerabilidades', '').strip() or None,
                        row.get('URL Fuente', '').strip() or None
                    ))
                    imported_count += 1

                except Exception as e:
                    errors.append(f'Error procesando "{library_name}": {str(e)}')

        elif file.filename.lower().endswith('.json'):
            # Parse JSON file
            content = file.read().decode('utf-8')
            libraries_data = json.loads(content)

            if not isinstance(libraries_data, list):
                raise ValueError("El archivo JSON debe contener una lista de librerías")

            for lib in libraries_data:
                try:
                    library_name = lib.get('library_name', '').strip()
                    if not library_name:
                        continue

                    library_type = lib.get('type', '').strip().lower()
                    if library_type not in ['js', 'css']:
                        library_type = 'js'  # default

                    # Check if library already exists
                    cursor.execute("SELECT id FROM global_libraries WHERE library_name = ?", (library_name,))
                    if cursor.fetchone():
                        errors.append(f'Librería "{library_name}" ya existe')
                        continue

                    cursor.execute('''
                        INSERT INTO global_libraries
                        (library_name, type, latest_safe_version, latest_version, description, vulnerability_info, source_url)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        library_name, library_type,
                        lib.get('latest_safe_version') or None,
                        lib.get('latest_version') or None,
                        lib.get('description') or None,
                        lib.get('vulnerability_info') or None,
                        lib.get('source_url') or None
                    ))
                    imported_count += 1

                except Exception as e:
                    errors.append(f'Error procesando "{library_name}": {str(e)}')

        conn.commit()
        conn.close()

        if imported_count > 0:
            flash(f'{imported_count} librerías importadas exitosamente', 'success')

        if errors:
            error_msg = f'{len(errors)} errores encontrados: ' + '; '.join(errors[:3])
            if len(errors) > 3:
                error_msg += f' y {len(errors) - 3} más...'
            flash(error_msg, 'warning')

        if imported_count == 0 and not errors:
            flash('No se encontraron librerías válidas para importar', 'warning')

    except Exception as e:
        flash(f'Error al importar catálogo: {str(e)}', 'error')

    return redirect(url_for('global_libraries'))

# Statistics Export/Import Routes
@app.route('/export-statistics/<format>')
@login_required
def export_statistics(format):
    try:
        conn = get_db_connection()

        # Get all scans first
        all_scans = conn.execute('''
            SELECT
                s.id,
                s.url,
                s.title,
                s.scan_date,
                s.status_code,
                c.name as project_name,
                COUNT(DISTINCT l2.id) as library_count,
                COUNT(DISTINCT f.id) as file_count
            FROM scans s
            LEFT JOIN projects c ON s.project_id = c.id
            LEFT JOIN libraries l2 ON s.id = l2.scan_id
            LEFT JOIN file_urls f ON s.id = f.scan_id
            GROUP BY s.id
            ORDER BY s.scan_date DESC
        ''').fetchall()

        # Filter vulnerable scans using has_vulnerability function
        vulnerable_scans = []
        for scan in all_scans:
            scan_dict = dict(scan)

            # Get libraries for this scan to count vulnerabilities correctly
            libraries = conn.execute('''
                SELECT l.version, COALESCE(l.latest_safe_version, gl.latest_safe_version) as safe_version
                FROM libraries l
                LEFT JOIN global_libraries gl ON l.library_name = gl.library_name AND l.type = gl.type
                WHERE l.scan_id = ?
                AND COALESCE(l.latest_safe_version, gl.latest_safe_version) IS NOT NULL
                AND l.version IS NOT NULL
                AND l.version != ''
            ''', (scan['id'],)).fetchall()

            vulnerability_count = sum(1 for lib in libraries
                                    if has_vulnerability(lib['version'], lib['safe_version']))

            if vulnerability_count > 0:
                scan_dict['vulnerability_count'] = vulnerability_count
                vulnerable_scans.append(scan_dict)

        # Also get detailed vulnerability info for each scan
        vulnerabilities = conn.execute('''
            SELECT
                l.scan_id,
                l.library_name,
                l.version,
                l.latest_safe_version,
                l.description,
                l.source_url
            FROM libraries l
            WHERE l.version IS NOT NULL
                AND l.latest_safe_version IS NOT NULL
                AND l.version != l.latest_safe_version
                AND l.version < l.latest_safe_version
            ORDER BY l.scan_id, l.library_name
        ''').fetchall()

        conn.close()

        if format == 'csv':
            # Create CSV response
            output = io.StringIO()
            writer = csv.writer(output)

            # Write header for scans
            writer.writerow(['URL', 'Título', 'Proyecto', 'Fecha Escaneo', 'Estado', 'Vulnerabilidades', 'Total Librerías', 'Total Archivos'])

            # Write scan data
            for scan in vulnerable_scans:
                writer.writerow([
                    scan['url'],
                    scan['title'] or '',
                    scan['project_name'] or 'Sin proyecto',
                    scan['scan_date'],
                    scan['status_code'],
                    scan['vulnerability_count'],
                    scan['library_count'],
                    scan['file_count']
                ])

            # Add separator
            writer.writerow([])
            writer.writerow(['--- Detalle de Vulnerabilidades ---'])
            writer.writerow(['ID Escaneo', 'Librería', 'Versión Actual', 'Versión Segura', 'Descripción', 'URL Fuente'])

            # Write vulnerability details
            for vuln in vulnerabilities:
                writer.writerow([
                    vuln['scan_id'],
                    vuln['library_name'],
                    vuln['version'],
                    vuln['latest_safe_version'],
                    vuln['description'] or '',
                    vuln['source_url'] or ''
                ])

            # Prepare response
            output.seek(0)
            response = Response(output.getvalue(), mimetype='text/csv')
            response.headers['Content-Disposition'] = f'attachment; filename=estadisticas_vulnerabilidades_{get_chile_time().strftime("%Y%m%d_%H%M%S")}.csv'
            return response

        elif format == 'json':
            # Prepare JSON data
            data = {
                'export_date': get_chile_time().isoformat(),
                'summary': {
                    'total_vulnerable_scans': len(vulnerable_scans),
                    'total_vulnerabilities': sum(scan['vulnerability_count'] for scan in vulnerable_scans)
                },
                'vulnerable_scans': [dict(scan) for scan in vulnerable_scans],
                'vulnerability_details': []
            }

            # Group vulnerabilities by scan_id
            scan_vulns = {}
            for vuln in vulnerabilities:
                scan_id = vuln['scan_id']
                if scan_id not in scan_vulns:
                    scan_vulns[scan_id] = []
                scan_vulns[scan_id].append({
                    'library_name': vuln['library_name'],
                    'current_version': vuln['version'],
                    'safe_version': vuln['latest_safe_version'],
                    'description': vuln['description'],
                    'source_url': vuln['source_url']
                })

            data['vulnerability_details'] = scan_vulns

            # Return JSON response
            response = Response(json.dumps(data, indent=2, ensure_ascii=False), mimetype='application/json')
            response.headers['Content-Disposition'] = f'attachment; filename=estadisticas_vulnerabilidades_{get_chile_time().strftime("%Y%m%d_%H%M%S")}.json'
            return response

        else:
            flash('Formato de exportación no válido', 'error')
            return redirect(url_for('statistics'))

    except Exception as e:
        flash(f'Error al exportar estadísticas: {str(e)}', 'error')
        return redirect(url_for('statistics'))

@app.route('/import-statistics', methods=['POST'])
@login_required
def import_statistics():
    try:
        if 'file' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('statistics'))

        file = request.files['file']
        if file.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('statistics'))

        if not file.filename.lower().endswith(('.csv', '.json')):
            flash('Solo se permiten archivos CSV o JSON', 'error')
            return redirect(url_for('statistics'))

        conn = get_db_connection()
        cursor = conn.cursor()
        imported_scans = 0
        updated_scans = 0
        errors = []

        if file.filename.lower().endswith('.json'):
            # Parse JSON file
            content = file.read().decode('utf-8')
            data = json.loads(content)

            if 'vulnerable_scans' in data:
                for scan_data in data['vulnerable_scans']:
                    try:
                        # Check if scan already exists
                        existing = cursor.execute('SELECT id FROM scans WHERE url = ?', (scan_data['url'],)).fetchone()

                        if existing:
                            # Update existing scan
                            scan_id = existing['id']
                            updated_scans += 1
                        else:
                            # Create new scan
                            cursor.execute('''
                                INSERT INTO scans (url, title, scan_date, status_code)
                                VALUES (?, ?, ?, ?)
                            ''', (
                                scan_data['url'],
                                scan_data.get('title', ''),
                                scan_data.get('scan_date', format_chile_time(fmt='%Y-%m-%d %H:%M:%S')),
                                scan_data.get('status_code', 200)
                            ))
                            scan_id = cursor.lastrowid
                            imported_scans += 1

                        # Import vulnerability details if available
                        if 'vulnerability_details' in data and str(scan_id) in data['vulnerability_details']:
                            for vuln in data['vulnerability_details'][str(scan_id)]:
                                # Check if library already exists for this scan
                                existing_lib = cursor.execute(
                                    'SELECT id FROM libraries WHERE scan_id = ? AND library_name = ?',
                                    (scan_id, vuln['library_name'])
                                ).fetchone()

                                if not existing_lib:
                                    cursor.execute('''
                                        INSERT INTO libraries (scan_id, library_name, version, latest_safe_version, description, source_url)
                                        VALUES (?, ?, ?, ?, ?, ?)
                                    ''', (
                                        scan_id,
                                        vuln['library_name'],
                                        vuln.get('current_version'),
                                        vuln.get('safe_version'),
                                        vuln.get('description'),
                                        vuln.get('source_url')
                                    ))

                    except Exception as e:
                        errors.append(f'Error procesando escaneo "{scan_data.get("url", "unknown")}": {str(e)}')

        elif file.filename.lower().endswith('.csv'):
            # Parse CSV file
            content = file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(content))

            for row in csv_reader:
                try:
                    # Skip separator rows
                    if '---' in str(row.get('URL', '')):
                        continue

                    url = row.get('URL', '').strip()
                    if not url or url == 'ID Escaneo':  # Skip vulnerability detail headers
                        continue

                    # Check if scan already exists
                    existing = cursor.execute('SELECT id FROM scans WHERE url = ?', (url,)).fetchone()

                    if existing:
                        scan_id = existing['id']
                        updated_scans += 1
                    else:
                        # Create new scan
                        cursor.execute('''
                            INSERT INTO scans (url, title, scan_date, status_code)
                            VALUES (?, ?, ?, ?)
                        ''', (
                            url,
                            row.get('Título', ''),
                            row.get('Fecha Escaneo', format_chile_time(fmt='%Y-%m-%d %H:%M:%S')),
                            int(row.get('Estado', 200)) if row.get('Estado', '').isdigit() else 200
                        ))
                        imported_scans += 1

                except Exception as e:
                    errors.append(f'Error procesando fila CSV: {str(e)}')

        conn.commit()
        conn.close()

        # Prepare success message
        if imported_scans > 0:
            flash(f'{imported_scans} escaneos importados exitosamente', 'success')
        if updated_scans > 0:
            flash(f'{updated_scans} escaneos actualizados', 'info')

        if errors:
            error_msg = f'{len(errors)} errores encontrados: ' + '; '.join(errors[:3])
            if len(errors) > 3:
                error_msg += f' y {len(errors) - 3} más...'
            flash(error_msg, 'warning')

        if imported_scans == 0 and updated_scans == 0 and not errors:
            flash('No se encontraron datos válidos para importar', 'warning')

    except json.JSONDecodeError as e:
        flash(f'Error al parsear archivo JSON: {str(e)}', 'error')
    except Exception as e:
        flash(f'Error al importar estadísticas: {str(e)}', 'error')

    return redirect(url_for('statistics'))

if __name__ == '__main__':
    # Initialize database tables on startup
    init_database()
    create_default_admin()

    # Configure for development vs production
    debug_mode = os.environ.get('FLASK_ENV') != 'production'
    host = os.environ.get('FLASK_HOST', '0.0.0.0')
    port = int(os.environ.get('FLASK_PORT', '5000'))

    if debug_mode:
        print("⚠️  Running in DEBUG mode - NOT suitable for production!")
        print("   Set FLASK_ENV=production to disable debug mode")
    else:
        print("🔒 Running in PRODUCTION mode")

    app.run(debug=debug_mode, host=host, port=port)

@app.after_request
def add_security_headers(response):
    # Comprehensive security headers
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=(), payment=(), usb=(), magnetometer=(), gyroscope=(), speaker=()'

    # Strict CSP policy (may need adjustment for specific CDNs)
    csp_policy = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' cdn.jsdelivr.net stackpath.bootstrapcdn.com; "
        "style-src 'self' 'unsafe-inline' cdn.jsdelivr.net stackpath.bootstrapcdn.com fonts.googleapis.com; "
        "font-src 'self' fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self'; "
        "object-src 'none'; "
        "frame-ancestors 'none'; "
        "base-uri 'self';"
    )
    response.headers['Content-Security-Policy'] = csp_policy

    # Add HSTS in production only
    if not app.debug:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains; preload'

    return response
